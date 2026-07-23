from __future__ import annotations

import argparse
import csv
import json
import math
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.patches import Rectangle
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mlbotnav.visual_entry_low_anchor_suggester import _bar_width_days, _draw_candles, _load_ohlcv, _source_csv, _style_axis
from mlbotnav.visual_entry_stas2_market_phase_audit import (
    PHASES,
    SESSION_TIME_BUCKETS,
    _day_type,
    _discover_stas1_runs,
    _effective_session,
    _iter_days,
    _phase_from_pct,
    _phase_rank,
    _read_csv_rows,
    _rel,
    _safe_mean,
    _session_for_timestamp,
    _write_csv,
    _write_json,
)


STATUS = "STAS2_MARKET_PHASE_REVIEW_CONTINUOUS_WAVE_READY_NO_ML_NO_OPTUNA_REVIEW_ONLY"
PRE_ENTRY_WINDOWS_MIN = [5, 15, 30, 60]


LONG_WAVE_PHASES = [
    ("Нет LONG-волны", 0.10),
    ("Микроход", 0.20),
    ("Малая LONG-волна", 0.35),
    ("Рабочая LONG-волна", 0.60),
    ("Сильная LONG-волна", 1.00),
    ("Очень сильная LONG-волна", math.inf),
]


SHORT_WAVE_PHASES = [
    ("Нет SHORT-волны", 0.10),
    ("Микроход", 0.20),
    ("Малая SHORT-волна", 0.35),
    ("Рабочая SHORT-волна", 0.60),
    ("Сильная SHORT-волна", 1.00),
    ("Очень сильная SHORT-волна", math.inf),
]


MACRO_WAVE_REVERSAL_PCT = 1.0
MACRO_WAVE_CONTINUOUS_SCHEMA_VERSION = 1
OVERVIEW_STRIP_HEIGHT_SCALE = 0.50
OVERVIEW_HEIGHT_RATIOS = [4.9, 0.19, 0.23, 0.23, 0.26, 1.15]
OVERVIEW_COMBO_HEIGHT_RATIOS = [4.9, 0.16, 0.19, 0.19, 0.21, 1.35, 1.05]
DAY_TIME_TICK_HOURS = tuple(range(0, 25, 2))


SESSION_COLORS = {
    "ASIA_PACIFIC": "#1e88e5",
    "PRE_LONDON": "#78909c",
    "LONDON_ONLY": "#43a047",
    "LONDON_NY_OVERLAP": "#f9a825",
    "NY_ONLY": "#8e24aa",
    "POST_NY": "#546e7a",
}


def _utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _run_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _fmt_ts(ts: pd.Timestamp) -> str:
    utc = ts.tz_convert("UTC") if ts.tzinfo is not None else ts.tz_localize("UTC")
    return utc.strftime("%Y-%m-%dT%H:%M:%SZ")


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        out = float(value)
        if math.isnan(out) or math.isinf(out):
            return default
        return out
    except Exception:
        return default


def _fmt_pct(value: Any) -> str:
    if value is None or value == "":
        return ""
    return f"{_safe_float(value):.2f}"


def _is_good(row: dict[str, Any]) -> bool:
    return not str(row.get("review_label") or "").startswith("BAD")


def _label_color(row: dict[str, Any]) -> str:
    return "#00e676" if _is_good(row) else "#ff5252"


def _setup_color(row: dict[str, Any]) -> str:
    rank = int(_safe_float(row.get("entry_setup_quality_rank"), -1.0))
    if rank <= 0:
        return "#ff5252"
    if rank == 1:
        return "#ffb300"
    if rank == 2:
        return "#64dd17"
    return "#00e676"


def _long_wave_from_pct(value: float) -> str:
    for name, upper in LONG_WAVE_PHASES:
        if value < upper:
            return name
    return LONG_WAVE_PHASES[-1][0]


def _long_wave_rank(name: str) -> int:
    for idx, (phase, _upper) in enumerate(LONG_WAVE_PHASES):
        if phase == name:
            return idx
    return -1


def _short_wave_from_pct(value: float) -> str:
    for name, upper in SHORT_WAVE_PHASES:
        if value < upper:
            return name
    return SHORT_WAVE_PHASES[-1][0]


def _short_wave_rank(name: str) -> int:
    for idx, (phase, _upper) in enumerate(SHORT_WAVE_PHASES):
        if phase == name:
            return idx
    return -1


def _hour_direction_bias(long_pct: float, short_pct: float) -> str:
    if long_pct < 0.35 and short_pct < 0.35:
        return "FLAT"
    if long_pct >= 0.60 and short_pct >= 0.60 and abs(long_pct - short_pct) < 0.25:
        return "TWO_WAY_VOLATILE"
    if long_pct >= short_pct:
        return "LONG_DOMINANT" if long_pct >= 0.35 else "WEAK_LONG"
    return "SHORT_DOMINANT" if short_pct >= 0.35 else "WEAK_SHORT"


def _parse_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value if str(item)]
    if value is None or value == "":
        return []
    text = str(value).strip()
    if not text:
        return []
    try:
        parsed = json.loads(text)
    except Exception:
        parsed = None
    if isinstance(parsed, list):
        return [str(item) for item in parsed if str(item)]
    return [part.strip() for part in text.split(",") if part.strip()]


def _stas1_signal_fields(row: dict[str, Any]) -> dict[str, Any]:
    risk_flags = _parse_list(row.get("risk_flags"))
    out: dict[str, Any] = {
        "stas1_risk_flags": risk_flags,
        "stas1_anchor_age_bars": int(_safe_float(row.get("anchor_age_bars"), -1.0)),
        "stas1_confirmation_time_utc": row.get("confirmation_time_utc") or "",
    }
    for key in [
        "feature_range_pos_10",
        "feature_range_pos_20",
        "feature_range_pos_60",
        "feature_range_pos_120",
        "feature_ret_15m_pct",
        "feature_ret_30m_pct",
        "feature_ret_60m_pct",
        "feature_volume_ratio20",
        "feature_lower_wick_to_body",
        "feature_reclaim_from_anchor_low_bps",
        "feature_range_compression20",
        "feature_rsi14",
        "feature_macd_hist",
    ]:
        out[f"stas1_{key}"] = _safe_float(row.get(key), "")
    return out


def _entry_setup_quality(row: dict[str, Any]) -> dict[str, Any]:
    flags = set(_parse_list(row.get("stas1_risk_flags")))
    range_pos10 = _safe_float(row.get("stas1_feature_range_pos_10"), 1.0)
    range_pos20 = _safe_float(row.get("stas1_feature_range_pos_20"), 1.0)
    volume_ratio = _safe_float(row.get("stas1_feature_volume_ratio20"), 0.0)
    lower_wick = _safe_float(row.get("stas1_feature_lower_wick_to_body"), 0.0)
    reclaim_bps = _safe_float(row.get("stas1_feature_reclaim_from_anchor_low_bps"), 0.0)
    ret15 = _safe_float(row.get("stas1_feature_ret_15m_pct"), 0.0)
    ret60 = _safe_float(row.get("stas1_feature_ret_60m_pct"), 0.0)
    pre15_bg_rank = int(_safe_float(row.get("pre_15m_background_phase_rank"), -1.0))
    pre15_pullback = _safe_float(row.get("pre_15m_long_wave_pullback_from_high_pct"), 0.0)
    pre30_wave = _safe_float(row.get("pre_30m_long_wave_up_from_low_pct"), 0.0)
    pre30_pullback = _safe_float(row.get("pre_30m_long_wave_pullback_from_high_pct"), 0.0)

    after_impulse = (
        "wide_signal_range" in flags
        or (pre15_bg_rank >= _phase_rank("Большая") and pre15_pullback >= 0.25)
        or (pre30_wave >= 0.70 and pre30_pullback >= 0.45)
        or (ret15 <= -0.20 and ret60 >= 0.25 and pre30_wave >= 0.50)
    )
    no_clear_low = "anchor_without_clear_wick" in flags or lower_wick < 0.50
    weak_confirmation = "low_volume_confirmation" in flags or volume_ratio < 0.55
    not_near_local_low = range_pos10 > 0.35 and range_pos20 > 0.35

    if after_impulse and no_clear_low:
        code = "NOISE_AFTER_SPIKE_NO_LOW"
        rank = 0
        label = "NOISE: после выноса"
        reason = "До входа был резкий вынос/широкая свеча, а сам low не подтвержден фитилем."
    elif weak_confirmation and no_clear_low:
        code = "NOISE_WEAK_LOW_CONFIRMATION"
        rank = 0
        label = "NOISE: low слабый"
        reason = "Low без нормального фитиля и без объема; кандидат нужен для ручного удаления/negative review."
    elif after_impulse:
        code = "CAUTION_AFTER_SPIKE"
        rank = 1
        label = "WARN: после выноса"
        reason = "Это может быть откат после сильного импульса, а не чистый локальный low."
    elif no_clear_low:
        code = "CAUTION_NO_CLEAR_LOW"
        rank = 1
        label = "WARN: low без фитиля"
        reason = "Anchor-low есть, но форма свечи слабо подтверждает отбой."
    elif not_near_local_low:
        code = "CAUTION_NOT_LOCAL_LOW"
        rank = 1
        label = "WARN: не у low"
        reason = "Цена не достаточно близко к локальному low внутри 10/20m диапазона."
    elif range_pos10 <= 0.20 and range_pos20 <= 0.25 and volume_ratio >= 0.70 and lower_wick >= 0.50:
        code = "CLEAN_LOCAL_LOW"
        rank = 3
        label = "CLEAN: local low"
        reason = "Вход расположен у локального low, есть фитиль/подтверждение и нет pre-entry spike-risk."
    elif range_pos10 <= 0.35 or reclaim_bps >= 5.0:
        code = "WORKING_LOCAL_LOW"
        rank = 2
        label = "OK: local low"
        reason = "Рабочий low-кандидат, но без статуса clean."
    else:
        code = "REVIEW_ONLY"
        rank = 1
        label = "WARN: review"
        reason = "Смешанный контекст; смотреть глазами перед переносом в следующий этап."

    return {
        "entry_setup_quality_code": code,
        "entry_setup_quality_rank": rank,
        "entry_setup_quality_label": label,
        "entry_setup_quality_reason": reason,
        "entry_setup_decision_input": (
            "pre-entry only: Stas1 risk_flags/features at signal close + Stas2 pre_15m/pre_30m context; "
            "no TP/exit/MFE/MAE/future candles"
        ),
    }


def _load_days(root: Path, days: list[str], symbol: str, timeframe: str) -> tuple[pd.DataFrame, dict[str, pd.DataFrame], list[str]]:
    frames: list[pd.DataFrame] = []
    frames_by_day: dict[str, pd.DataFrame] = {}
    missing: list[str] = []
    for day in days:
        source = _source_csv(root, day, timeframe, symbol)
        if not source.exists():
            missing.append(_rel(root, source))
            continue
        df = _load_ohlcv(source)
        df["day_utc"] = day
        df["source_csv"] = _rel(root, source)
        df["candle_range_pct"] = (df["high"] - df["low"]) / df["open"].replace(0, np.nan) * 100.0
        df["candle_body_pct"] = (df["close"] - df["open"]).abs() / df["open"].replace(0, np.nan) * 100.0
        df["close_step_abs_pct"] = df["close"].pct_change().abs().fillna(0.0) * 100.0
        frames.append(df)
        frames_by_day[day] = df
    if not frames:
        raise FileNotFoundError("no OHLCV sources found for requested period")
    all_df = pd.concat(frames, ignore_index=True).sort_values("open_time_utc").reset_index(drop=True)
    return all_df, frames_by_day, missing


def _phase_context_for_window(df: pd.DataFrame, entry_ts: pd.Timestamp, minutes: int) -> dict[str, Any]:
    start = entry_ts - pd.Timedelta(minutes=minutes)
    prior = df[(df["open_time_utc"] >= start) & (df["open_time_utc"] < entry_ts)].sort_values("open_time_utc")
    if prior.empty:
        return {
            f"pre_{minutes}m_rows": 0,
            f"pre_{minutes}m_first_open_time_utc": "",
            f"pre_{minutes}m_last_open_time_utc": "",
            f"pre_{minutes}m_range_pct": "",
            f"pre_{minutes}m_close_move_pct": "",
            f"pre_{minutes}m_path_pct": "",
            f"pre_{minutes}m_median_candle_pct": "",
            f"pre_{minutes}m_phase_metric_pct": "",
            f"pre_{minutes}m_phase": "NO_DATA",
            f"pre_{minutes}m_phase_rank": -1,
            f"pre_{minutes}m_background_phase_metric_pct": "",
            f"pre_{minutes}m_background_phase": "NO_DATA",
            f"pre_{minutes}m_background_phase_rank": -1,
            f"pre_{minutes}m_window_low_time_utc": "",
            f"pre_{minutes}m_window_low_price": "",
            f"pre_{minutes}m_post_low_high_time_utc": "",
            f"pre_{minutes}m_post_low_high_price": "",
            f"pre_{minutes}m_long_wave_up_from_low_pct": "",
            f"pre_{minutes}m_long_wave_up_from_low_to_last_close_pct": "",
            f"pre_{minutes}m_long_wave_up_from_low_to_post_low_high_pct": "",
            f"pre_{minutes}m_up_from_low_pct": "",
            f"pre_{minutes}m_long_wave_pullback_from_post_low_high_pct": "",
            f"pre_{minutes}m_long_wave_pullback_from_high_pct": "",
            f"pre_{minutes}m_pullback_from_high_pct": "",
            f"pre_{minutes}m_long_wave_phase": "NO_DATA",
            f"pre_{minutes}m_long_wave_rank": -1,
            f"pre_{minutes}m_long_wave_input": "NO_DATA",
        }
    open_price = float(prior.iloc[0]["open"])
    close_price = float(prior.iloc[-1]["close"])
    high = float(prior["high"].max())
    low = float(prior["low"].min())
    range_pct = (high - low) / max(open_price, 1e-12) * 100.0
    close_move_pct = (close_price - open_price) / max(open_price, 1e-12) * 100.0
    path_pct = float(prior["close_step_abs_pct"].sum())
    median_candle_pct = _safe_mean(prior["candle_range_pct"])
    phase_metric_pct = max(range_pct, abs(close_move_pct) * 1.25, median_candle_pct * 8.0)
    phase = _phase_from_pct(phase_metric_pct)
    low_idx = prior["low"].astype(float).idxmin()
    low_row = prior.loc[low_idx]
    low_time = pd.Timestamp(low_row["open_time_utc"])
    low_price = float(low_row["low"])
    after_low = prior.loc[low_idx:].copy()
    high_after_low_idx = after_low["high"].astype(float).idxmax()
    high_after_low_row = after_low.loc[high_after_low_idx]
    high_after_low_time = pd.Timestamp(high_after_low_row["open_time_utc"])
    high_after_low = float(high_after_low_row["high"])
    long_wave_up_to_last_close_pct = max(0.0, (close_price - low_price) / max(low_price, 1e-12) * 100.0)
    long_wave_up_to_post_low_high_pct = max(0.0, (high_after_low - low_price) / max(low_price, 1e-12) * 100.0)
    long_wave_pullback_pct = max(0.0, (high_after_low - close_price) / max(high_after_low, 1e-12) * 100.0)
    long_wave_phase = _long_wave_from_pct(long_wave_up_to_post_low_high_pct)
    return {
        f"pre_{minutes}m_rows": int(len(prior)),
        f"pre_{minutes}m_first_open_time_utc": _fmt_ts(pd.Timestamp(prior.iloc[0]["open_time_utc"])),
        f"pre_{minutes}m_last_open_time_utc": _fmt_ts(pd.Timestamp(prior.iloc[-1]["open_time_utc"])),
        f"pre_{minutes}m_range_pct": round(range_pct, 6),
        f"pre_{minutes}m_close_move_pct": round(close_move_pct, 6),
        f"pre_{minutes}m_path_pct": round(path_pct, 6),
        f"pre_{minutes}m_median_candle_pct": round(median_candle_pct, 6),
        f"pre_{minutes}m_phase_metric_pct": round(phase_metric_pct, 6),
        f"pre_{minutes}m_phase": phase,
        f"pre_{minutes}m_phase_rank": _phase_rank(phase),
        f"pre_{minutes}m_background_phase_metric_pct": round(phase_metric_pct, 6),
        f"pre_{minutes}m_background_phase": phase,
        f"pre_{minutes}m_background_phase_rank": _phase_rank(phase),
        f"pre_{minutes}m_window_low_time_utc": _fmt_ts(low_time),
        f"pre_{minutes}m_window_low_price": round(low_price, 8),
        f"pre_{minutes}m_post_low_high_time_utc": _fmt_ts(high_after_low_time),
        f"pre_{minutes}m_post_low_high_price": round(high_after_low, 8),
        f"pre_{minutes}m_long_wave_up_from_low_pct": round(long_wave_up_to_post_low_high_pct, 6),
        f"pre_{minutes}m_long_wave_up_from_low_to_last_close_pct": round(long_wave_up_to_last_close_pct, 6),
        f"pre_{minutes}m_long_wave_up_from_low_to_post_low_high_pct": round(long_wave_up_to_post_low_high_pct, 6),
        f"pre_{minutes}m_up_from_low_pct": round(long_wave_up_to_post_low_high_pct, 6),
        f"pre_{minutes}m_long_wave_pullback_from_post_low_high_pct": round(long_wave_pullback_pct, 6),
        f"pre_{minutes}m_long_wave_pullback_from_high_pct": round(long_wave_pullback_pct, 6),
        f"pre_{minutes}m_pullback_from_high_pct": round(long_wave_pullback_pct, 6),
        f"pre_{minutes}m_long_wave_phase": long_wave_phase,
        f"pre_{minutes}m_long_wave_rank": _long_wave_rank(long_wave_phase),
        f"pre_{minutes}m_long_wave_input": "window low -> subsequent high/last close from bars where open_time_utc < entry_time_utc",
    }


def _range_context(df: pd.DataFrame, start_ts: pd.Timestamp, entry_ts: pd.Timestamp, prefix: str) -> dict[str, Any]:
    prior = df[(df["open_time_utc"] >= start_ts) & (df["open_time_utc"] < entry_ts)].sort_values("open_time_utc")
    if prior.empty:
        return {
            f"{prefix}_rows": 0,
            f"{prefix}_first_open_time_utc": "",
            f"{prefix}_last_open_time_utc": "",
            f"{prefix}_range_pct": "",
            f"{prefix}_close_move_pct": "",
        }
    open_price = float(prior.iloc[0]["open"])
    close_price = float(prior.iloc[-1]["close"])
    high = float(prior["high"].max())
    low = float(prior["low"].min())
    return {
        f"{prefix}_rows": int(len(prior)),
        f"{prefix}_first_open_time_utc": _fmt_ts(pd.Timestamp(prior.iloc[0]["open_time_utc"])),
        f"{prefix}_last_open_time_utc": _fmt_ts(pd.Timestamp(prior.iloc[-1]["open_time_utc"])),
        f"{prefix}_range_pct": round((high - low) / max(open_price, 1e-12) * 100.0, 6),
        f"{prefix}_close_move_pct": round((close_price - open_price) / max(open_price, 1e-12) * 100.0, 6),
    }


def _session_start_for_entry(entry_ts: pd.Timestamp, session_code: str) -> pd.Timestamp:
    day_start = entry_ts.floor("D")
    for code, _label, start_minute, _end_minute in SESSION_TIME_BUCKETS:
        if code == session_code:
            return day_start + pd.Timedelta(minutes=start_minute)
    return day_start


def _load_entry_context_rows(
    *,
    root: Path,
    df: pd.DataFrame,
    start_day: str,
    end_day: str,
    stas1_run_dirs: list[str],
) -> tuple[list[dict[str, Any]], list[str]]:
    start = pd.Timestamp(start_day, tz="UTC")
    end = pd.Timestamp(end_day, tz="UTC") + pd.Timedelta(days=1)
    runs = _discover_stas1_runs(root, stas1_run_dirs)
    rows: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for run in runs:
        for row in _read_csv_rows(run.csv_path):
            raw_entry = row.get("entry_time_utc") or ""
            if not raw_entry:
                continue
            entry_ts = pd.Timestamp(raw_entry)
            entry_ts = entry_ts.tz_convert("UTC") if entry_ts.tzinfo is not None else entry_ts.tz_localize("UTC")
            if entry_ts < start or entry_ts >= end:
                continue
            key = (str(row.get("record_id") or ""), _fmt_ts(entry_ts))
            if key in seen:
                continue
            seen.add(key)

            session_code, session_label = _session_for_timestamp(entry_ts)
            day_type = _day_type(entry_ts)
            effective_code, effective_label, real_tradfi_open = _effective_session(session_code, session_label, day_type)
            day_start = entry_ts.floor("D")
            session_start = _session_start_for_entry(entry_ts, session_code)

            out = {
                "source_run": _rel(root, run.run_dir),
                "record_id": row.get("record_id") or "",
                "candidate_id": row.get("candidate_id") or "",
                "day_utc": entry_ts.strftime("%Y-%m-%d"),
                "anchor_time_utc": row.get("anchor_time_utc") or "",
                "entry_time_utc": _fmt_ts(entry_ts),
                "entry_open_price": _safe_float(row.get("entry_open_price")),
                "entry_price_5bps": _safe_float(row.get("entry_price_5bps")),
                "anchor_low_price": _safe_float(row.get("anchor_low_price")),
                "review_label": row.get("review_label") or "",
                "outcome_status": row.get("outcome_status") or "",
                "is_good_stas1_review": _is_good(row),
                "suggested_type": row.get("suggested_type") or "",
                "score": _safe_float(row.get("score")),
                "day_type": day_type,
                "is_weekend": day_type == "weekend",
                "session_time_bucket_code": session_code,
                "session_time_bucket_label": session_label,
                "effective_session_code": effective_code,
                "effective_session_label": effective_label,
                "real_tradfi_session_open": real_tradfi_open,
            }
            out.update(_stas1_signal_fields(row))
            for minutes in PRE_ENTRY_WINDOWS_MIN:
                out.update(_phase_context_for_window(df, entry_ts, minutes))
            out.update(_range_context(df, session_start, entry_ts, "session_so_far"))
            out.update(_range_context(df, day_start, entry_ts, "day_so_far"))
            out.update(_entry_setup_quality(out))
            latest_context_ts = ""
            for context_key in [
                "pre_5m_last_open_time_utc",
                "pre_15m_last_open_time_utc",
                "pre_30m_last_open_time_utc",
                "pre_60m_last_open_time_utc",
                "session_so_far_last_open_time_utc",
                "day_so_far_last_open_time_utc",
            ]:
                value = str(out.get(context_key) or "")
                if value > latest_context_ts:
                    latest_context_ts = value
            out["context_max_open_time_utc"] = latest_context_ts
            out["context_before_entry_check"] = bool(latest_context_ts and latest_context_ts < _fmt_ts(entry_ts))
            out["phase_decision_input"] = "background phase: pre_5m/pre_15m/pre_30m/pre_60m only; all use open_time_utc < entry_time_utc"
            out["long_wave_decision_input"] = "LONG wave: window low -> subsequent high/last_close from pre-entry windows only; all use open_time_utc < entry_time_utc"
            out["stas2_boundary"] = "pre_entry_context_only_no_tp_no_exit_no_future_bars"
            rows.append(out)
    return sorted(rows, key=lambda item: (str(item["entry_time_utc"]), str(item["record_id"]))), [_rel(root, run.run_dir) for run in runs]


def _hour_phase_rows(df: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for hour_ts, group in df.groupby(df["open_time_utc"].dt.floor("h"), sort=True):
        group = group.sort_values("open_time_utc")
        open_price = float(group.iloc[0]["open"])
        close_price = float(group.iloc[-1]["close"])
        high = float(group["high"].max())
        low = float(group["low"].min())
        range_pct = (high - low) / max(open_price, 1e-12) * 100.0
        move_pct = (close_price - open_price) / max(open_price, 1e-12) * 100.0
        path_pct = float(group["close_step_abs_pct"].sum())
        median_candle_pct = _safe_mean(group["candle_range_pct"])
        phase_metric_pct = max(range_pct, abs(move_pct) * 1.25, median_candle_pct * 8.0)
        phase = _phase_from_pct(phase_metric_pct)
        low_idx = group["low"].astype(float).idxmin()
        low_row = group.loc[low_idx]
        low_time = pd.Timestamp(low_row["open_time_utc"])
        low_price = float(low_row["low"])
        after_low = group.loc[low_idx:].copy()
        high_after_low_idx = after_low["high"].astype(float).idxmax()
        high_after_low_row = after_low.loc[high_after_low_idx]
        high_after_low_time = pd.Timestamp(high_after_low_row["open_time_utc"])
        high_after_low = float(high_after_low_row["high"])
        hour_long_wave_up_from_low_pct = max(0.0, (high_after_low - low_price) / max(low_price, 1e-12) * 100.0)
        hour_long_wave_close_from_low_pct = max(0.0, (close_price - low_price) / max(low_price, 1e-12) * 100.0)
        hour_long_wave_pullback_pct = max(0.0, (high_after_low - close_price) / max(high_after_low, 1e-12) * 100.0)
        hour_long_wave_phase = _long_wave_from_pct(hour_long_wave_up_from_low_pct)
        high_idx = group["high"].astype(float).idxmax()
        high_row = group.loc[high_idx]
        high_time = pd.Timestamp(high_row["open_time_utc"])
        high_price = float(high_row["high"])
        after_high = group.loc[high_idx:].copy()
        low_after_high_idx = after_high["low"].astype(float).idxmin()
        low_after_high_row = after_high.loc[low_after_high_idx]
        low_after_high_time = pd.Timestamp(low_after_high_row["open_time_utc"])
        low_after_high = float(low_after_high_row["low"])
        hour_short_wave_down_from_high_pct = max(0.0, (high_price - low_after_high) / max(high_price, 1e-12) * 100.0)
        hour_short_wave_close_from_high_pct = max(0.0, (high_price - close_price) / max(high_price, 1e-12) * 100.0)
        hour_short_wave_bounce_pct = max(0.0, (close_price - low_after_high) / max(low_after_high, 1e-12) * 100.0)
        hour_short_wave_phase = _short_wave_from_pct(hour_short_wave_down_from_high_pct)
        session_code, session_label = _session_for_timestamp(hour_ts + pd.Timedelta(minutes=30))
        day_type = _day_type(hour_ts)
        effective_code, effective_label, real_tradfi_open = _effective_session(session_code, session_label, day_type)
        rows.append(
            {
                "day_utc": hour_ts.strftime("%Y-%m-%d"),
                "hour_start_utc": _fmt_ts(hour_ts),
                "hour_utc": int(hour_ts.hour),
                "day_type": day_type,
                "session_time_bucket_code": session_code,
                "session_time_bucket_label": session_label,
                "effective_session_code": effective_code,
                "effective_session_label": effective_label,
                "real_tradfi_session_open": real_tradfi_open,
                "hour_range_pct": round(range_pct, 6),
                "hour_close_move_pct": round(move_pct, 6),
                "hour_path_pct": round(path_pct, 6),
                "phase_metric_pct": round(phase_metric_pct, 6),
                "market_phase": phase,
                "market_phase_rank": _phase_rank(phase),
                "hour_background_phase_metric_pct": round(phase_metric_pct, 6),
                "hour_background_phase": phase,
                "hour_background_phase_rank": _phase_rank(phase),
                "hour_long_wave_low_time_utc": _fmt_ts(low_time),
                "hour_long_wave_low_price": round(low_price, 8),
                "hour_long_wave_post_low_high_time_utc": _fmt_ts(high_after_low_time),
                "hour_long_wave_post_low_high_price": round(high_after_low, 8),
                "hour_long_wave_up_from_low_pct": round(hour_long_wave_up_from_low_pct, 6),
                "hour_long_wave_close_from_low_pct": round(hour_long_wave_close_from_low_pct, 6),
                "hour_long_wave_pullback_from_post_low_high_pct": round(hour_long_wave_pullback_pct, 6),
                "hour_long_wave_phase": hour_long_wave_phase,
                "hour_long_wave_rank": _long_wave_rank(hour_long_wave_phase),
                "hour_long_wave_boundary": "closed_hour_overview_only_not_entry_feature",
                "hour_short_wave_high_time_utc": _fmt_ts(high_time),
                "hour_short_wave_high_price": round(high_price, 8),
                "hour_short_wave_post_high_low_time_utc": _fmt_ts(low_after_high_time),
                "hour_short_wave_post_high_low_price": round(low_after_high, 8),
                "hour_short_wave_down_from_high_pct": round(hour_short_wave_down_from_high_pct, 6),
                "hour_short_wave_close_from_high_pct": round(hour_short_wave_close_from_high_pct, 6),
                "hour_short_wave_bounce_from_post_high_low_pct": round(hour_short_wave_bounce_pct, 6),
                "hour_short_wave_phase": hour_short_wave_phase,
                "hour_short_wave_rank": _short_wave_rank(hour_short_wave_phase),
                "hour_short_wave_boundary": "closed_hour_overview_only_not_entry_feature",
                "hour_direction_bias": _hour_direction_bias(hour_long_wave_up_from_low_pct, hour_short_wave_down_from_high_pct),
                "audit_boundary": "closed_hour_overview_only_not_entry_feature",
            }
        )
    return rows


def _pct_up(low_price: float, high_price: float) -> float:
    return max(0.0, (high_price - low_price) / max(low_price, 1e-12) * 100.0)


def _pct_down(high_price: float, low_price: float) -> float:
    return max(0.0, (high_price - low_price) / max(high_price, 1e-12) * 100.0)


def _bar_step(df: pd.DataFrame) -> pd.Timedelta:
    if len(df) > 1:
        diffs = df.sort_values("open_time_utc")["open_time_utc"].diff().dropna()
        step = diffs.median() if not diffs.empty else pd.Timedelta(minutes=1)
        if pd.notna(step) and pd.Timedelta(0) < step <= pd.Timedelta(hours=1):
            return step
    return pd.Timedelta(minutes=1)


def _segment_stats(df: pd.DataFrame, start_time: pd.Timestamp, end_time: pd.Timestamp, *, include_end: bool = False) -> dict[str, float]:
    ordered = df.sort_values("open_time_utc")
    if include_end:
        segment = ordered[(ordered["open_time_utc"] >= start_time) & (ordered["open_time_utc"] <= end_time)]
    else:
        segment = ordered[(ordered["open_time_utc"] >= start_time) & (ordered["open_time_utc"] < end_time)]
    if segment.empty:
        return {"start_price": 0.0, "end_price": 0.0, "high": 0.0, "low": 0.0, "range_pct": 0.0, "close_move_pct": 0.0}
    start_price = float(segment.iloc[0]["open"])
    end_price = float(segment.iloc[-1]["close"])
    high = float(segment["high"].max())
    low = float(segment["low"].min())
    range_pct = (high - low) / max(start_price, 1e-12) * 100.0
    close_move_pct = (end_price - start_price) / max(start_price, 1e-12) * 100.0
    return {
        "start_price": start_price,
        "end_price": end_price,
        "high": high,
        "low": low,
        "range_pct": range_pct,
        "close_move_pct": close_move_pct,
    }


def _crossed_hour_count(start_time: pd.Timestamp, end_time: pd.Timestamp) -> int:
    if end_time <= start_time:
        return 0
    safe_end = end_time - pd.Timedelta(microseconds=1)
    return int((safe_end.floor("h") - start_time.floor("h")).total_seconds() // 3600) + 1


def _crossed_day_count(start_time: pd.Timestamp, end_time: pd.Timestamp) -> int:
    if end_time <= start_time:
        return 0
    safe_end = end_time - pd.Timedelta(microseconds=1)
    return int((safe_end.floor("D") - start_time.floor("D")).total_seconds() // 86400) + 1


def _continuous_wave_rows(df: pd.DataFrame, reversal_pct: float = MACRO_WAVE_REVERSAL_PCT) -> list[dict[str, Any]]:
    group = df.sort_values("open_time_utc").reset_index(drop=True)
    if len(group) < 2:
        return []
    rows: list[dict[str, Any]] = []
    first = group.iloc[0]
    data_start = pd.Timestamp(first["open_time_utc"])
    data_end = pd.Timestamp(group.iloc[-1]["open_time_utc"]) + _bar_step(group)
    candidate_low_price = float(first["low"])
    candidate_low_time = data_start
    candidate_high_price = float(first["high"])
    candidate_high_time = data_start
    direction = ""
    wave_start_time = candidate_low_time
    wave_start_price = candidate_low_price
    extreme_time = candidate_high_time
    extreme_price = candidate_high_price
    wave_no = 1

    def add_wave(*, direction: str, start_time: pd.Timestamp, start_price: float, end_time: pd.Timestamp, end_price: float, status: str) -> None:
        nonlocal wave_no
        if end_time <= start_time:
            return
        move_pct = _pct_up(start_price, end_price) if direction == "LONG" else _pct_down(start_price, end_price)
        if move_pct <= 0.0:
            return
        duration_min = (end_time - start_time).total_seconds() / 60.0
        rows.append(
            {
                "continuous_wave_schema_version": MACRO_WAVE_CONTINUOUS_SCHEMA_VERSION,
                "continuous_wave_id": f"CW_{wave_no:06d}",
                "continuous_wave_no": wave_no,
                "continuous_wave_segment_kind": "WAVE",
                "continuous_wave_direction": direction,
                "continuous_wave_status": status,
                "continuous_wave_start_time_utc": _fmt_ts(start_time),
                "continuous_wave_start_price": round(start_price, 8),
                "continuous_wave_end_time_utc": _fmt_ts(end_time),
                "continuous_wave_end_price": round(end_price, 8),
                "continuous_wave_move_pct": round(move_pct, 6),
                "continuous_wave_move_basis": "directional_pivot_move_pct",
                "continuous_wave_range_pct": round(move_pct, 6),
                "continuous_wave_close_move_pct": "",
                "continuous_wave_duration_min": round(duration_min, 3),
                "continuous_wave_start_day_utc": start_time.strftime("%Y-%m-%d"),
                "continuous_wave_end_day_utc": end_time.strftime("%Y-%m-%d"),
                "continuous_wave_crossed_day_count": _crossed_day_count(start_time, end_time),
                "continuous_wave_crossed_hour_count": _crossed_hour_count(start_time, end_time),
                "continuous_wave_reversal_threshold_pct": reversal_pct,
                "continuous_wave_is_last_leg": status == "ACTIVE",
                "continuous_wave_boundary": "continuous_macro_wave_hindsight_review_not_entry_feature",
            }
        )
        wave_no += 1

    for item in group.itertuples(index=False):
        ts = pd.Timestamp(item.open_time_utc)
        high = float(item.high)
        low = float(item.low)

        if not direction:
            if low < candidate_low_price:
                candidate_low_price = low
                candidate_low_time = ts
            if high > candidate_high_price:
                candidate_high_price = high
                candidate_high_time = ts
            if candidate_low_time <= candidate_high_time and _pct_up(candidate_low_price, candidate_high_price) >= reversal_pct:
                direction = "LONG"
                wave_start_time = candidate_low_time
                wave_start_price = candidate_low_price
                extreme_time = candidate_high_time
                extreme_price = candidate_high_price
                continue
            if candidate_high_time <= candidate_low_time and _pct_down(candidate_high_price, candidate_low_price) >= reversal_pct:
                direction = "SHORT"
                wave_start_time = candidate_high_time
                wave_start_price = candidate_high_price
                extreme_time = candidate_low_time
                extreme_price = candidate_low_price
                continue
            continue

        if direction == "LONG":
            if high >= extreme_price:
                extreme_price = high
                extreme_time = ts
            if ts > extreme_time and _pct_down(extreme_price, low) >= reversal_pct:
                add_wave(
                    direction="LONG",
                    start_time=wave_start_time,
                    start_price=wave_start_price,
                    end_time=extreme_time,
                    end_price=extreme_price,
                    status="CONFIRMED",
                )
                direction = "SHORT"
                wave_start_time = extreme_time
                wave_start_price = extreme_price
                extreme_time = ts
                extreme_price = low
        else:
            if low <= extreme_price:
                extreme_price = low
                extreme_time = ts
            if ts > extreme_time and _pct_up(extreme_price, high) >= reversal_pct:
                add_wave(
                    direction="SHORT",
                    start_time=wave_start_time,
                    start_price=wave_start_price,
                    end_time=extreme_time,
                    end_price=extreme_price,
                    status="CONFIRMED",
                )
                direction = "LONG"
                wave_start_time = extreme_time
                wave_start_price = extreme_price
                extreme_time = ts
                extreme_price = high

    if direction:
        add_wave(
            direction=direction,
            start_time=wave_start_time,
            start_price=wave_start_price,
            end_time=extreme_time,
            end_price=extreme_price,
            status="ACTIVE",
        )

    wave_rows = sorted(
        rows,
        key=lambda item: (str(item.get("continuous_wave_start_time_utc") or ""), str(item.get("continuous_wave_id") or "")),
    )
    out: list[dict[str, Any]] = []
    gap_no = 1
    cursor = data_start

    def add_gap(start_time: pd.Timestamp, end_time: pd.Timestamp, reason: str) -> None:
        nonlocal gap_no
        if end_time <= start_time:
            return
        duration_min = (end_time - start_time).total_seconds() / 60.0
        if duration_min < 1.0:
            return
        stats = _segment_stats(group, start_time, end_time)
        out.append(
            {
                "continuous_wave_schema_version": MACRO_WAVE_CONTINUOUS_SCHEMA_VERSION,
                "continuous_wave_id": f"CU_{gap_no:06d}",
                "continuous_wave_no": f"GAP{gap_no:02d}",
                "continuous_wave_segment_kind": "GAP",
                "continuous_wave_direction": "GAP",
                "continuous_wave_status": "UNCONFIRMED",
                "continuous_wave_gap_reason": reason,
                "continuous_wave_start_time_utc": _fmt_ts(start_time),
                "continuous_wave_start_price": round(stats["start_price"], 8) if stats["start_price"] else "",
                "continuous_wave_end_time_utc": _fmt_ts(end_time),
                "continuous_wave_end_price": round(stats["end_price"], 8) if stats["end_price"] else "",
                "continuous_wave_move_pct": round(stats["range_pct"], 6),
                "continuous_wave_move_basis": "gap_range_pct",
                "continuous_wave_range_pct": round(stats["range_pct"], 6),
                "continuous_wave_close_move_pct": round(stats["close_move_pct"], 6),
                "continuous_wave_duration_min": round(duration_min, 3),
                "continuous_wave_start_day_utc": start_time.strftime("%Y-%m-%d"),
                "continuous_wave_end_day_utc": (end_time - pd.Timedelta(microseconds=1)).strftime("%Y-%m-%d"),
                "continuous_wave_crossed_day_count": _crossed_day_count(start_time, end_time),
                "continuous_wave_crossed_hour_count": _crossed_hour_count(start_time, end_time),
                "continuous_wave_reversal_threshold_pct": reversal_pct,
                "continuous_wave_is_last_leg": False,
                "continuous_wave_boundary": "unconfirmed_continuous_gap_review_not_entry_feature",
            }
        )
        gap_no += 1

    for row in wave_rows:
        start = pd.Timestamp(row["continuous_wave_start_time_utc"])
        end = pd.Timestamp(row["continuous_wave_end_time_utc"])
        if start > cursor:
            add_gap(cursor, start, "before_next_confirmed_wave")
        out.append(row)
        if end > cursor:
            cursor = end
    if data_end > cursor:
        add_gap(cursor, data_end, "after_last_confirmed_or_active_wave")
    return sorted(out, key=lambda item: (str(item.get("continuous_wave_start_time_utc") or ""), str(item.get("continuous_wave_id") or "")))


def _macro_wave_rows(
    df: pd.DataFrame,
    reversal_pct: float = MACRO_WAVE_REVERSAL_PCT,
    continuous_rows: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    if continuous_rows is None:
        continuous_rows = _continuous_wave_rows(df, reversal_pct=reversal_pct)
    group = df.sort_values("open_time_utc").reset_index(drop=True)
    if group.empty:
        return []
    rows: list[dict[str, Any]] = []
    days = sorted(str(day) for day in group["day_utc"].dropna().unique())
    for continuous in continuous_rows:
        full_start = pd.Timestamp(continuous["continuous_wave_start_time_utc"])
        full_end = pd.Timestamp(continuous["continuous_wave_end_time_utc"])
        if full_end <= full_start:
            continue
        segment_kind = str(continuous.get("continuous_wave_segment_kind") or "WAVE")
        direction = str(continuous.get("continuous_wave_direction") or "")
        full_move_pct = _safe_float(continuous.get("continuous_wave_move_pct"), 0.0)
        full_duration_min = _safe_float(continuous.get("continuous_wave_duration_min"), 0.0)
        for day in days:
            day_start = pd.Timestamp(day, tz="UTC")
            day_end = day_start + pd.Timedelta(days=1)
            visible_start = max(full_start, day_start)
            visible_end = min(full_end, day_end)
            if visible_end <= visible_start:
                continue
            include_end = visible_end == full_end and visible_end < day_end
            stats = _segment_stats(group, visible_start, visible_end, include_end=include_end)
            full_start_price = _safe_float(continuous.get("continuous_wave_start_price"), stats["start_price"])
            full_end_price = _safe_float(continuous.get("continuous_wave_end_price"), stats["end_price"])
            visible_start_price = full_start_price if visible_start == full_start else stats["start_price"]
            if segment_kind == "WAVE" and direction == "LONG":
                visible_end_price = full_end_price if visible_end == full_end else stats["high"]
                visible_move_pct = _pct_up(visible_start_price, visible_end_price) if visible_start_price and visible_end_price else 0.0
            elif segment_kind == "WAVE" and direction == "SHORT":
                visible_end_price = full_end_price if visible_end == full_end else stats["low"]
                visible_move_pct = _pct_down(visible_start_price, visible_end_price) if visible_start_price and visible_end_price else 0.0
            else:
                visible_move_pct = stats["range_pct"]
                visible_end_price = stats["end_price"]
            duration_min = (visible_end - visible_start).total_seconds() / 60.0
            carry_from_prev = full_start < day_start
            carry_to_next = full_end > day_end
            wave_no = continuous.get("continuous_wave_no", "")
            slice_id = f"{continuous.get('continuous_wave_id')}_{day.replace('-', '')}"
            rows.append(
                {
                    "day_utc": day,
                    "macro_wave_id": slice_id,
                    "macro_wave_no": wave_no,
                    "macro_wave_segment_kind": segment_kind,
                    "macro_wave_direction": direction,
                    "macro_wave_status": continuous.get("continuous_wave_status", ""),
                    "macro_wave_gap_reason": continuous.get("continuous_wave_gap_reason", ""),
                    "macro_wave_start_time_utc": _fmt_ts(visible_start),
                    "macro_wave_start_price": round(visible_start_price, 8) if visible_start_price else "",
                    "macro_wave_end_time_utc": _fmt_ts(visible_end),
                    "macro_wave_end_price": round(visible_end_price, 8) if visible_end_price else "",
                    "macro_wave_move_pct": round(visible_move_pct, 6),
                    "macro_wave_visible_move_pct": round(visible_move_pct, 6),
                    "macro_wave_full_move_pct": round(full_move_pct, 6),
                    "macro_wave_move_basis": "visible_directional_slice_pct" if segment_kind == "WAVE" else "gap_range_pct",
                    "macro_wave_range_pct": round(stats["range_pct"], 6),
                    "macro_wave_visible_range_pct": round(stats["range_pct"], 6),
                    "macro_wave_close_move_pct": round(stats["close_move_pct"], 6) if segment_kind != "WAVE" else "",
                    "macro_wave_duration_min": round(duration_min, 3),
                    "macro_wave_full_duration_min": round(full_duration_min, 3),
                    "macro_wave_start_hour_utc": _fmt_ts(visible_start.floor("h")),
                    "macro_wave_end_hour_utc": _fmt_ts((visible_end - pd.Timedelta(microseconds=1)).floor("h")),
                    "macro_wave_crossed_hour_count": _crossed_hour_count(visible_start, visible_end),
                    "macro_wave_crossed_day_count": _crossed_day_count(full_start, full_end),
                    "macro_wave_reversal_threshold_pct": reversal_pct,
                    "macro_wave_is_last_leg": bool(continuous.get("continuous_wave_is_last_leg")),
                    "macro_wave_carry_from_prev_day": carry_from_prev,
                    "macro_wave_carry_to_next_day": carry_to_next,
                    "macro_wave_full_start_time_utc": _fmt_ts(full_start),
                    "macro_wave_full_end_time_utc": _fmt_ts(full_end),
                    "macro_wave_full_start_day_utc": continuous.get("continuous_wave_start_day_utc", ""),
                    "macro_wave_full_end_day_utc": continuous.get("continuous_wave_end_day_utc", ""),
                    "continuous_wave_id": continuous.get("continuous_wave_id", ""),
                    "macro_wave_boundary": "continuous_macro_wave_day_slice_hindsight_review_not_entry_feature"
                    if segment_kind == "WAVE"
                    else "unconfirmed_gap_day_slice_not_macro_wave_not_entry_feature",
                }
            )
    return sorted(rows, key=lambda item: (str(item.get("day_utc") or ""), str(item.get("macro_wave_start_time_utc") or ""), str(item.get("macro_wave_id") or "")))


def _entry_summary(rows: list[dict[str, Any]], keys: list[str], phase_field: str = "pre_60m_phase") -> list[dict[str, Any]]:
    buckets: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        buckets[tuple(row.get(key, "") for key in keys)].append(row)
    out: list[dict[str, Any]] = []
    for key, items in sorted(buckets.items(), key=lambda item: item[0]):
        phase_counts = Counter(str(item.get(phase_field) or "") for item in items)
        good = sum(1 for item in items if bool(item.get("is_good_stas1_review")))
        record = {keys[idx]: key[idx] for idx in range(len(keys))}
        ranks = [_safe_float(item.get(phase_field.replace("_phase", "_phase_rank")), -1.0) for item in items]
        record.update(
            {
                "entries": len(items),
                "stas1_good": good,
                "stas1_bad": len(items) - good,
                "stas1_good_share": round(good / max(len(items), 1), 6),
                "avg_phase_rank": round(sum(rank for rank in ranks if rank >= 0) / max(sum(1 for rank in ranks if rank >= 0), 1), 6),
                "dominant_phase": phase_counts.most_common(1)[0][0] if phase_counts else "",
                "phase_counts": dict(phase_counts),
                "boundary": "Stas1 GOOD/BAD is review outcome context, not ML label",
            }
        )
        out.append(record)
    return out


def _entry_phase_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for minutes in PRE_ENTRY_WINDOWS_MIN:
        key = f"pre_{minutes}m_phase"
        buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            buckets[str(row.get(key) or "")].append(row)
        for phase, items in sorted(buckets.items(), key=lambda item: _phase_rank(item[0])):
            good = sum(1 for item in items if bool(item.get("is_good_stas1_review")))
            out.append(
                {
                    "window": f"pre_{minutes}m",
                    "phase": phase,
                    "phase_rank": _phase_rank(phase),
                    "entries": len(items),
                    "stas1_good": good,
                    "stas1_bad": len(items) - good,
                    "stas1_good_share": round(good / max(len(items), 1), 6),
                    "boundary": "phase is computed before entry; no TP/exit here",
                }
            )
    return out


def _entry_long_wave_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for minutes in PRE_ENTRY_WINDOWS_MIN:
        key = f"pre_{minutes}m_long_wave_phase"
        pct_key = f"pre_{minutes}m_long_wave_up_from_low_pct"
        pullback_key = f"pre_{minutes}m_long_wave_pullback_from_high_pct"
        buckets: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            buckets[str(row.get(key) or "")].append(row)
        for phase, items in sorted(buckets.items(), key=lambda item: _long_wave_rank(item[0])):
            good = sum(1 for item in items if bool(item.get("is_good_stas1_review")))
            up_values = [_safe_float(item.get(pct_key), -1.0) for item in items]
            pullback_values = [_safe_float(item.get(pullback_key), -1.0) for item in items]
            valid_up = [value for value in up_values if value >= 0]
            valid_pullback = [value for value in pullback_values if value >= 0]
            out.append(
                {
                    "window": f"pre_{minutes}m",
                    "long_wave_phase": phase,
                    "long_wave_rank": _long_wave_rank(phase),
                    "entries": len(items),
                    "stas1_good": good,
                    "stas1_bad": len(items) - good,
                    "stas1_good_share": round(good / max(len(items), 1), 6),
                    "avg_up_from_low_pct": round(sum(valid_up) / max(len(valid_up), 1), 6),
                    "avg_pullback_from_high_pct": round(sum(valid_pullback) / max(len(valid_pullback), 1), 6),
                    "boundary": "LONG wave is computed before entry; no TP/exit here",
                }
            )
    return out


def _entry_setup_summary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    buckets: dict[tuple[int, str, str], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        key = (
            int(_safe_float(row.get("entry_setup_quality_rank"), -1.0)),
            str(row.get("entry_setup_quality_code") or ""),
            str(row.get("entry_setup_quality_label") or ""),
        )
        buckets[key].append(row)
    out: list[dict[str, Any]] = []
    for (rank, code, label), items in sorted(buckets.items(), key=lambda item: (item[0][0], item[0][1])):
        good = sum(1 for item in items if bool(item.get("is_good_stas1_review")))
        reasons = Counter(str(item.get("entry_setup_quality_reason") or "") for item in items)
        out.append(
            {
                "entry_setup_quality_rank": rank,
                "entry_setup_quality_code": code,
                "entry_setup_quality_label": label,
                "entries": len(items),
                "stas1_good": good,
                "stas1_bad": len(items) - good,
                "stas1_good_share": round(good / max(len(items), 1), 6),
                "dominant_reason": reasons.most_common(1)[0][0] if reasons else "",
                "boundary": "setup quality is pre-entry visual review context, not TP/exit and not ML label",
            }
        )
    return out


def _day_summary(rows: list[dict[str, Any]], hour_rows: list[dict[str, Any]], macro_wave_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
    hours_by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
    waves_by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_day[str(row["day_utc"])].append(row)
    for row in hour_rows:
        hours_by_day[str(row["day_utc"])].append(row)
    for row in macro_wave_rows:
        waves_by_day[str(row["day_utc"])].append(row)
    days = sorted(set(by_day) | set(hours_by_day) | set(waves_by_day))
    out: list[dict[str, Any]] = []
    for day in days:
        entries = by_day.get(day, [])
        hours = hours_by_day.get(day, [])
        waves = waves_by_day.get(day, [])
        confirmed_waves = [item for item in waves if str(item.get("macro_wave_segment_kind") or "WAVE") == "WAVE"]
        gap_segments = [item for item in waves if str(item.get("macro_wave_segment_kind") or "") == "GAP"]
        good = sum(1 for item in entries if bool(item.get("is_good_stas1_review")))
        avg_rank = sum(float(item["market_phase_rank"]) for item in hours) / max(len(hours), 1) if hours else 0.0
        avg_long_wave_rank = sum(float(item.get("hour_long_wave_rank", -1)) for item in hours) / max(len(hours), 1) if hours else 0.0
        avg_short_wave_rank = sum(float(item.get("hour_short_wave_rank", -1)) for item in hours) / max(len(hours), 1) if hours else 0.0
        out.append(
            {
                "day_utc": day,
                "day_type": _day_type(pd.Timestamp(day, tz="UTC")),
                "entries": len(entries),
                "stas1_good": good,
                "stas1_bad": len(entries) - good,
                "setup_clean": sum(1 for item in entries if int(_safe_float(item.get("entry_setup_quality_rank"), -1.0)) >= 3),
                "setup_working": sum(1 for item in entries if int(_safe_float(item.get("entry_setup_quality_rank"), -1.0)) == 2),
                "setup_warn": sum(1 for item in entries if int(_safe_float(item.get("entry_setup_quality_rank"), -1.0)) == 1),
                "setup_noise": sum(1 for item in entries if int(_safe_float(item.get("entry_setup_quality_rank"), -1.0)) <= 0),
                "avg_hour_phase_rank": round(avg_rank, 6),
                "dominant_hour_phase": Counter(str(item.get("market_phase") or "") for item in hours).most_common(1)[0][0] if hours else "",
                "strong_plus_hours": sum(1 for item in hours if int(item.get("market_phase_rank", -1)) >= _phase_rank("Сильная")),
                "dead_weak_hours": sum(1 for item in hours if int(item.get("market_phase_rank", -1)) <= _phase_rank("Слабая")),
                "avg_hour_long_wave_rank": round(avg_long_wave_rank, 6),
                "dominant_hour_long_wave": Counter(str(item.get("hour_long_wave_phase") or "") for item in hours).most_common(1)[0][0] if hours else "",
                "working_long_wave_hours": sum(1 for item in hours if int(item.get("hour_long_wave_rank", -1)) >= _long_wave_rank("Рабочая LONG-волна")),
                "micro_or_no_long_wave_hours": sum(1 for item in hours if int(item.get("hour_long_wave_rank", -1)) <= _long_wave_rank("Микроход")),
                "avg_hour_short_wave_rank": round(avg_short_wave_rank, 6),
                "dominant_hour_short_wave": Counter(str(item.get("hour_short_wave_phase") or "") for item in hours).most_common(1)[0][0] if hours else "",
                "working_short_wave_hours": sum(1 for item in hours if int(item.get("hour_short_wave_rank", -1)) >= _short_wave_rank("Рабочая SHORT-волна")),
                "hour_direction_bias_counts": dict(Counter(str(item.get("hour_direction_bias") or "") for item in hours)),
                "macro_wave_count": len(confirmed_waves),
                "macro_gap_segment_count": len(gap_segments),
                "macro_gap_total_duration_min": round(sum(_safe_float(item.get("macro_wave_duration_min"), 0.0) for item in gap_segments), 3),
                "macro_long_wave_count": sum(1 for item in confirmed_waves if str(item.get("macro_wave_direction")) == "LONG"),
                "macro_short_wave_count": sum(1 for item in confirmed_waves if str(item.get("macro_wave_direction")) == "SHORT"),
                "max_macro_wave_move_pct": round(max([_safe_float(item.get("macro_wave_move_pct"), 0.0) for item in confirmed_waves] or [0.0]), 6),
                "max_macro_gap_range_pct": round(max([_safe_float(item.get("macro_wave_range_pct"), 0.0) for item in gap_segments] or [0.0]), 6),
                "macro_wave_boundary": "resolved_macro_wave_hindsight_review_not_entry_feature",
            }
        )
    return out


def _columns_for(rows: list[dict[str, Any]]) -> list[str]:
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    return columns


def _xlsx_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, (bool, int, float)) or value is None:
        return value
    return str(value)


def _write_xlsx(path: Path, sheets: list[tuple[str, list[dict[str, Any]]]]) -> None:
    wb = Workbook()
    first = True
    for title, rows in sheets:
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = title[:31]
        columns = _columns_for(rows)
        if not columns:
            columns = ["empty"]
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row_idx, row in enumerate(rows, 2):
            for col_idx, column in enumerate(columns, 1):
                ws.cell(row=row_idx, column=col_idx, value=_xlsx_value(row.get(column)))
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_idx, column in enumerate(columns, 1):
            sample = [str(row.get(column, "")) for row in rows[:200]]
            width = min(max([len(column), *(len(value) for value in sample)]) + 2, 52)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    wb.save(path)


def _session_spans_for_day(day: str) -> list[dict[str, Any]]:
    start = pd.Timestamp(day, tz="UTC")
    spans: list[dict[str, Any]] = []
    for code, label, start_minute, end_minute in SESSION_TIME_BUCKETS:
        spans.append(
            {
                "code": code,
                "label": label,
                "start": start + pd.Timedelta(minutes=start_minute),
                "end": start + pd.Timedelta(minutes=end_minute),
                "color": SESSION_COLORS.get(code, "#607d8b"),
            }
        )
    return spans


def _shade_sessions(ax: Any, day: str, *, label_top: bool = False) -> None:
    ymin, ymax = ax.get_ylim()
    for span in _session_spans_for_day(day):
        start = span["start"].tz_convert(None)
        end = span["end"].tz_convert(None)
        ax.axvspan(start, end, color=span["color"], alpha=0.075, zorder=0)
        if label_top:
            center = start + (end - start) / 2
            ax.text(
                center,
                ymax,
                span["label"],
                color="#cfd8dc",
                fontsize=7,
                ha="center",
                va="top",
                rotation=0,
                alpha=0.84,
            )
    ax.set_ylim(ymin, ymax)


def _phase_color(rank: int) -> Any:
    cmap = plt.get_cmap("viridis", len(PHASES))
    return cmap(min(max(rank, 0), len(PHASES) - 1))


def _long_wave_color(rank: int) -> Any:
    cmap = plt.get_cmap("plasma", len(LONG_WAVE_PHASES))
    return cmap(min(max(rank, 0), len(LONG_WAVE_PHASES) - 1))


def _short_wave_color(rank: int) -> Any:
    cmap = plt.get_cmap("magma", len(SHORT_WAVE_PHASES))
    return cmap(min(max(rank, 0), len(SHORT_WAVE_PHASES) - 1))


def _macro_wave_color(direction: str) -> str:
    if direction == "LONG":
        return "#2fd17c"
    if direction == "SHORT":
        return "#ff7043"
    return "#607d8b"


def _strip_text_color(fill_color: Any) -> str:
    red, green, blue = fill_color[:3]
    luminance = 0.2126 * red + 0.7152 * green + 0.0722 * blue
    return "#101418" if luminance >= 0.58 else "white"


def _day_time_ticks(day: str) -> tuple[list[pd.Timestamp], list[str]]:
    start = pd.Timestamp(day)
    ticks = [start + pd.Timedelta(hours=hour) for hour in DAY_TIME_TICK_HOURS]
    labels = [tick.strftime("%H:%M") for tick in ticks]
    return ticks, labels


def _set_day_time_axis(ax: Any, day: str) -> None:
    ticks, labels = _day_time_ticks(day)
    ax.set_xlim(ticks[0].to_pydatetime(), ticks[-1].to_pydatetime())
    ax.set_xticks([tick.to_pydatetime() for tick in ticks])
    tick_labels = ax.set_xticklabels(labels, color="#cfd8dc", fontsize=8, rotation=25)
    for idx, label in enumerate(tick_labels):
        if idx == 0:
            label.set_ha("left")
        elif idx == len(tick_labels) - 1:
            label.set_ha("right")
        else:
            label.set_ha("center")
    ax.tick_params(axis="x", colors="#cfd8dc", labelsize=8, pad=2)


def _render_rank_strip(
    ax: Any,
    rows: list[dict[str, Any]],
    day: str,
    *,
    rank_field: str,
    label: str,
    color_kind: str,
    pct_field: str = "",
) -> None:
    ax.set_facecolor("#101418")
    ax.set_yticks([])
    ax.set_ylim(0, 1)
    ax.grid(False)
    for spine in ax.spines.values():
        spine.set_color("#3a444b")
    for row in rows:
        hour_start = pd.Timestamp(row["hour_start_utc"]).tz_convert(None)
        hour_end = hour_start + pd.Timedelta(hours=1)
        rank = int(row.get(rank_field, -1))
        if color_kind == "long":
            color = _long_wave_color(rank)
        elif color_kind == "short":
            color = _short_wave_color(rank)
        else:
            color = _phase_color(rank)
        ax.add_patch(
            Rectangle(
                (mdates.date2num(hour_start), 0.08),
                mdates.date2num(hour_end) - mdates.date2num(hour_start),
                0.84,
                facecolor=color,
                edgecolor="#101418",
                linewidth=0.45,
                alpha=0.95,
            )
        )
        text = str(rank)
        if pct_field:
            text = f"{rank}\n{_safe_float(row.get(pct_field), 0.0):.2f}%"
        ax.text(
            hour_start + pd.Timedelta(minutes=30),
            0.52,
            text,
            color=_strip_text_color(color),
            ha="center",
            va="center",
            fontsize=6,
            fontweight="bold",
        )
    start = pd.Timestamp(day)
    end = start + pd.Timedelta(days=1)
    ax.set_xlim(start.to_pydatetime(), end.to_pydatetime())
    ax.set_ylabel(label, color="white")


def _render_macro_wave_strip(ax: Any, wave_rows: list[dict[str, Any]], day: str) -> None:
    ax.set_facecolor("#101418")
    ax.set_yticks([])
    ax.set_ylim(0, 1)
    ax.grid(False)
    for spine in ax.spines.values():
        spine.set_color("#3a444b")
    for row in wave_rows:
        start = pd.Timestamp(row["macro_wave_start_time_utc"]).tz_convert(None)
        end = pd.Timestamp(row["macro_wave_end_time_utc"]).tz_convert(None)
        if end <= start:
            continue
        segment_kind = str(row.get("macro_wave_segment_kind") or "WAVE")
        direction = str(row.get("macro_wave_direction") or "")
        color = _macro_wave_color(direction)
        ax.add_patch(
            Rectangle(
                (mdates.date2num(start), 0.08),
                mdates.date2num(end) - mdates.date2num(start),
                0.84,
                facecolor=color,
                edgecolor="#101418",
                linewidth=0.50,
                alpha=0.88,
            )
        )
        duration_min = _safe_float(row.get("macro_wave_duration_min"), 0.0)
        visible_pct_for_label = _safe_float(row.get("macro_wave_visible_move_pct"), _safe_float(row.get("macro_wave_move_pct"), 0.0))
        strong_short_wave_label = segment_kind == "WAVE" and visible_pct_for_label >= MACRO_WAVE_REVERSAL_PCT and duration_min >= 5.0
        min_label_minutes = 8.0 if segment_kind == "GAP" else 15.0
        if duration_min >= min_label_minutes or strong_short_wave_label:
            if direction == "LONG":
                color_tuple = (0.18, 0.82, 0.49)
            elif direction == "SHORT":
                color_tuple = (1.0, 0.44, 0.26)
            else:
                color_tuple = (0.38, 0.49, 0.55)
            if segment_kind == "GAP":
                text = f"GAP\n{_safe_float(row.get('macro_wave_move_pct'), 0.0):.2f}%"
            else:
                wave_no = int(_safe_float(row.get("macro_wave_no"), 0.0))
                prefix = f"W{wave_no:02d} {direction}"
                carry_from_prev = bool(row.get("macro_wave_carry_from_prev_day"))
                carry_to_next = bool(row.get("macro_wave_carry_to_next_day"))
                if carry_from_prev:
                    prefix = "< " + prefix
                if carry_to_next:
                    prefix = prefix + " >"
                visible_pct = visible_pct_for_label
                full_pct = _safe_float(row.get("macro_wave_full_move_pct"), visible_pct)
                if duration_min < min_label_minutes:
                    text = f"{visible_pct:.2f}%"
                elif (carry_from_prev or carry_to_next) and duration_min >= 35.0:
                    text = f"{prefix}\nvis {visible_pct:.2f}%\nfull {full_pct:.2f}%"
                else:
                    text = f"{prefix}\n{visible_pct:.2f}%"
            ax.text(
                start + (end - start) / 2,
                0.52,
                text,
                color=_strip_text_color(color_tuple),
                ha="center",
                va="center",
                fontsize=4.5 if duration_min < min_label_minutes else (5 if duration_min < 35.0 or "\nfull " in text else 6),
                fontweight="bold",
            )
        ax.axvline(start, color=color, alpha=0.42, linewidth=0.75)
        ax.axvline(end, color=color, alpha=0.42, linewidth=0.75)
    start_day = pd.Timestamp(day)
    end_day = start_day + pd.Timedelta(days=1)
    ax.set_xlim(start_day.to_pydatetime(), end_day.to_pydatetime())
    ax.set_ylabel("WAVE", color="white")


def _render_day_overview(
    *,
    df: pd.DataFrame,
    entry_rows: list[dict[str, Any]],
    hour_rows: list[dict[str, Any]],
    macro_wave_rows: list[dict[str, Any]],
    day: str,
    symbol: str,
    timeframe: str,
    out_path: Path,
) -> None:
    fig, (ax_price, ax_bg_phase, ax_long_wave, ax_short_wave, ax_macro_wave, ax_vol) = plt.subplots(
        6,
        1,
        figsize=(32, 17.4),
        sharex=True,
        gridspec_kw={"height_ratios": OVERVIEW_HEIGHT_RATIOS},
    )
    fig.patch.set_facecolor("#101418")
    _style_axis(ax_price)
    _style_axis(ax_vol)
    _draw_candles(ax_price, df, timeframe, linewidth=0.30)
    _shade_sessions(ax_price, day, label_top=True)
    colors = ["#26a69a" if c >= o else "#ef5350" for o, c in zip(df["open"], df["close"])]
    ax_vol.bar(df["open_time_utc"].dt.tz_convert(None), df["volume"], width=_bar_width_days(timeframe), color=colors, alpha=0.70)
    _shade_sessions(ax_vol, day)
    _render_rank_strip(
        ax_bg_phase,
        hour_rows,
        day,
        rank_field="hour_background_phase_rank",
        label="Фон",
        color_kind="phase",
        pct_field="hour_background_phase_metric_pct",
    )
    _render_rank_strip(
        ax_long_wave,
        hour_rows,
        day,
        rank_field="hour_long_wave_rank",
        label="LONG",
        color_kind="long",
        pct_field="hour_long_wave_up_from_low_pct",
    )
    _render_rank_strip(
        ax_short_wave,
        hour_rows,
        day,
        rank_field="hour_short_wave_rank",
        label="SHORT",
        color_kind="short",
        pct_field="hour_short_wave_down_from_high_pct",
    )
    _render_macro_wave_strip(ax_macro_wave, macro_wave_rows, day)
    for row in macro_wave_rows:
        start = pd.Timestamp(row["macro_wave_start_time_utc"]).tz_convert(None)
        end = pd.Timestamp(row["macro_wave_end_time_utc"]).tz_convert(None)
        color = _macro_wave_color(str(row.get("macro_wave_direction") or ""))
        ax_price.axvline(start, color=color, alpha=0.12, linewidth=1.05, zorder=1)
        ax_price.axvline(end, color=color, alpha=0.12, linewidth=1.05, zorder=1)

    for row in entry_rows:
        entry_time = pd.Timestamp(row["entry_time_utc"]).tz_convert(None)
        anchor_raw = row.get("anchor_time_utc") or row["entry_time_utc"]
        anchor_time = pd.Timestamp(anchor_raw).tz_convert(None)
        anchor_low = _safe_float(row.get("anchor_low_price"))
        entry_open = _safe_float(row.get("entry_open_price"))
        color = _label_color(row)
        is_good = bool(row.get("is_good_stas1_review"))
        ax_price.axvline(entry_time, color=color, alpha=0.18 if is_good else 0.10, linewidth=0.9, zorder=3)
        ax_price.scatter([anchor_time], [anchor_low], s=18 if is_good else 12, color="#ff5252", edgecolors="#0b0f12", linewidths=0.35, alpha=0.86, zorder=6)
        if is_good:
            ax_price.scatter(
                [entry_time],
                [entry_open],
                marker="^",
                s=60,
                color=color,
                edgecolors="white",
                linewidths=0.45,
                alpha=0.92,
                zorder=7,
            )
        else:
            ax_price.scatter(
                [entry_time],
                [entry_open],
                marker="x",
                s=66,
                color=color,
                linewidths=1.15,
                alpha=0.92,
                zorder=7,
            )
    start = pd.Timestamp(day)
    end = start + pd.Timedelta(days=1)
    ax_price.set_xlim(start.to_pydatetime(), end.to_pydatetime())
    day_type = _day_type(pd.Timestamp(day, tz="UTC"))
    ax_price.set_title(
        f"STAS2 {symbol} {timeframe} {day} | {day_type} | фон + LONG/SHORT + WAVE review | без TP/exit/future",
        color="white",
        fontsize=15,
    )
    ax_price.set_ylabel("Price", color="white")
    ax_vol.set_ylabel("Volume", color="white")
    _set_day_time_axis(ax_vol, day)
    fig.tight_layout()
    fig.savefig(out_path, dpi=155, facecolor=fig.get_facecolor())
    plt.close(fig)


def _row_context_text(row: dict[str, Any]) -> str:
    parts = [
        f"{row['candidate_id']} {row['review_label']}",
        f"{row['effective_session_label']} | {row['day_type']}",
        f"setup {row.get('entry_setup_quality_label', '')} | {row.get('entry_setup_quality_reason', '')}",
        f"stas1 risk_flags {json.dumps(row.get('stas1_risk_flags') or [], ensure_ascii=False)}",
    ]
    for minutes in PRE_ENTRY_WINDOWS_MIN:
        phase = row.get(f"pre_{minutes}m_background_phase", row.get(f"pre_{minutes}m_phase", ""))
        rng = row.get(f"pre_{minutes}m_range_pct", "")
        move = row.get(f"pre_{minutes}m_close_move_pct", "")
        wave = row.get(f"pre_{minutes}m_long_wave_up_from_low_pct", "")
        pullback = row.get(f"pre_{minutes}m_long_wave_pullback_from_post_low_high_pct", "")
        wave_phase = row.get(f"pre_{minutes}m_long_wave_phase", "")
        parts.append(f"{minutes}m bg {phase} range {_fmt_pct(rng)}% move {_fmt_pct(move)}%")
        parts.append(f"     pre-wave {wave_phase} up {_fmt_pct(wave)}% pb {_fmt_pct(pullback)}%")
    parts.append(f"session so far {_fmt_pct(row.get('session_so_far_range_pct', ''))}% | day so far {_fmt_pct(row.get('day_so_far_range_pct', ''))}%")
    return "\n".join(parts)


def _render_entry_context_pages(
    *,
    frames_by_day: dict[str, pd.DataFrame],
    rows: list[dict[str, Any]],
    symbol: str,
    timeframe: str,
    out_dir: Path,
    file_prefix: str = "STAS2_ENTRY_CONTEXT_PAGE",
    render_limit: int = 0,
) -> list[Path]:
    sorted_rows = sorted(rows, key=lambda item: (str(item["entry_time_utc"]), str(item["record_id"])))
    if render_limit > 0:
        sorted_rows = sorted_rows[:render_limit]
    paths: list[Path] = []
    per_page = 6
    for page_start in range(0, len(sorted_rows), per_page):
        page_rows = sorted_rows[page_start : page_start + per_page]
        fig, axes = plt.subplots(3, 2, figsize=(22, 22), sharex=False)
        fig.patch.set_facecolor("#101418")
        flat = list(axes.flatten())
        for ax, row in zip(flat, page_rows):
            _style_axis(ax)
            day = str(row["day_utc"])
            df = frames_by_day.get(day)
            if df is None:
                ax.axis("off")
                continue
            entry_ts = pd.Timestamp(row["entry_time_utc"])
            entry_ts = entry_ts.tz_convert("UTC") if entry_ts.tzinfo is not None else entry_ts.tz_localize("UTC")
            start = entry_ts - pd.Timedelta(minutes=75)
            win = df[(df["open_time_utc"] >= start) & (df["open_time_utc"] < entry_ts)].reset_index(drop=True)
            if win.empty:
                ax.axis("off")
                continue
            _draw_candles(ax, win, timeframe, linewidth=0.42)
            _shade_sessions(ax, day)
            entry_naive = entry_ts.tz_convert(None)
            entry_open = _safe_float(row.get("entry_open_price"))
            color = _label_color(row)
            ax.axvline(entry_naive, color=color, linewidth=1.15, alpha=0.70)
            if bool(row.get("is_good_stas1_review")):
                ax.scatter(
                    [entry_naive],
                    [entry_open],
                    marker="^",
                    s=118,
                    color=color,
                    edgecolors="white",
                    linewidths=0.6,
                    zorder=8,
                )
            else:
                ax.scatter(
                    [entry_naive],
                    [entry_open],
                    marker="x",
                    s=118,
                    color=color,
                    linewidths=1.8,
                    zorder=8,
                )
            y0, y1 = ax.get_ylim()
            ax.text(
                0.012,
                0.985,
                _row_context_text(row),
                transform=ax.transAxes,
                ha="left",
                va="top",
                color="#f5f5f5",
                fontsize=7,
                bbox={"facecolor": "#101418", "edgecolor": "#455a64", "alpha": 0.78, "boxstyle": "round,pad=0.28"},
            )
            ax.set_xlim(start.tz_convert(None).to_pydatetime(), entry_naive.to_pydatetime())
            ax.set_ylim(y0, y1)
            ax.set_title(
                f"{symbol} {day} | pre-entry only | entry {entry_ts.strftime('%H:%M')}",
                color="white",
                fontsize=9,
            )
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            ax.tick_params(axis="x", labelrotation=25)
        for ax in flat[len(page_rows) :]:
            ax.axis("off")
            ax.set_facecolor("#101418")
        fig.suptitle("STAS2 entry context: свечи после entry не рисуются, TP/exit относятся к STAS3", color="white", fontsize=15)
        fig.tight_layout(rect=[0, 0, 1, 0.985])
        page_no = len(paths) + 1
        path = out_dir / f"{file_prefix}_{page_no:02d}.png"
        fig.savefig(path, dpi=150, facecolor=fig.get_facecolor())
        plt.close(fig)
        paths.append(path)
    return paths


def _render_run_index(*, run_dir: Path, day_summary: list[dict[str, Any]], out_path: Path, symbol: str, timeframe: str) -> None:
    fig, ax = plt.subplots(figsize=(15, max(7, 1.0 + len(day_summary) * 0.72)))
    fig.patch.set_facecolor("#101418")
    ax.set_facecolor("#101418")
    ax.axis("off")
    ax.set_title(f"STAS2 Market Phase Review | {symbol} {timeframe} | index", color="white", fontsize=15, loc="left")
    lines = [
        "Открыть папку дня, затем 00_OVERVIEW и страницы ENTRY_CONTEXT.",
        "Граница: Stas2 показывает только контекст до входа; TP/exit будут в Stas3.",
        "",
        "day | type | entries | GOOD/BAD | avg bg | avg LONG | avg SHORT | macro W | strong+ | weak",
    ]
    for row in day_summary:
        lines.append(
            f"{row['day_utc']} | {row['day_type']} | {row['entries']} | {row['stas1_good']}/{row['stas1_bad']} | "
            f"{float(row['avg_hour_phase_rank']):.2f} {row['dominant_hour_phase']} | "
            f"{float(row.get('avg_hour_long_wave_rank', 0.0)):.2f} {row.get('dominant_hour_long_wave', '')} | "
            f"{float(row.get('avg_hour_short_wave_rank', 0.0)):.2f} {row.get('dominant_hour_short_wave', '')} | "
            f"{row.get('macro_wave_count', 0)} | {row['strong_plus_hours']} | {row['dead_weak_hours']}"
        )
    ax.text(0.02, 0.93, "\n".join(lines), transform=ax.transAxes, va="top", ha="left", color="#eceff1", fontsize=10, family="monospace")
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, facecolor=fig.get_facecolor())
    plt.close(fig)


def _build_browse_by_day(
    *,
    run_dir: Path,
    frames_by_day: dict[str, pd.DataFrame],
    rows: list[dict[str, Any]],
    hour_rows: list[dict[str, Any]],
    macro_wave_rows: list[dict[str, Any]],
    day_summary: list[dict[str, Any]],
    day_overviews: dict[str, Path],
    symbol: str,
    timeframe: str,
    render_limit: int,
) -> dict[str, Any]:
    browse_dir = run_dir / "BROWSE_BY_DAY"
    browse_dir.mkdir(parents=True, exist_ok=True)
    rows_by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
    hours_by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
    waves_by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        rows_by_day[str(row["day_utc"])].append(row)
    for row in hour_rows:
        hours_by_day[str(row["day_utc"])].append(row)
    for row in macro_wave_rows:
        waves_by_day[str(row["day_utc"])].append(row)

    days_payload: list[dict[str, Any]] = []
    for day_row in day_summary:
        day = str(day_row["day_utc"])
        compact = day.replace("-", "")
        day_dir = browse_dir / day
        day_dir.mkdir(parents=True, exist_ok=True)
        day_entries = sorted(rows_by_day.get(day, []), key=lambda item: str(item["entry_time_utc"]))
        overview_dst = day_dir / f"00_{compact}_OVERVIEW.png"
        src = day_overviews.get(day)
        if src and src.exists():
            overview_dst.write_bytes(src.read_bytes())
        else:
            overview_dst = Path("")
        pages = _render_entry_context_pages(
            frames_by_day=frames_by_day,
            rows=day_entries,
            symbol=symbol,
            timeframe=timeframe,
            out_dir=day_dir,
            file_prefix=f"{compact}_ENTRY_CONTEXT_PAGE",
            render_limit=render_limit,
        )
        records_csv = day_dir / f"{compact}_STAS2_RECORDS.csv"
        _write_csv(records_csv, day_entries)
        macro_waves_csv = day_dir / f"{compact}_STAS2_MACRO_WAVES.csv"
        _write_csv(macro_waves_csv, waves_by_day.get(day, []))
        days_payload.append(
            {
                "day_utc": day,
                "folder": day_dir,
                "overview": overview_dst if str(overview_dst) else None,
                "entry_context_pages": pages,
                "records_csv": records_csv,
                "macro_waves_csv": macro_waves_csv,
                "entries": len(day_entries),
                "stas1_good": sum(1 for item in day_entries if bool(item.get("is_good_stas1_review"))),
                "stas1_bad": sum(1 for item in day_entries if not bool(item.get("is_good_stas1_review"))),
                "hour_rows": len(hours_by_day.get(day, [])),
                "macro_wave_rows": len(waves_by_day.get(day, [])),
            }
        )

    index_png = browse_dir / "00_RUN_INDEX.png"
    index_md = browse_dir / "00_RUN_INDEX_RU.md"
    _render_run_index(run_dir=run_dir, day_summary=day_summary, out_path=index_png, symbol=symbol, timeframe=timeframe)
    lines = [
        "# STAS2 browse by day",
        "",
        "Назначение: смотреть фазы, сессии и pre-entry контекст на графиках по дням.",
        "",
        "Порядок: открыть папку дня, открыть `00_OVERVIEW`, затем листать `ENTRY_CONTEXT_PAGE`.",
        "",
        "| День | Entries | GOOD | BAD | Папка |",
        "|---|---:|---:|---:|---|",
    ]
    for item in days_payload:
        lines.append(f"| {item['day_utc']} | {item['entries']} | {item['stas1_good']} | {item['stas1_bad']} | `{item['folder'].name}` |")
    index_md.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return {"browse_dir": browse_dir, "index_png": index_png, "index_md": index_md, "days": days_payload}


def _write_report(path: Path, payload: dict[str, Any]) -> None:
    summary = payload["summary"]
    artifacts = payload["artifacts"]
    lines = [
        "# STAS2 Market Phase Review",
        "",
        f"Статус: `{payload['status']}`.",
        "",
        "Назначение: перенести фазы рынка, сессии и проценты движения цены на отдельные Stas2-графики поверх Stas1 входов, не меняя Stas1.",
        "",
        "## Главное правило",
        "",
        "Stas2 считает и рисует только то, что доступно до входа. Свечи после `entry_time_utc`, TP, exit, MFE/MAE и процентная лестница сделки относятся к Stas3.",
        "",
        "## Что построено",
        "",
        f"- Диапазон: `{payload['date_range']['start_day']}` .. `{payload['date_range']['end_day']}`.",
        f"- Обработано Stas1 entry rows: `{summary['entry_rows']}`.",
        f"- GOOD по Stas1 review: `{summary['stas1_good']}`.",
        f"- BAD по Stas1 review: `{summary['stas1_bad']}`.",
        f"- Setup clean/working/warn/noise: `{summary.get('setup_clean', 0)}` / `{summary.get('setup_working', 0)}` / `{summary.get('setup_warn', 0)}` / `{summary.get('setup_noise', 0)}`.",
        f"- Continuous-wave rows: `{summary.get('continuous_wave_rows', 0)}` = waves `{summary.get('continuous_confirmed_wave_rows', 0)}` + gaps `{summary.get('continuous_gap_rows', 0)}`.",
        f"- Macro-wave day slices: `{summary.get('macro_wave_rows', 0)}` = wave slices `{summary.get('macro_confirmed_wave_rows', 0)}` + gap slices `{summary.get('macro_gap_rows', 0)}`.",
        f"- Дней с графиками: `{summary['days_processed']}`.",
        "",
        "## Pre-entry поля",
        "",
        "- `pre_5m_*`, `pre_15m_*`, `pre_30m_*`, `pre_60m_*`: range, close-move, path и фаза до входа.",
        "- `pre_*_background_phase`: общий фон окна: range/volatility/path. Если здесь `Слабая`, это не значит, что LONG-волны не было.",
        "- `pre_*_long_wave_*`: LONG-волна внутри окна до входа: low -> subsequent high, откат от этой вершины к последнему закрытому close.",
        "- `hour_short_wave_*`: закрытый часовой обзор SHORT-движения: high -> subsequent low внутри часа.",
        "- `continuous_wave_*`: переменные swing-волны без привязки к дню. Это review/hindsight слой для глаз и разметки, а не causal feature для входа.",
        "- `macro_wave_*`: дневные срезы `continuous_wave_*` для overview PNG. Если волна идет через 00:00, на одном дне будет `carry_to_next`, на следующем `carry_from_prev`.",
        "- `entry_setup_quality_*`: отдельная pre-entry оценка чистоты конкретной точки: clean/working/warn/noise. Это не меняет Stas1 outcome и не является ML-label.",
        "- `stas1_risk_flags` и `stas1_feature_*`: исходные pre-entry признаки Stas1, по которым объясняется setup-quality.",
        "- `session_so_far_*`: ход с начала текущей UTC-корзины до входа.",
        "- `day_so_far_*`: ход с начала UTC-дня до входа.",
        "- `effective_session`: `day_type + session_time_bucket`, чтобы выходной Лондон не смешивался с будним Лондоном.",
        "",
        "На дневном overview полосы `Фон`, `LONG` и `SHORT` остаются audit-визуалом закрытого часа. Полоса `WAVE` показывает дневной срез непрерывных волн, поэтому волна может начинаться вчера и продолжаться сегодня. Точки входа рисуются как в Stas1, а текстовые названия возле точек не рисуются, чтобы график не шумел. `setup_quality` смотреть в CSV/XLSX и closeup-тексте.",
        "",
        "## Артефакты",
        "",
    ]
    for key, value in artifacts.items():
        lines.append(f"- `{key}`: `{value}`")
    lines.extend(
        [
            "",
            "## Граница Stas2/Stas3",
            "",
            "- Stas2: контекст до входа, визуальная фаза, сессия, процент движения до входа.",
            "- Stas2 continuous WAVE: review/hindsight слой для визуальной сверки больших движений; не переносить в ML как готовый входной признак без отдельной causal-разметки.",
            "- Stas3: жизнь сделки после входа, 5m-блоки, TP/exit, процентная лестница, зависание сделки, drawdown.",
            "- Stas2 не запускает ML, Optuna, scorer, target-lock или API.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_review(
    *,
    start_day: str,
    end_day: str,
    symbol: str,
    timeframe: str,
    out_dir: Path,
    run_label: str,
    stas1_run_dirs: list[str],
    render_limit: int,
) -> dict[str, Any]:
    root = Path(__file__).resolve().parents[2]
    if not out_dir.is_absolute():
        out_dir = root / out_dir
    run_id = f"{run_label}_{_run_stamp()}" if run_label else f"stas2_review_{start_day.replace('-', '')}_{end_day.replace('-', '')}_{_run_stamp()}"
    run_dir = out_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    days = _iter_days(start_day, end_day)
    df, frames_by_day, missing_sources = _load_days(root, days, symbol, timeframe)
    entry_rows, stas1_runs_used = _load_entry_context_rows(root=root, df=df, start_day=start_day, end_day=end_day, stas1_run_dirs=stas1_run_dirs)
    hour_rows = _hour_phase_rows(df)
    continuous_wave_rows = _continuous_wave_rows(df)
    macro_wave_rows = _macro_wave_rows(df, continuous_rows=continuous_wave_rows)
    session_summary = _entry_summary(
        entry_rows,
        ["day_type", "session_time_bucket_code", "session_time_bucket_label", "effective_session_label"],
        "pre_60m_background_phase",
    )
    phase_summary = _entry_phase_summary(entry_rows)
    long_wave_summary = _entry_long_wave_summary(entry_rows)
    setup_summary = _entry_setup_summary(entry_rows)
    day_summary = _day_summary(entry_rows, hour_rows, macro_wave_rows)

    records_csv = run_dir / "STAS2_RECORDS.csv"
    hour_csv = run_dir / "STAS2_HOURLY_PHASES.csv"
    continuous_wave_csv = run_dir / "STAS2_CONTINUOUS_WAVES.csv"
    macro_wave_csv = run_dir / "STAS2_MACRO_WAVES.csv"
    session_csv = run_dir / "STAS2_ENTRY_SESSION_SUMMARY.csv"
    phase_csv = run_dir / "STAS2_ENTRY_PRE_PHASE_SUMMARY.csv"
    long_wave_csv = run_dir / "STAS2_ENTRY_LONG_WAVE_SUMMARY.csv"
    setup_csv = run_dir / "STAS2_ENTRY_SETUP_QUALITY_SUMMARY.csv"
    day_csv = run_dir / "STAS2_DAY_SUMMARY.csv"
    xlsx_path = run_dir / "STAS2_MARKET_PHASE_TABLES.xlsx"
    payload_path = run_dir / "STAS2_PAYLOAD.json"
    report_path = run_dir / "STAS2_REPORT_RU.md"

    _write_csv(records_csv, entry_rows)
    _write_csv(hour_csv, hour_rows)
    _write_csv(continuous_wave_csv, continuous_wave_rows)
    _write_csv(macro_wave_csv, macro_wave_rows)
    _write_csv(session_csv, session_summary)
    _write_csv(phase_csv, phase_summary)
    _write_csv(long_wave_csv, long_wave_summary)
    _write_csv(setup_csv, setup_summary)
    _write_csv(day_csv, day_summary)
    _write_xlsx(
        xlsx_path,
        [
            ("Entry context", entry_rows),
            ("Hourly phases", hour_rows),
            ("Continuous waves", continuous_wave_rows),
            ("Macro waves", macro_wave_rows),
            ("Entry sessions", session_summary),
            ("Pre phase summary", phase_summary),
            ("Long wave summary", long_wave_summary),
            ("Setup quality summary", setup_summary),
            ("Days", day_summary),
        ],
    )

    rows_by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
    hours_by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
    waves_by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in entry_rows:
        rows_by_day[str(row["day_utc"])].append(row)
    for row in hour_rows:
        hours_by_day[str(row["day_utc"])].append(row)
    for row in macro_wave_rows:
        waves_by_day[str(row["day_utc"])].append(row)

    day_overviews: dict[str, Path] = {}
    for day, frame in frames_by_day.items():
        path = run_dir / f"STAS2_DAY_OVERVIEW_{day.replace('-', '')}.png"
        _render_day_overview(
            df=frame,
            entry_rows=rows_by_day.get(day, []),
            hour_rows=hours_by_day.get(day, []),
            macro_wave_rows=waves_by_day.get(day, []),
            day=day,
            symbol=symbol,
            timeframe=timeframe,
            out_path=path,
        )
        day_overviews[day] = path

    entry_pages = _render_entry_context_pages(
        frames_by_day=frames_by_day,
        rows=entry_rows,
        symbol=symbol,
        timeframe=timeframe,
        out_dir=run_dir,
        render_limit=render_limit,
    )
    browse = _build_browse_by_day(
        run_dir=run_dir,
        frames_by_day=frames_by_day,
        rows=entry_rows,
        hour_rows=hour_rows,
        macro_wave_rows=macro_wave_rows,
        day_summary=day_summary,
        day_overviews=day_overviews,
        symbol=symbol,
        timeframe=timeframe,
        render_limit=render_limit,
    )

    good = sum(1 for row in entry_rows if bool(row.get("is_good_stas1_review")))
    artifacts = {
        "run_dir": _rel(root, run_dir),
        "report_ru": _rel(root, report_path),
        "json": _rel(root, payload_path),
        "records_csv": _rel(root, records_csv),
        "hourly_csv": _rel(root, hour_csv),
        "continuous_wave_csv": _rel(root, continuous_wave_csv),
        "macro_wave_csv": _rel(root, macro_wave_csv),
        "session_summary_csv": _rel(root, session_csv),
        "phase_summary_csv": _rel(root, phase_csv),
        "long_wave_summary_csv": _rel(root, long_wave_csv),
        "setup_quality_summary_csv": _rel(root, setup_csv),
        "day_summary_csv": _rel(root, day_csv),
        "xlsx": _rel(root, xlsx_path),
        "day_overviews": [_rel(root, path) for path in day_overviews.values()],
        "entry_context_pages": [_rel(root, path) for path in entry_pages],
        "browse_dir": _rel(root, browse["browse_dir"]),
        "browse_index_png": _rel(root, browse["index_png"]),
        "browse_index_md": _rel(root, browse["index_md"]),
        "browse_days": [
            {
                "day_utc": item["day_utc"],
                "folder": _rel(root, item["folder"]),
                "overview": _rel(root, item["overview"]) if item["overview"] else "",
                "entry_context_pages": [_rel(root, path) for path in item["entry_context_pages"]],
                "records_csv": _rel(root, item["records_csv"]),
                "macro_waves_csv": _rel(root, item["macro_waves_csv"]),
            }
            for item in browse["days"]
        ],
    }
    payload = {
        "schema_version": 2,
        "status": STATUS,
        "created_utc": _utc_now(),
        "run_id": run_id,
        "symbol": symbol,
        "timeframe": timeframe,
        "date_range": {"start_day": start_day, "end_day": end_day},
        "summary": {
            "days_requested": len(days),
            "days_processed": len(frames_by_day),
            "entry_rows": len(entry_rows),
            "stas1_good": good,
            "stas1_bad": len(entry_rows) - good,
            "setup_clean": sum(1 for row in entry_rows if int(_safe_float(row.get("entry_setup_quality_rank"), -1.0)) >= 3),
            "setup_working": sum(1 for row in entry_rows if int(_safe_float(row.get("entry_setup_quality_rank"), -1.0)) == 2),
            "setup_warn": sum(1 for row in entry_rows if int(_safe_float(row.get("entry_setup_quality_rank"), -1.0)) == 1),
            "setup_noise": sum(1 for row in entry_rows if int(_safe_float(row.get("entry_setup_quality_rank"), -1.0)) <= 0),
            "missing_sources": len(missing_sources),
            "continuous_wave_rows": len(continuous_wave_rows),
            "continuous_confirmed_wave_rows": sum(1 for row in continuous_wave_rows if str(row.get("continuous_wave_segment_kind") or "WAVE") == "WAVE"),
            "continuous_gap_rows": sum(1 for row in continuous_wave_rows if str(row.get("continuous_wave_segment_kind") or "") == "GAP"),
            "macro_wave_rows": len(macro_wave_rows),
            "macro_confirmed_wave_rows": sum(1 for row in macro_wave_rows if str(row.get("macro_wave_segment_kind") or "WAVE") == "WAVE"),
            "macro_gap_rows": sum(1 for row in macro_wave_rows if str(row.get("macro_wave_segment_kind") or "") == "GAP"),
        },
        "pre_entry_windows_min": PRE_ENTRY_WINDOWS_MIN,
        "session_time_buckets_utc": [
            {
                "session_time_bucket_code": code,
                "session_time_bucket_label": label,
                "start_time_utc": f"{start // 60:02d}:{start % 60:02d}",
                "end_time_utc": "24:00" if end == 1440 else f"{end // 60:02d}:{end % 60:02d}",
            }
            for code, label, start, end in SESSION_TIME_BUCKETS
        ],
        "phase_ladder": [
            {"rank": idx, "phase": name, "upper_phase_metric_pct": None if math.isinf(upper) else upper}
            for idx, (name, upper) in enumerate(PHASES)
        ],
        "long_wave_ladder": [
            {"rank": idx, "long_wave_phase": name, "upper_up_from_low_pct": None if math.isinf(upper) else upper}
            for idx, (name, upper) in enumerate(LONG_WAVE_PHASES)
        ],
        "short_wave_ladder": [
            {"rank": idx, "short_wave_phase": name, "upper_down_from_high_pct": None if math.isinf(upper) else upper}
            for idx, (name, upper) in enumerate(SHORT_WAVE_PHASES)
        ],
        "macro_wave_settings": {
            "reversal_threshold_pct": MACRO_WAVE_REVERSAL_PCT,
            "continuous_schema_version": MACRO_WAVE_CONTINUOUS_SCHEMA_VERSION,
            "boundary": "continuous_macro_wave_hindsight_review_not_entry_feature",
        },
        "phase_interpretation": {
            "background_phase": "общий фон волатильности/range окна или закрытого часа",
            "long_wave_phase": "направленная LONG-волна low -> subsequent high внутри pre-entry окна или закрытого часа",
            "short_wave_phase": "направленная SHORT-волна high -> subsequent low внутри закрытого часа",
            "continuous_wave": "глобальная swing-волна без привязки к дню; день показывает только видимый срез",
            "macro_wave": "дневной срез continuous-wave по развороту 1%; review-only, не causal feature для входа",
            "entry_setup_quality": "pre-entry оценка чистоты конкретного low-входа; это не TP/exit и не ML-label",
        },
        "stas1_runs_used": stas1_runs_used,
        "missing_sources": missing_sources,
        "entry_rows": entry_rows,
        "hour_rows": hour_rows,
        "continuous_wave_rows": continuous_wave_rows,
        "macro_wave_rows": macro_wave_rows,
        "entry_session_summary": session_summary,
        "entry_phase_summary": phase_summary,
        "entry_long_wave_summary": long_wave_summary,
        "entry_setup_summary": setup_summary,
        "day_summary": day_summary,
        "artifacts": artifacts,
        "guardrails": [
            "NO_ML_EXPORT_TRAINING",
            "NO_OPTUNA",
            "NO_SCORER",
            "NO_TARGET_LOCK",
            "NO_API",
            "PRE_ENTRY_ONLY",
            "NO_TP_EXIT_PERCENT_LADDER_IN_STAS2",
            "ALL_FEATURE_WINDOWS_USE_OPEN_TIME_UTC_LT_ENTRY_TIME_UTC",
        ],
    }
    _write_json(payload_path, payload)
    _write_report(report_path, payload)
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description="STAS2 visual market phase review, pre-entry only, no ML/Optuna.")
    parser.add_argument("--start-day", default="2026-05-02")
    parser.add_argument("--end-day", default="2026-05-03")
    parser.add_argument("--symbol", default="SOLUSDT")
    parser.add_argument("--timeframe", default="1m")
    parser.add_argument("--out-dir", default="STAS2_MARKET_PHASE_REVIEW/runs")
    parser.add_argument("--run-label", default="stas2_review_v0")
    parser.add_argument("--stas1-run-dir", action="append", default=[])
    parser.add_argument("--render-limit", type=int, default=0, help="0 means render all entry context rows")
    args = parser.parse_args()

    payload = run_review(
        start_day=args.start_day,
        end_day=args.end_day,
        symbol=args.symbol,
        timeframe=args.timeframe,
        out_dir=Path(args.out_dir),
        run_label=args.run_label,
        stas1_run_dirs=args.stas1_run_dir,
        render_limit=args.render_limit,
    )
    print(
        json.dumps(
            {
                "status": payload["status"],
                "run_id": payload["run_id"],
                "summary": payload["summary"],
                "artifacts": payload["artifacts"],
                "missing_sources": payload["missing_sources"],
                "stas1_runs_used": payload["stas1_runs_used"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
