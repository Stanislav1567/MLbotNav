from __future__ import annotations

import argparse
import csv
import json
import math
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import matplotlib.dates as mdates
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mlbotnav.visual_entry_low_anchor_suggester import _bar_width_days, _draw_candles, _load_ohlcv, _source_csv, _style_axis
from mlbotnav.visual_entry_stas2_market_phase_audit import _iter_days, _rel


STATUS = "STAS3_V2_PERCENT_LADDER_LONG_ONLY_READY_NO_ML_NO_OPTUNA_POST_ENTRY_AUDIT"


def _build_v2_percent_ladder() -> list[float]:
    fast = [round(x / 10.0, 1) for x in range(3, 10)]
    medium = [round(x / 10.0, 1) for x in range(10, 21)]
    extension = [round(x / 10.0, 1) for x in range(22, 201, 2)]
    out: list[float] = []
    for value in [*fast, *medium, *extension]:
        if value not in out:
            out.append(value)
    return out


PERCENT_LADDER = _build_v2_percent_ladder()
POST_WINDOWS_MIN = [5, 15, 30, 60]
TP_COLOR = "#ffd54f"
TP_FALLBACK_COLOR = "#b0bec5"
SIGNAL_COLOR = "#ff7043"
ENTRY_COLOR = "#26c6da"
TP_MOVE_COLOR = "#ff1744"
MFE_COLOR = "#76ff03"
MAE_COLOR = "#ba68c8"
MFE_MOVE_COLOR = "#00e676"
PATH_CHECKPOINTS = [0.3, 0.5, 1.0, 2.0]
V2_SOURCE_STAS2_RUN = "STAS2_MARKET_PHASE_REVIEW/runs/stas2_20260510_20260512_continuous_wave_v2_20260709_081330"
V2_WORK_DAYS = ["2026-05-10", "2026-05-11", "2026-05-12"]


def _utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _run_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _fmt_ts(ts: pd.Timestamp | None) -> str:
    if ts is None:
        return ""
    utc = ts.tz_convert("UTC") if ts.tzinfo is not None else ts.tz_localize("UTC")
    return utc.strftime("%Y-%m-%dT%H:%M:%SZ")


def _parse_utc_ts(value: Any) -> pd.Timestamp | None:
    if value is None or value == "":
        return None
    try:
        ts = pd.Timestamp(value)
    except Exception:
        return None
    if pd.isna(ts):
        return None
    return ts.tz_convert("UTC") if ts.tzinfo is not None else ts.tz_localize("UTC")


def _minutes_between(start: pd.Timestamp | None, end: pd.Timestamp | None) -> float | str:
    if start is None or end is None:
        return ""
    return round((end - start).total_seconds() / 60.0, 3)


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None or value == "":
            return default
        out = float(value)
        if math.isnan(out) or math.isinf(out):
            return default
        return out
    except Exception:
        return default


def _phase_key(pct: float) -> str:
    return f"{pct:.1f}".replace(".", "p")


def _visual_tp_fields(row: dict[str, Any]) -> dict[str, Any]:
    pct = _safe_float(row.get("reasonable_tp_pct"), 0.0)
    source = "phase TP review"
    color = TP_COLOR
    linestyle = "-"
    if pct <= 0:
        pct = _safe_float(row.get("ideal_review_tp_pct"), 0.0)
        source = "ideal review TP"
        color = TP_FALLBACK_COLOR
        linestyle = "--"
    if pct <= 0:
        return {
            "pct": 0.0,
            "key": "",
            "target": 0.0,
            "hit_time_utc": "",
            "time_to_min": "",
            "source": "NO_TP_REVIEW",
            "color": color,
            "linestyle": linestyle,
            "is_reasonable": False,
        }
    key = _phase_key(pct)
    target = _safe_float(row.get(f"target_{key}pct_price"), 0.0)
    hit_time = row.get(f"hit_time_{key}pct_utc") or ""
    time_to = row.get(f"time_to_{key}pct_min") or ""
    return {
        "pct": pct,
        "key": key,
        "target": target,
        "hit_time_utc": hit_time,
        "time_to_min": time_to,
        "source": source,
        "color": color,
        "linestyle": linestyle,
        "is_reasonable": _safe_float(row.get("reasonable_tp_pct"), 0.0) > 0,
    }


def _visual_signal_entry_fields(row: dict[str, Any]) -> dict[str, Any]:
    signal_time = row.get("anchor_time_utc") or row.get("stas1_confirmation_time_utc") or row.get("entry_time_utc") or ""
    entry_time = row.get("entry_time_utc") or ""
    signal_price = _safe_float(row.get("anchor_low_price"), 0.0)
    entry_open = _safe_float(row.get("entry_open_price"), 0.0)
    entry_price = _safe_float(row.get("entry_price_5bps"), entry_open)
    if signal_price <= 0:
        signal_price = entry_open if entry_open > 0 else entry_price
    return {
        "signal_time_utc": signal_time,
        "entry_time_utc": entry_time,
        "signal_price": signal_price,
        "entry_open_price": entry_open,
        "entry_price": entry_price,
    }


def _visual_mfe_mae_fields(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "mfe_time_utc": row.get("actual_mfe_time_utc") or "",
        "mfe_price": _safe_float(row.get("mfe_max_price"), _safe_float(row.get("max_high_hold"), 0.0)),
        "mfe_pct": _safe_float(row.get("actual_mfe_pct"), 0.0),
        "mae_time_utc": row.get("actual_mae_time_utc") or "",
        "mae_price": _safe_float(row.get("mae_min_price"), _safe_float(row.get("min_low_hold"), 0.0)),
        "mae_pct": _safe_float(row.get("actual_mae_pct"), 0.0),
    }


def _draw_trade_move_arrow(
    ax: Any,
    start_time: Any,
    start_price: float,
    end_time: Any,
    end_price: float,
    *,
    linewidth: float,
    alpha: float,
    color: str = TP_MOVE_COLOR,
) -> None:
    ax.annotate(
        "",
        xy=(end_time, end_price),
        xytext=(start_time, start_price),
        arrowprops={
            "arrowstyle": "->",
            "color": color,
            "linewidth": linewidth,
            "alpha": alpha,
            "shrinkA": 0,
            "shrinkB": 0,
            "mutation_scale": 12,
        },
        zorder=11,
    )


def _fmt_pct_value(value: Any) -> str:
    if value is None or value == "":
        return ""
    number = _safe_float(value, 0.0)
    if abs(number - round(number)) < 1e-9:
        return f"{number:.0f}"
    return f"{number:.1f}"


def _median(values: list[float]) -> float:
    clean = [float(value) for value in values if not math.isnan(float(value)) and not math.isinf(float(value))]
    return round(float(np.median(clean)), 6) if clean else 0.0


def _percentile(values: list[float], pct: float) -> float:
    clean = [float(value) for value in values if not math.isnan(float(value)) and not math.isinf(float(value))]
    return round(float(np.percentile(clean, pct)), 6) if clean else 0.0


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def _read_csv_rows_optional(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    return _read_csv_rows(path)


def _load_stas2_context_tables(stas2_run_dir: Path) -> dict[str, list[dict[str, Any]]]:
    return {
        "hourly": _read_csv_rows_optional(stas2_run_dir / "STAS2_HOURLY_PHASES.csv"),
        "macro": _read_csv_rows_optional(stas2_run_dir / "STAS2_MACRO_WAVES.csv"),
        "continuous": _read_csv_rows_optional(stas2_run_dir / "STAS2_CONTINUOUS_WAVES.csv"),
    }


def _ts_in_interval(ts: pd.Timestamp, start_raw: Any, end_raw: Any) -> bool:
    start = _parse_utc_ts(start_raw)
    end = _parse_utc_ts(end_raw)
    if start is None or end is None:
        return False
    return start <= ts < end


def _find_hour_context(entry_time: pd.Timestamp, hourly_rows: list[dict[str, Any]]) -> dict[str, Any]:
    hour_start = entry_time.floor("h")
    hour_key = _fmt_ts(hour_start)
    for row in hourly_rows:
        if row.get("hour_start_utc") == hour_key:
            return row
    return {}


def _find_macro_context(entry_time: pd.Timestamp, macro_rows: list[dict[str, Any]]) -> dict[str, Any]:
    for row in macro_rows:
        if _ts_in_interval(entry_time, row.get("macro_wave_start_time_utc"), row.get("macro_wave_end_time_utc")):
            return row
    return {}


def _find_continuous_context(entry_time: pd.Timestamp, continuous_rows: list[dict[str, Any]]) -> dict[str, Any]:
    for row in continuous_rows:
        if _ts_in_interval(entry_time, row.get("continuous_wave_start_time_utc"), row.get("continuous_wave_end_time_utc")):
            return row
    return {}


def _pick(row: dict[str, Any], key: str, default: Any = "") -> Any:
    value = row.get(key, default)
    return default if value is None else value


def _stas2_context_payload(
    source_row: dict[str, Any],
    entry_time: pd.Timestamp,
    context_tables: dict[str, list[dict[str, Any]]] | None,
) -> dict[str, Any]:
    context_tables = context_tables or {}
    hour = _find_hour_context(entry_time, context_tables.get("hourly", []))
    macro = _find_macro_context(entry_time, context_tables.get("macro", []))
    continuous = _find_continuous_context(entry_time, context_tables.get("continuous", []))

    payload: dict[str, Any] = {
        "context_scope": "PRE_ENTRY_CAUSAL_PLUS_REVIEW_CONTEXT",
        "direction_scope": "LONG_ONLY",
        "long_only_flag": True,
        "short_context_only_flag": True,
        "wave_context_scope": "HINDSIGHT_WAVE_REVIEW" if macro or continuous else "NONE",
        "volume_context_status": "PARTIAL",
        "volume_ratio20": _pick(source_row, "stas1_feature_volume_ratio20"),
    }

    for key in [
        "session_time_bucket_code",
        "session_time_bucket_label",
        "effective_session_code",
        "effective_session_label",
        "real_tradfi_session_open",
        "day_type",
        "pre_5m_range_pct",
        "pre_15m_range_pct",
        "pre_30m_range_pct",
        "pre_60m_range_pct",
        "pre_5m_path_pct",
        "pre_15m_path_pct",
        "pre_30m_path_pct",
        "pre_60m_path_pct",
        "pre_5m_long_wave_phase",
        "pre_15m_long_wave_phase",
        "pre_30m_long_wave_phase",
        "pre_60m_long_wave_phase",
        "pre_5m_long_wave_up_from_low_pct",
        "pre_15m_long_wave_up_from_low_pct",
        "pre_30m_long_wave_up_from_low_pct",
        "pre_60m_long_wave_up_from_low_pct",
        "pre_15m_background_phase",
        "pre_30m_background_phase",
        "pre_60m_background_phase",
        "pre_15m_background_phase_rank",
        "pre_30m_background_phase_rank",
        "pre_60m_background_phase_rank",
    ]:
        payload[key] = _pick(source_row, key)

    for key in [
        "hour_background_phase",
        "hour_background_phase_rank",
        "hour_range_pct",
        "hour_path_pct",
        "hour_close_move_pct",
        "hour_long_wave_phase",
        "hour_long_wave_rank",
        "hour_long_wave_up_from_low_pct",
        "hour_long_wave_pullback_from_post_low_high_pct",
        "hour_short_wave_phase",
        "hour_short_wave_rank",
        "hour_short_wave_down_from_high_pct",
        "hour_direction_bias",
    ]:
        payload[key] = _pick(hour, key)

    for key in [
        "macro_wave_id",
        "macro_wave_no",
        "macro_wave_segment_kind",
        "macro_wave_direction",
        "macro_wave_status",
        "macro_wave_start_time_utc",
        "macro_wave_end_time_utc",
        "macro_wave_visible_move_pct",
        "macro_wave_full_move_pct",
        "macro_wave_visible_range_pct",
        "macro_wave_full_duration_min",
        "macro_wave_carry_from_prev_day",
        "macro_wave_carry_to_next_day",
        "continuous_wave_id",
    ]:
        payload[key] = _pick(macro, key)

    for key in [
        "continuous_wave_segment_kind",
        "continuous_wave_direction",
        "continuous_wave_status",
        "continuous_wave_start_time_utc",
        "continuous_wave_end_time_utc",
        "continuous_wave_move_pct",
        "continuous_wave_range_pct",
        "continuous_wave_duration_min",
        "continuous_wave_boundary",
    ]:
        payload[key] = _pick(continuous, key)

    if not payload.get("continuous_wave_id") and continuous:
        payload["continuous_wave_id"] = _pick(continuous, "continuous_wave_id")
    return payload


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(rows)


def _write_csv_columns(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows([{col: row.get(col, "") for col in columns} for row in rows])


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _load_days(root: Path, start_day: str, end_day: str, hold_hours: float, symbol: str, timeframe: str) -> tuple[pd.DataFrame, list[str]]:
    tail_days = max(0, int(math.ceil(max(0.0, hold_hours) / 24.0)))
    days = _iter_days(start_day, (pd.Timestamp(end_day) + pd.Timedelta(days=tail_days)).strftime("%Y-%m-%d"))
    frames: list[pd.DataFrame] = []
    missing: list[str] = []
    for day in days:
        source = _source_csv(root, day, timeframe, symbol)
        if not source.exists():
            missing.append(_rel(root, source))
            continue
        frame = _load_ohlcv(source)
        frame["source_day_utc"] = day
        frames.append(frame)
    if not frames:
        raise FileNotFoundError("no OHLCV sources found for requested period")
    out = pd.concat(frames, ignore_index=True).sort_values("open_time_utc").reset_index(drop=True)
    return out, missing


def _latest_stas2_run(root: Path) -> Path:
    runs_dir = root / "STAS2_MARKET_PHASE_REVIEW" / "runs"
    candidates = [p for p in runs_dir.glob("*") if p.is_dir() and (p / "STAS2_RECORDS.csv").exists()]
    if not candidates:
        raise FileNotFoundError("STAS2_RECORDS.csv not found under STAS2_MARKET_PHASE_REVIEW/runs")
    return sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def _hit_payload(future: pd.DataFrame, entry_time: pd.Timestamp, entry_price: float) -> tuple[dict[str, Any], float, str, str]:
    payload: dict[str, Any] = {}
    hit_pcts: list[float] = []
    first_hit_pct = 0.0
    first_hit_time = ""
    for pct in PERCENT_LADDER:
        key = _phase_key(pct)
        target_price = entry_price * (1.0 + pct / 100.0)
        hit_rows = future[future["high"] >= target_price]
        hit_time = pd.Timestamp(hit_rows.iloc[0]["open_time_utc"]) if not hit_rows.empty else None
        hit = hit_time is not None
        if hit:
            hit_pcts.append(pct)
            if first_hit_pct <= 0:
                first_hit_pct = pct
                first_hit_time = _fmt_ts(hit_time)
            before_hit = future[future["open_time_utc"] <= hit_time]
            mae_before_hit = float(before_hit["low"].min()) if not before_hit.empty else entry_price
            mae_before_hit_pct: float | str = round((mae_before_hit / entry_price - 1.0) * 100.0, 6)
        else:
            mae_before_hit_pct = ""
        payload[f"target_{key}pct_price"] = round(target_price, 8)
        payload[f"hit_{key}pct"] = bool(hit)
        payload[f"hit_time_{key}pct_utc"] = _fmt_ts(hit_time)
        payload[f"time_to_{key}pct_min"] = round((hit_time - entry_time).total_seconds() / 60.0, 3) if hit_time is not None else ""
        payload[f"mae_before_{key}pct_pct"] = mae_before_hit_pct
    max_hit = max(hit_pcts) if hit_pcts else 0.0
    return payload, max_hit, (f"{first_hit_pct:.1f}%" if first_hit_pct > 0 else ""), first_hit_time


def _window_payload(future: pd.DataFrame, entry_time: pd.Timestamp, entry_price: float) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for minutes in POST_WINDOWS_MIN:
        end = entry_time + pd.Timedelta(minutes=minutes)
        win = future[(future["open_time_utc"] >= entry_time) & (future["open_time_utc"] < end)]
        prefix = f"post_{minutes}m"
        if win.empty:
            payload.update(
                {
                    f"{prefix}_rows": 0,
                    f"{prefix}_mfe_pct": "",
                    f"{prefix}_mae_pct": "",
                    f"{prefix}_mfe_time_utc": "",
                    f"{prefix}_mae_time_utc": "",
                    f"{prefix}_close_move_pct": "",
                    f"{prefix}_range_pct": "",
                }
            )
            continue
        max_high = float(win["high"].max())
        min_low = float(win["low"].min())
        last_close = float(win.iloc[-1]["close"])
        high_row = win.loc[win["high"].idxmax()]
        low_row = win.loc[win["low"].idxmin()]
        payload.update(
            {
                f"{prefix}_rows": int(len(win)),
                f"{prefix}_mfe_pct": round((max_high / entry_price - 1.0) * 100.0, 6),
                f"{prefix}_mae_pct": round((min_low / entry_price - 1.0) * 100.0, 6),
                f"{prefix}_mfe_time_utc": _fmt_ts(pd.Timestamp(high_row["open_time_utc"])),
                f"{prefix}_mae_time_utc": _fmt_ts(pd.Timestamp(low_row["open_time_utc"])),
                f"{prefix}_close_move_pct": round((last_close / entry_price - 1.0) * 100.0, 6),
                f"{prefix}_range_pct": round((max_high / min_low - 1.0) * 100.0 if min_low > 0 else 0.0, 6),
            }
        )
    return payload


def _trade_path_payload(future: pd.DataFrame, entry_price: float, first_hit_time: str) -> dict[str, Any]:
    if future.empty or entry_price <= 0:
        return {
            "hold_window_rows": 0,
            "mfe_hold_pct": "",
            "mae_hold_pct": "",
            "actual_mfe_pct": "",
            "actual_mae_pct": "",
            "actual_mfe_time_utc": "",
            "actual_mae_time_utc": "",
            "actual_end_time_utc": "",
            "max_high_hold": "",
            "min_low_hold": "",
            "close_move_hold_pct": "",
            "drawdown_before_first_hit_pct": "",
            "first_post_entry_event": "",
        }
    max_high = float(future["high"].max())
    min_low = float(future["low"].min())
    last_close = float(future.iloc[-1]["close"])
    high_row = future.loc[future["high"].idxmax()]
    low_row = future.loc[future["low"].idxmin()]
    end_time = pd.Timestamp(future.iloc[-1]["open_time_utc"])
    first_hit_ts = pd.Timestamp(first_hit_time).tz_convert("UTC") if first_hit_time else None
    before_first_hit = future[future["open_time_utc"] <= first_hit_ts] if first_hit_ts is not None else future
    drawdown_before_first_hit = float(before_first_hit["low"].min()) if not before_first_hit.empty else entry_price
    drawdown_rows = future[future["low"] < entry_price]
    first_drawdown_time = pd.Timestamp(drawdown_rows.iloc[0]["open_time_utc"]) if not drawdown_rows.empty else None
    if first_hit_ts is not None and first_drawdown_time is not None:
        first_event = "DRAWDOWN_BEFORE_FIRST_HIT" if first_drawdown_time < first_hit_ts else "LADDER_HIT_BEFORE_DRAWDOWN"
    elif first_hit_ts is not None:
        first_event = "LADDER_HIT_NO_DRAWDOWN_FIRST"
    elif first_drawdown_time is not None:
        first_event = "DRAWDOWN_NO_LADDER_HIT"
    else:
        first_event = "NO_DRAWDOWN_NO_LADDER_HIT"
    return {
        "hold_window_rows": int(len(future)),
        "actual_move_pct": round((max_high / entry_price - 1.0) * 100.0, 6),
        "actual_mfe_pct": round((max_high / entry_price - 1.0) * 100.0, 6),
        "actual_mae_pct": round((min_low / entry_price - 1.0) * 100.0, 6),
        "actual_mfe_time_utc": _fmt_ts(pd.Timestamp(high_row["open_time_utc"])),
        "actual_mae_time_utc": _fmt_ts(pd.Timestamp(low_row["open_time_utc"])),
        "actual_end_time_utc": _fmt_ts(end_time),
        "mfe_hold_pct": round((max_high / entry_price - 1.0) * 100.0, 6),
        "mae_hold_pct": round((min_low / entry_price - 1.0) * 100.0, 6),
        "max_high_hold": round(max_high, 8),
        "min_low_hold": round(min_low, 8),
        "close_move_hold_pct": round((last_close / entry_price - 1.0) * 100.0, 6),
        "drawdown_before_first_hit_pct": round((drawdown_before_first_hit / entry_price - 1.0) * 100.0, 6),
        "first_post_entry_event": first_event,
    }


def _low_signal_to_mfe_payload(
    future: pd.DataFrame,
    entry_time: pd.Timestamp,
    entry_price: float,
    signal_time_raw: Any,
    signal_price: float,
) -> dict[str, Any]:
    if future.empty or entry_price <= 0:
        return {
            "signal_to_entry_move_pct": "",
            "mfe_max_price": "",
            "mfe_max_time_utc": "",
            "mfe_from_signal_pct": "",
            "mfe_from_entry_pct": "",
            "signal_to_mfe_min": "",
            "entry_to_mfe_min": "",
            "mae_min_price": "",
            "mae_min_time_utc": "",
            "mae_from_entry_pct": "",
            "entry_to_mae_min": "",
            "mae_before_mfe_price": "",
            "mae_before_mfe_time_utc": "",
            "mae_before_mfe_pct": "",
        }

    signal_time = _parse_utc_ts(signal_time_raw) or entry_time
    signal_base = signal_price if signal_price > 0 else entry_price
    high_row = future.loc[future["high"].idxmax()]
    low_row = future.loc[future["low"].idxmin()]
    mfe_price = float(high_row["high"])
    mfe_time = pd.Timestamp(high_row["open_time_utc"])
    mae_price = float(low_row["low"])
    mae_time = pd.Timestamp(low_row["open_time_utc"])
    before_mfe = future[future["open_time_utc"] <= mfe_time]
    if before_mfe.empty:
        mae_before_mfe_price = entry_price
        mae_before_mfe_time = entry_time
    else:
        mae_before_mfe_row = before_mfe.loc[before_mfe["low"].idxmin()]
        mae_before_mfe_price = float(mae_before_mfe_row["low"])
        mae_before_mfe_time = pd.Timestamp(mae_before_mfe_row["open_time_utc"])

    return {
        "signal_to_entry_move_pct": round((entry_price / signal_base - 1.0) * 100.0, 6) if signal_base > 0 else "",
        "mfe_max_price": round(mfe_price, 8),
        "mfe_max_time_utc": _fmt_ts(mfe_time),
        "mfe_from_signal_pct": round((mfe_price / signal_base - 1.0) * 100.0, 6) if signal_base > 0 else "",
        "mfe_from_entry_pct": round((mfe_price / entry_price - 1.0) * 100.0, 6),
        "signal_to_mfe_min": _minutes_between(signal_time, mfe_time),
        "entry_to_mfe_min": _minutes_between(entry_time, mfe_time),
        "mae_min_price": round(mae_price, 8),
        "mae_min_time_utc": _fmt_ts(mae_time),
        "mae_from_entry_pct": round((mae_price / entry_price - 1.0) * 100.0, 6),
        "entry_to_mae_min": _minutes_between(entry_time, mae_time),
        "mae_before_mfe_price": round(mae_before_mfe_price, 8),
        "mae_before_mfe_time_utc": _fmt_ts(mae_before_mfe_time),
        "mae_before_mfe_pct": round((mae_before_mfe_price / entry_price - 1.0) * 100.0, 6),
    }


def _entry_to_tp_path_payload(future: pd.DataFrame, entry_price: float, hit_payload: dict[str, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    if future.empty or entry_price <= 0:
        for pct in PATH_CHECKPOINTS:
            key = _phase_key(pct)
            payload.update(
                {
                    f"mae_before_{key}pct_pct": "",
                    f"mfe_after_{key}pct_pct": "",
                    f"extra_after_{key}pct_pct": "",
                }
            )
        payload.update(
            {
                "after_mfe_min_pct": "",
                "giveback_after_mfe_pct": "",
                "giveback_after_mfe_to_close_pct": "",
                "giveback_after_mfe_to_min_pct": "",
            }
        )
        return payload

    high_row = future.loc[future["high"].idxmax()]
    mfe_time = pd.Timestamp(high_row["open_time_utc"])
    mfe_pct = round((float(high_row["high"]) / entry_price - 1.0) * 100.0, 6)
    last_close = float(future.iloc[-1]["close"])
    close_move_pct = round((last_close / entry_price - 1.0) * 100.0, 6)
    after_mfe = future[future["open_time_utc"] >= mfe_time]
    after_mfe_min_pct = ""
    giveback_to_min = ""
    if not after_mfe.empty:
        after_mfe_min = float(after_mfe["low"].min())
        after_mfe_min_pct = round((after_mfe_min / entry_price - 1.0) * 100.0, 6)
        giveback_to_min = round(mfe_pct - after_mfe_min_pct, 6)

    for pct in PATH_CHECKPOINTS:
        key = _phase_key(pct)
        hit_time = _parse_utc_ts(hit_payload.get(f"hit_time_{key}pct_utc"))
        if hit_time is None:
            payload.update(
                {
                    f"mae_before_{key}pct_pct": "",
                    f"mfe_after_{key}pct_pct": "",
                    f"extra_after_{key}pct_pct": "",
                }
            )
            continue
        before_hit = future[future["open_time_utc"] <= hit_time]
        after_hit = future[future["open_time_utc"] >= hit_time]
        mae_before_hit = float(before_hit["low"].min()) if not before_hit.empty else entry_price
        mfe_after_hit = float(after_hit["high"].max()) if not after_hit.empty else entry_price
        mfe_after_pct = round((mfe_after_hit / entry_price - 1.0) * 100.0, 6)
        payload.update(
            {
                f"mae_before_{key}pct_pct": round((mae_before_hit / entry_price - 1.0) * 100.0, 6),
                f"mfe_after_{key}pct_pct": mfe_after_pct,
                f"extra_after_{key}pct_pct": round(mfe_after_pct - pct, 6),
            }
        )

    payload.update(
        {
            "after_mfe_min_pct": after_mfe_min_pct,
            "giveback_after_mfe_pct": round(mfe_pct - close_move_pct, 6),
            "giveback_after_mfe_to_close_pct": round(mfe_pct - close_move_pct, 6),
            "giveback_after_mfe_to_min_pct": giveback_to_min,
        }
    )
    return payload


def _tp_profile_label(row: dict[str, Any]) -> str:
    background = str(row.get("phase_at_entry") or row.get("pre_15m_background_phase") or "")
    long_wave = str(row.get("long_wave_at_entry") or row.get("pre_15m_long_wave_phase") or "")
    setup = str(row.get("setup_at_entry") or row.get("entry_setup_quality_label") or "")
    return f"{background} | {long_wave} | {setup}"


def _stas3_verdict(row: dict[str, Any]) -> tuple[str, str]:
    max_hit = _safe_float(row.get("max_ladder_hit_pct"), 0.0)
    mae_15 = _safe_float(row.get("post_15m_mae_pct"), 0.0)
    t1 = row.get("time_to_1p0pct_min")
    t1_min = _safe_float(t1, -1.0) if t1 != "" else -1.0
    setup_rank = int(_safe_float(row.get("entry_setup_quality_rank"), -1.0))
    if max_hit >= 1.0 and 0 <= t1_min <= 120 and mae_15 >= -0.35 and setup_rank >= 2:
        return "FAST_CLEAN_1PCT", "1% быстро, ранняя просадка умеренная, setup не noise."
    if max_hit >= 1.0 and t1_min > 360:
        return "LATE_PUMP_DEPENDENT", "1% был достигнут поздно; точка зависит от дальнего пампа/удержания."
    if max_hit >= 1.0:
        return "WORKING_BUT_NEEDS_REVIEW", "1% достигнут, но скорость/просадка/setup требуют ручной проверки."
    if max_hit >= 0.5:
        return "PARTIAL_ONLY", "Дошло только до промежуточной фазы, полного 1% нет."
    if mae_15 <= -0.60:
        return "EARLY_DRAWDOWN_NO_ROOM", "В первые 15 минут сильная просадка и нет нормальной лестницы."
    return "NO_LADDER_HIT", "Процентная лестница почти не сработала в окне удержания."


def _tp_mismatch(row: dict[str, Any]) -> tuple[str, str]:
    setup_rank = int(_safe_float(row.get("entry_setup_quality_rank"), -1.0))
    max_hit = _safe_float(row.get("max_ladder_hit_pct"), 0.0)
    actual_move = _safe_float(row.get("actual_move_pct"), max_hit)
    rec_fast = _safe_float(row.get("reasonable_tp_fast_pct"), 0.0)
    rec_hold = _safe_float(row.get("reasonable_tp_hold_pct"), 0.0)
    t1 = row.get("time_to_1p0pct_min")
    t1_min = _safe_float(t1, -1.0) if t1 != "" else -1.0
    hit_1pct = bool(row.get("hit_1p0pct"))

    if setup_rank <= 0:
        if hit_1pct:
            return "NOISE_ENTRY_BUT_FUTURE_HIT", "Setup уже помечен как noise; будущий hit не делает вход хорошим."
        return "NOISE_ENTRY_NO_ROOM", "Setup noise и движение после входа не подтвердило нормальную лестницу."
    if hit_1pct and t1_min > 360:
        return "GOOD_OR_OK_ENTRY_BUT_1PCT_LATE", "1% достигнут слишком поздно; для базового TP цель завышена."
    if not hit_1pct and actual_move >= 0.3 and (rec_fast > 0 or rec_hold > 0) and rec_hold < 1.0:
        return "GOOD_OR_OK_ENTRY_WRONG_1PCT_TP", "Вход мог быть рабочим, но реалистичная цель ниже 1%."
    if rec_fast > 0 and rec_fast < 1.0 and hit_1pct:
        return "HIT_1PCT_BUT_PHASE_TP_LOWER", "1% случился, но фазовая fast-статистика предлагает меньший TP."
    if rec_fast >= 1.0 and hit_1pct and 0 <= t1_min <= 360:
        return "TP_ALIGNED_WITH_PHASE", "1% согласуется с текущей фазовой лестницей."
    if max_hit < 0.3:
        return "NO_TP_ROOM", "Даже минимальная рабочая фаза 0.3% не подтверждена."
    return "REVIEW_MANUALLY", "Нужна ручная проверка: движение есть, но автоматический вывод неоднозначен."


def _exit_review_bucket(row: dict[str, Any]) -> tuple[str, str]:
    hit_0p3 = bool(row.get("hit_0p3pct"))
    actual_mfe = _safe_float(row.get("actual_mfe_pct"), 0.0)
    entry_to_mfe = _safe_float(row.get("entry_to_mfe_min"), -1.0) if row.get("entry_to_mfe_min") != "" else -1.0
    mae_before_mfe = _safe_float(row.get("mae_before_mfe_pct"), 0.0)
    time_to_0p3 = _safe_float(row.get("time_to_0p3pct_min"), -1.0) if row.get("time_to_0p3pct_min") != "" else -1.0
    time_to_1p0 = _safe_float(row.get("time_to_1p0pct_min"), -1.0) if row.get("time_to_1p0pct_min") != "" else -1.0
    extra_after_0p3 = _safe_float(row.get("extra_after_0p3pct_pct"), 0.0)
    extra_after_1p0 = _safe_float(row.get("extra_after_1p0pct_pct"), 0.0)
    mae_before_1p0 = _safe_float(row.get("mae_before_1p0pct_pct"), 0.0)

    if not hit_0p3:
        return "NO_0P3_ROOM_REVIEW", "Review-only: даже 0.3% не подтвердился в окне удержания."
    if actual_mfe >= 2.0 and 0 <= time_to_1p0 <= 360 and extra_after_1p0 >= 1.0 and mae_before_1p0 >= -0.60:
        return "EARLY_1PCT_TRAIL_REVIEW", "Review-only: 1% пришел не позднее 6 часов, после него был запас MFE; кандидат на изучение трейла/частичного выхода."
    if actual_mfe >= 1.0 and mae_before_mfe < -0.60:
        return "BIG_MFE_BUT_DEEP_MAE_REVIEW", "Review-only: MFE большой, но до максимума была глубокая просадка."
    if actual_mfe >= 1.0 and entry_to_mfe > 360:
        return "LATE_MFE_PUMP_REVIEW", "Review-only: большой MFE появился поздно; это может быть поздний памп, а не качество входа."
    if actual_mfe >= 1.0 and extra_after_0p3 >= 0.7 and mae_before_mfe >= -0.35 and 0 <= entry_to_mfe <= 360:
        return "CLEAN_MFE_RUNNER_REVIEW", "Review-only: после 0.3% рынок дал продолжение, просадка до MFE умеренная, максимум не слишком поздний."
    if 0 <= time_to_0p3 <= 60 and actual_mfe < 0.5 and extra_after_0p3 < 0.2:
        return "FAST_0P3_ONLY_REVIEW", "Review-only: 0.3% был быстро, но продолжения почти не было."
    return "MIXED_EXIT_REVIEW", "Review-only: неоднозначный путь сделки, нужен визуальный разбор."


def _annotate_big_move_review(rows: list[dict[str, Any]]) -> None:
    for row in rows:
        reasonable_tp = _safe_float(row.get("reasonable_tp_pct"), 0.0)
        actual_mfe = _safe_float(row.get("actual_mfe_pct"), 0.0)
        row["tp_to_mfe_extra_pct"] = round(actual_mfe - reasonable_tp, 6) if reasonable_tp > 0 else ""
        bucket, reason = _exit_review_bucket(row)
        row["exit_review_bucket"] = bucket
        row["exit_review_reason"] = reason


def _ladder_stats_for_items(items: list[dict[str, Any]], *, fast_minutes: int) -> dict[str, Any]:
    out: dict[str, Any] = {}
    count = len(items)
    for pct in PERCENT_LADDER:
        key = _phase_key(pct)
        hit_items = [item for item in items if bool(item.get(f"hit_{key}pct"))]
        hit_times = [_safe_float(item.get(f"time_to_{key}pct_min"), 0.0) for item in hit_items if item.get(f"time_to_{key}pct_min") != ""]
        fast_items = [item for item in hit_items if _safe_float(item.get(f"time_to_{key}pct_min"), fast_minutes + 1.0) <= fast_minutes]
        out[f"hit_{key}pct_count"] = len(hit_items)
        out[f"hit_{key}pct_rate"] = round(len(hit_items) / count, 6) if count else 0.0
        out[f"fast_hit_{key}pct_count"] = len(fast_items)
        out[f"fast_hit_{key}pct_rate"] = round(len(fast_items) / count, 6) if count else 0.0
        out[f"median_time_to_{key}pct_min"] = _median(hit_times) if hit_times else ""
    return out


def _tp_band_label(tp_fast: float, tp_hold: float, data_status: str) -> str:
    if data_status != "OK":
        return "INSUFFICIENT_DATA_REVIEW_ONLY"
    basis = tp_fast if tp_fast > 0 else tp_hold
    if basis <= 0:
        return "NO_ENTRY_OR_OBSERVE"
    if basis <= 0.3:
        return "MINIMAL_0P3"
    if basis <= 0.6:
        return "SMALL_0P3_0P6"
    if basis <= 1.0:
        return "NORMAL_0P7_1P0"
    return "EXTENDED_ABOVE_1P0"


def _move_scale_bucket(max_hit: float) -> str:
    if max_hit <= 0:
        return "NO_LADDER_HIT"
    if max_hit < 1.0:
        return "FAST_TP"
    if max_hit <= 2.0:
        return "MEDIUM_MOVE"
    if max_hit <= 20.0:
        return "MEDIUM_EXTENSION_REVIEW"
    return "ABOVE_GRID_DIAGNOSTIC"


def _ideal_review_tp_fields(row: dict[str, Any], hold_hours: float) -> dict[str, Any]:
    max_feasible = _safe_float(row.get("max_ladder_hit_pct"), 0.0)
    if max_feasible <= 0:
        return {
            "max_feasible_review_tp_pct": "",
            "ideal_review_tp_pct": "",
            "ideal_review_tp_reason": "Нет достигнутого уровня рабочей V2-лестницы.",
            "ideal_review_tp_warning": "NO_LADDER_HIT_REVIEW_ONLY",
            "move_scale_bucket": "NO_LADDER_HIT",
        }

    candidates: list[float] = []
    warnings: list[str] = []
    for pct in PERCENT_LADDER:
        key = _phase_key(pct)
        if not bool(row.get(f"hit_{key}pct")):
            continue
        time_to = _safe_float(row.get(f"time_to_{key}pct_min"), hold_hours * 60.0 + 1.0)
        mae_before = _safe_float(row.get(f"mae_before_{key}pct_pct"), 0.0)
        if pct < 1.0:
            ok = time_to <= 120 and mae_before >= -0.60
        elif pct <= 2.0:
            ok = time_to <= 720 and mae_before >= -1.00
        else:
            ok = time_to <= hold_hours * 60.0 and mae_before >= -1.50
        if ok:
            candidates.append(pct)

    if not candidates:
        first_hit = _safe_float(str(row.get("first_ladder_hit") or "0").replace("%", ""), 0.0)
        ideal = first_hit if first_hit > 0 else max_feasible
        warnings.append("MAX_TP_REQUIRES_DEEP_MAE_OR_LATE_MOVE")
        reason = "Review-only: максимальный достигнутый TP есть, но путь до него слишком поздний или с просадкой; выбран первый достижимый уровень для разбора."
    else:
        ideal = max(candidates)
        reason = "Review-only: выбран самый высокий достигнутый уровень V2-лестницы, который прошел фильтр времени и MAE."

    max_time = _safe_float(row.get(f"time_to_{_phase_key(max_feasible)}pct_min"), 0.0)
    max_mae = _safe_float(row.get(f"mae_before_{_phase_key(max_feasible)}pct_pct"), 0.0)
    if max_time > 720 and max_feasible >= 1.0:
        warnings.append("LATE_TARGET")
    if max_mae < -1.0:
        warnings.append("DEEP_MAE_BEFORE_TARGET")
    if _safe_float(row.get("hour_short_wave_down_from_high_pct"), 0.0) >= _safe_float(row.get("hour_long_wave_up_from_low_pct"), 0.0) and _safe_float(row.get("hour_short_wave_down_from_high_pct"), 0.0) >= 1.0:
        warnings.append("STRONG_SHORT_RISK_CONTEXT")
    if str(row.get("macro_wave_direction") or "") == "SHORT":
        warnings.append("HINDSIGHT_WAVE_SHORT_CONTEXT")

    return {
        "max_feasible_review_tp_pct": max_feasible,
        "ideal_review_tp_pct": ideal,
        "ideal_review_tp_reason": reason,
        "ideal_review_tp_warning": "|".join(dict.fromkeys(warnings)),
        "move_scale_bucket": _move_scale_bucket(max_feasible),
    }


def _tp_reasonable_fields(
    items: list[dict[str, Any]],
    *,
    min_samples: int,
    fast_minutes: int,
    hit_rate_min: float,
    fast_hit_rate_min: float,
) -> dict[str, Any]:
    count = len(items)
    ladder_stats = _ladder_stats_for_items(items, fast_minutes=fast_minutes)
    actual_moves = [_safe_float(item.get("actual_move_pct"), 0.0) for item in items]
    mae_15 = [_safe_float(item.get("post_15m_mae_pct"), 0.0) for item in items]
    fast_candidates: list[float] = []
    hold_candidates: list[float] = []
    for pct in PERCENT_LADDER:
        key = _phase_key(pct)
        if _safe_float(ladder_stats.get(f"fast_hit_{key}pct_rate"), 0.0) >= fast_hit_rate_min:
            fast_candidates.append(pct)
        if _safe_float(ladder_stats.get(f"hit_{key}pct_rate"), 0.0) >= hit_rate_min:
            hold_candidates.append(pct)
    data_status = "OK" if count >= min_samples else "INSUFFICIENT_DATA"
    tp_fast = max(fast_candidates) if fast_candidates and data_status == "OK" else 0.0
    tp_hold = max(hold_candidates) if hold_candidates and data_status == "OK" else 0.0
    if data_status == "INSUFFICIENT_DATA":
        decision = "Наблюдение: мало строк для фазовой границы."
    elif tp_fast > 0:
        decision = f"Базовый TP по быстрым достижениям: около {_fmt_pct_value(tp_fast)}%."
    elif tp_hold > 0:
        decision = f"Только extended TP: до {_fmt_pct_value(tp_hold)}%, но без fast-подтверждения."
    else:
        decision = "Входы запрещать или оставлять как наблюдение: даже нижняя лестница не подтверждена."
    return {
        "data_status": data_status,
        "reasonable_tp_fast_pct": tp_fast if tp_fast > 0 else "",
        "reasonable_tp_hold_pct": tp_hold if tp_hold > 0 else "",
        "reasonable_tp_band": _tp_band_label(tp_fast, tp_hold, data_status),
        "tp_ladder_v1_decision": decision,
        "median_actual_move_pct": _median(actual_moves),
        "p25_actual_move_pct": _percentile(actual_moves, 25),
        "p75_actual_move_pct": _percentile(actual_moves, 75),
        "median_mae_15m_pct": _median(mae_15),
        **ladder_stats,
    }


def _annotate_rows_with_tp_recommendations(
    rows: list[dict[str, Any]],
    *,
    min_samples: int,
    fast_minutes: int,
    hit_rate_min: float,
    fast_hit_rate_min: float,
) -> dict[str, dict[str, Any]]:
    by_profile: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        by_profile[str(row.get("tp_phase_profile") or "")].append(row)
    profile_recommendations: dict[str, dict[str, Any]] = {}
    for profile, items in by_profile.items():
        fields = _tp_reasonable_fields(
            items,
            min_samples=min_samples,
            fast_minutes=fast_minutes,
            hit_rate_min=hit_rate_min,
            fast_hit_rate_min=fast_hit_rate_min,
        )
        fields["tp_phase_profile"] = profile
        fields["count"] = len(items)
        profile_recommendations[profile] = fields
    for row in rows:
        fields = profile_recommendations.get(str(row.get("tp_phase_profile") or ""), {})
        for key in ["data_status", "reasonable_tp_fast_pct", "reasonable_tp_hold_pct", "reasonable_tp_band", "tp_ladder_v1_decision"]:
            row[key] = fields.get(key, "")
        tp_fast = _safe_float(row.get("reasonable_tp_fast_pct"), 0.0)
        tp_hold = _safe_float(row.get("reasonable_tp_hold_pct"), 0.0)
        reasonable_tp = tp_fast if tp_fast > 0 else tp_hold
        entry_price = _safe_float(row.get("entry_price_for_calc"), _safe_float(row.get("entry_price_5bps"), 0.0))
        row["reasonable_tp_pct"] = reasonable_tp if reasonable_tp > 0 else ""
        row["reasonable_tp_price"] = round(entry_price * (1.0 + reasonable_tp / 100.0), 8) if entry_price > 0 and reasonable_tp > 0 else ""
        row["reasonable_tp_source"] = (
            "phase_profile_fast_hit_rate"
            if tp_fast > 0
            else "phase_profile_hold_hit_rate"
            if tp_hold > 0
            else "insufficient_or_no_phase_edge_review_only"
        )
        row["reasonable_tp_reason"] = fields.get("tp_ladder_v1_decision", "")
        row["tp_ladder_v2_decision"] = fields.get("tp_ladder_v1_decision", "")
        row["phase_profile_sample_count"] = fields.get("count", "")
        mismatch, reason = _tp_mismatch(row)
        row["tp_vs_1pct_label"] = mismatch
        row["tp_vs_1pct_reason"] = reason
        row["wrong_tp_flag"] = mismatch in {"GOOD_OR_OK_ENTRY_WRONG_1PCT_TP", "GOOD_OR_OK_ENTRY_BUT_1PCT_LATE", "HIT_1PCT_BUT_PHASE_TP_LOWER"}
        row["noise_flag"] = bool(row.get("noise_flag")) or str(mismatch).startswith("NOISE_ENTRY")
    return profile_recommendations


def compute_stas3_rows(
    stas2_rows: list[dict[str, Any]],
    market_df: pd.DataFrame,
    hold_hours: float,
    skipped_rows: list[dict[str, Any]] | None = None,
    context_tables: dict[str, list[dict[str, Any]]] | None = None,
) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    hold_delta = pd.Timedelta(hours=hold_hours)
    for row in sorted(stas2_rows, key=lambda item: (str(item.get("entry_time_utc")), str(item.get("record_id")))):
        if not str(row.get("entry_price_5bps") or "").strip():
            if skipped_rows is not None:
                skipped_rows.append(
                    {
                        "stas2_record_id": row.get("record_id"),
                        "candidate_id": row.get("candidate_id"),
                        "day_utc": row.get("day_utc"),
                        "entry_time_utc": row.get("entry_time_utc"),
                        "skip_reason": "missing_entry_price_5bps",
                    }
                )
            continue
        entry_time = pd.Timestamp(row["entry_time_utc"])
        entry_time = entry_time.tz_convert("UTC") if entry_time.tzinfo is not None else entry_time.tz_localize("UTC")
        entry_open = _safe_float(row.get("entry_open_price"))
        entry_price = _safe_float(row.get("entry_price_5bps"), 0.0)
        future = market_df[(market_df["open_time_utc"] >= entry_time) & (market_df["open_time_utc"] <= entry_time + hold_delta)].reset_index(drop=True)
        if future.empty or entry_price <= 0:
            if skipped_rows is not None:
                skipped_rows.append(
                    {
                        "stas2_record_id": row.get("record_id"),
                        "candidate_id": row.get("candidate_id"),
                        "day_utc": row.get("day_utc"),
                        "entry_time_utc": row.get("entry_time_utc"),
                        "skip_reason": "NO_FUTURE_BARS_OR_BAD_ENTRY_PRICE_5BPS",
                    }
                )
            continue
        hit_payload, max_hit, first_hit_label, first_hit_time = _hit_payload(future, entry_time, entry_price)
        phase_at_entry = row.get("pre_15m_background_phase") or row.get("pre_30m_background_phase") or ""
        long_wave_at_entry = row.get("pre_15m_long_wave_phase") or row.get("pre_30m_long_wave_phase") or ""
        setup_at_entry = row.get("entry_setup_quality_label") or ""
        signal_price = _safe_float(row.get("anchor_low_price"), entry_price)
        if signal_price <= 0:
            signal_price = entry_price
        base = {
            "stas1_source_run": row.get("source_run"),
            "stas1_record_id": row.get("record_id"),
            "stas2_record_id": row.get("record_id"),
            "record_id": row.get("record_id"),
            "candidate_id": row.get("candidate_id"),
            "day_utc": row.get("day_utc"),
            "anchor_time_utc": row.get("anchor_time_utc"),
            "stas1_confirmation_time_utc": row.get("stas1_confirmation_time_utc"),
            "anchor_low_price": round(signal_price, 8),
            "entry_time_utc": _fmt_ts(entry_time),
            "entry_open_price": round(entry_open, 8),
            "entry_price_5bps": round(entry_price, 8),
            "entry_price_for_calc": round(entry_price, 8),
            "entry_price_source": "entry_price_5bps",
            "entry_price_locked_flag": True,
            "entry_price_missing_flag": False,
            "review_label": row.get("review_label"),
            "outcome_status_stas1": row.get("outcome_status"),
            "entry_setup_quality_rank": row.get("entry_setup_quality_rank"),
            "entry_setup_quality_label": row.get("entry_setup_quality_label"),
            "entry_setup_quality_reason": row.get("entry_setup_quality_reason"),
            "day_type": row.get("day_type"),
            "session_time_bucket_code": row.get("session_time_bucket_code"),
            "session_time_bucket_label": row.get("session_time_bucket_label"),
            "effective_session_code": row.get("effective_session_code"),
            "effective_session_label": row.get("effective_session_label"),
            "pre_15m_background_phase": row.get("pre_15m_background_phase"),
            "pre_15m_long_wave_phase": row.get("pre_15m_long_wave_phase"),
            "pre_30m_background_phase": row.get("pre_30m_background_phase"),
            "pre_30m_long_wave_phase": row.get("pre_30m_long_wave_phase"),
            "phase_at_entry": phase_at_entry,
            "long_wave_at_entry": long_wave_at_entry,
            "setup_at_entry": setup_at_entry,
            "tp_phase_profile": f"{phase_at_entry} | {long_wave_at_entry} | {setup_at_entry}",
            "context_before_entry_check": row.get("context_before_entry_check"),
            "max_ladder_hit_pct": round(max_hit, 3),
            "first_ladder_hit": first_hit_label,
            "first_ladder_hit_time_utc": first_hit_time,
            "hold_hours": hold_hours,
        }
        base.update(_stas2_context_payload(row, entry_time, context_tables))
        base.update(hit_payload)
        base.update(_window_payload(future, entry_time, entry_price))
        base.update(_trade_path_payload(future, entry_price, first_hit_time))
        base.update(_low_signal_to_mfe_payload(future, entry_time, entry_price, base.get("anchor_time_utc") or base.get("stas1_confirmation_time_utc"), signal_price))
        base.update(_entry_to_tp_path_payload(future, entry_price, hit_payload))
        base.update(_ideal_review_tp_fields(base, hold_hours))
        base["tp_reason"] = base.get("ideal_review_tp_reason", "")
        base["tp_warning"] = base.get("ideal_review_tp_warning", "")
        verdict, reason = _stas3_verdict(base)
        base["stas3_verdict"] = verdict
        base["stas3_verdict_reason"] = reason
        base["noise_flag"] = str(row.get("entry_setup_quality_label") or "").lower().startswith("noise") or int(_safe_float(row.get("entry_setup_quality_rank"), 0.0)) <= 0
        base["wrong_tp_flag"] = False
        base["stas3_boundary"] = "stas3_v2_long_only_post_entry_audit_no_ml_no_optuna_no_api"
        out.append(base)
    return out


def _summary(rows: list[dict[str, Any]], group_fields: list[str]) -> list[dict[str, Any]]:
    buckets: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        buckets[tuple(str(row.get(field, "")) for field in group_fields)].append(row)
    out: list[dict[str, Any]] = []
    for key, items in sorted(buckets.items()):
        count = len(items)
        hit_1pct = sum(1 for item in items if bool(item.get("hit_1p0pct")))
        verdicts = Counter(str(item.get("stas3_verdict", "")) for item in items)
        out.append(
            {
                **{field: key[idx] for idx, field in enumerate(group_fields)},
                "count": count,
                "hit_1pct_count": hit_1pct,
                "hit_1pct_rate": round(hit_1pct / count, 6) if count else 0.0,
                "avg_max_ladder_hit_pct": round(float(np.mean([_safe_float(item.get("max_ladder_hit_pct")) for item in items])), 6),
                "avg_mae_15m_pct": round(float(np.mean([_safe_float(item.get("post_15m_mae_pct")) for item in items])), 6),
                "verdict_counts": json.dumps(dict(verdicts), ensure_ascii=False, sort_keys=True),
            }
        )
    return out


def _tp_summary(
    rows: list[dict[str, Any]],
    group_fields: list[str],
    *,
    min_samples: int,
    fast_minutes: int,
    hit_rate_min: float,
    fast_hit_rate_min: float,
) -> list[dict[str, Any]]:
    buckets: dict[tuple[str, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        buckets[tuple(str(row.get(field, "")) for field in group_fields)].append(row)
    out: list[dict[str, Any]] = []
    for key, items in sorted(buckets.items()):
        fields = _tp_reasonable_fields(
            items,
            min_samples=min_samples,
            fast_minutes=fast_minutes,
            hit_rate_min=hit_rate_min,
            fast_hit_rate_min=fast_hit_rate_min,
        )
        out.append(
            {
                **{field: key[idx] for idx, field in enumerate(group_fields)},
                "count": len(items),
                **fields,
            }
        )
    return out


def _project_rows(rows: list[dict[str, Any]], columns: list[str]) -> list[dict[str, Any]]:
    return [{col: row.get(col, "") for col in columns} for row in rows]


def _write_xlsx(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    header_fill = PatternFill("solid", fgColor="263238")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(sheet_name[:31])
        columns: list[str] = []
        for row in rows:
            for key in row:
                if key not in columns:
                    columns.append(key)
        if not columns:
            columns = ["empty"]
        ws.append(columns)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in rows:
            ws.append([row.get(col, "") for col in columns])
        ws.freeze_panes = "A2"
        for idx, col in enumerate(columns, start=1):
            width = min(max(10, len(col) + 2), 44)
            ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)


def _label_color(row: dict[str, Any]) -> str:
    verdict = str(row.get("stas3_verdict") or "")
    if verdict == "FAST_CLEAN_1PCT":
        return "#00e676"
    if verdict in {"WORKING_BUT_NEEDS_REVIEW", "PARTIAL_ONLY"}:
        return "#ffb300"
    if verdict == "LATE_PUMP_DEPENDENT":
        return "#42a5f5"
    return "#ff5252"


def _render_day_overview(*, df: pd.DataFrame, rows: list[dict[str, Any]], day: str, symbol: str, timeframe: str, out_path: Path) -> None:
    fig, (ax_price, ax_vol) = plt.subplots(2, 1, figsize=(32, 13), sharex=True, gridspec_kw={"height_ratios": [4.8, 1.15]})
    fig.patch.set_facecolor("#101418")
    _style_axis(ax_price)
    _style_axis(ax_vol)
    _draw_candles(ax_price, df, timeframe, linewidth=0.30)
    colors = ["#26a69a" if c >= o else "#ef5350" for o, c in zip(df["open"], df["close"])]
    ax_vol.bar(df["open_time_utc"].dt.tz_convert(None), df["volume"], width=_bar_width_days(timeframe), color=colors, alpha=0.70)
    start = pd.Timestamp(day)
    end = start + pd.Timedelta(days=1)
    for row in rows:
        se = _visual_signal_entry_fields(row)
        entry_time = pd.Timestamp(se["entry_time_utc"]).tz_convert(None)
        entry_price = _safe_float(se["entry_price"])
        signal_time = pd.Timestamp(se["signal_time_utc"]).tz_convert(None) if se["signal_time_utc"] else entry_time
        signal_price = _safe_float(se["signal_price"], entry_price)
        color = _label_color(row)
        ax_price.axvline(entry_time, color=color, alpha=0.18, linewidth=0.9)
        ax_price.plot([signal_time, entry_time], [signal_price, entry_price], color=ENTRY_COLOR, alpha=0.55, linewidth=0.85, linestyle=":", zorder=6)
        ax_price.scatter([signal_time], [signal_price], marker="o", s=38, color=SIGNAL_COLOR, edgecolors="white", linewidths=0.35, zorder=8)
        ax_price.scatter([entry_time], [entry_price], marker="^", s=64, color=ENTRY_COLOR, edgecolors="white", linewidths=0.45, zorder=9)
        tp = _visual_tp_fields(row)
        if tp["target"] > 0:
            hit_time_raw = tp["hit_time_utc"]
            if hit_time_raw:
                hit_time = pd.Timestamp(hit_time_raw).tz_convert(None)
                ax_price.plot(
                    [entry_time, hit_time],
                    [tp["target"], tp["target"]],
                    color=tp["color"],
                    alpha=0.40,
                    linewidth=0.85,
                    linestyle=tp["linestyle"],
                    zorder=5,
                )
                _draw_trade_move_arrow(ax_price, entry_time, entry_price, hit_time, tp["target"], linewidth=1.15, alpha=0.78)
                ax_price.scatter(
                    [hit_time],
                    [tp["target"]],
                    marker="*",
                    s=92,
                    color=tp["color"],
                    edgecolors="#101418",
                    linewidths=0.45,
                    zorder=9,
                )
            else:
                ax_price.hlines(tp["target"], entry_time, entry_time + pd.Timedelta(minutes=45), color=tp["color"], alpha=0.45, linewidth=1.1, linestyle=tp["linestyle"])
        mm = _visual_mfe_mae_fields(row)
        mfe_ts = _parse_utc_ts(mm["mfe_time_utc"])
        if mfe_ts is not None:
            mfe_time = mfe_ts.tz_convert(None)
            if start <= mfe_time < end and mm["mfe_price"] > 0:
                ax_price.scatter([mfe_time], [mm["mfe_price"]], marker="D", s=42, color=MFE_COLOR, edgecolors="#101418", linewidths=0.35, zorder=8)
    ax_price.set_xlim(start.to_pydatetime(), end.to_pydatetime())
    ax_price.set_title(f"STAS3 {symbol} {timeframe} {day} | post-entry percent ladder / MFE / MAE | NO ML/OPTUNA/API", color="white", fontsize=15)
    ax_price.set_ylabel("Price", color="white")
    ax_vol.set_ylabel("Volume", color="white")
    ax_vol.xaxis.set_major_locator(mdates.HourLocator(byhour=range(0, 24, 2)))
    ax_vol.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
    fig.autofmt_xdate()
    fig.tight_layout()
    fig.savefig(out_path, dpi=155, facecolor=fig.get_facecolor())
    plt.close(fig)


def _row_text(row: dict[str, Any]) -> str:
    tp = _visual_tp_fields(row)
    se = _visual_signal_entry_fields(row)
    return "\n".join(
        [
            f"{row.get('candidate_id')} | {row.get('stas3_verdict')}",
            f"{row.get('effective_session_label')} | setup {row.get('entry_setup_quality_label')}",
            f"signal {se['signal_price']:.4f} -> entry {se['entry_price']:.4f}",
            f"max feasible {row.get('max_feasible_review_tp_pct')}% | ideal {row.get('ideal_review_tp_pct')}%",
            f"1% time {row.get('time_to_1p0pct_min')} | short-risk {row.get('hour_short_wave_down_from_high_pct')}%",
            f"{tp['source']} +{tp['pct']:g}% | time {tp['time_to_min']}",
            f"{row.get('tp_vs_1pct_label')}",
            f"exit review {row.get('exit_review_bucket')}",
            f"MFE diagnostic {row.get('actual_mfe_pct')}% at {row.get('entry_to_mfe_min')}m",
            f"5m MFE/MAE {row.get('post_5m_mfe_pct')} / {row.get('post_5m_mae_pct')}%",
            f"15m MFE/MAE {row.get('post_15m_mfe_pct')} / {row.get('post_15m_mae_pct')}%",
        ]
    )


def _render_entry_pages(
    *,
    market_df: pd.DataFrame,
    rows: list[dict[str, Any]],
    symbol: str,
    timeframe: str,
    out_dir: Path,
    post_minutes: int,
    render_limit: int,
) -> list[Path]:
    sorted_rows = sorted(rows, key=lambda row: (str(row["entry_time_utc"]), str(row.get("record_id"))))
    if render_limit > 0:
        sorted_rows = sorted_rows[:render_limit]
    paths: list[Path] = []
    per_page = 6
    for page_start in range(0, len(sorted_rows), per_page):
        page_rows = sorted_rows[page_start : page_start + per_page]
        fig, axes = plt.subplots(3, 2, figsize=(22, 22), sharex=False)
        fig.patch.set_facecolor("#101418")
        flat = list(axes.flatten())
        for ax, row in zip(flat, page_rows):
            _style_axis(ax)
            se = _visual_signal_entry_fields(row)
            entry_ts = pd.Timestamp(se["entry_time_utc"]).tz_convert("UTC")
            signal_ts = pd.Timestamp(se["signal_time_utc"]).tz_convert("UTC") if se["signal_time_utc"] else entry_ts
            start = min(signal_ts, entry_ts) - pd.Timedelta(minutes=15)
            end = entry_ts + pd.Timedelta(minutes=post_minutes)
            win = market_df[(market_df["open_time_utc"] >= start) & (market_df["open_time_utc"] <= end)].reset_index(drop=True)
            if win.empty:
                ax.axis("off")
                continue
            _draw_candles(ax, win, timeframe, linewidth=0.42)
            entry_naive = entry_ts.tz_convert(None)
            signal_naive = signal_ts.tz_convert(None)
            entry_price = _safe_float(se["entry_price"])
            signal_price = _safe_float(se["signal_price"], entry_price)
            color = _label_color(row)
            ax.axvline(signal_naive, color=SIGNAL_COLOR, linewidth=1.05, alpha=0.62, linestyle=":", zorder=6)
            ax.axvline(entry_naive, color=ENTRY_COLOR, linewidth=1.25, alpha=0.82, zorder=7)
            ax.plot([signal_naive, entry_naive], [signal_price, entry_price], color=ENTRY_COLOR, alpha=0.76, linewidth=1.35, linestyle=":", zorder=7)
            ax.scatter([signal_naive], [signal_price], marker="o", s=92, color=SIGNAL_COLOR, edgecolors="white", linewidths=0.65, zorder=9)
            ax.scatter([entry_naive], [entry_price], marker="^", s=135, color=ENTRY_COLOR, edgecolors="white", linewidths=0.75, zorder=10)
            ax.text(
                signal_naive,
                signal_price,
                f" SIGNAL {signal_price:.4f}",
                ha="right",
                va="top",
                color=SIGNAL_COLOR,
                fontsize=8,
                fontweight="bold",
                bbox={"facecolor": "#101418", "edgecolor": SIGNAL_COLOR, "alpha": 0.82, "boxstyle": "round,pad=0.16"},
            )
            ax.text(
                entry_naive,
                entry_price,
                f" ENTRY {entry_price:.4f}",
                ha="left",
                va="bottom",
                color=ENTRY_COLOR,
                fontsize=8,
                fontweight="bold",
                bbox={"facecolor": "#101418", "edgecolor": ENTRY_COLOR, "alpha": 0.82, "boxstyle": "round,pad=0.16"},
            )
            tp = _visual_tp_fields(row)
            for pct in [0.3, 0.5, 1.0, 2.0, 3.0, 7.0]:
                key = _phase_key(pct)
                target = _safe_float(row.get(f"target_{key}pct_price"))
                if target > 0:
                    ax.axhline(target, color="#cfd8dc", linewidth=0.75, alpha=0.28)
                    ax.text(0.992, target, f"+{pct:g}%", transform=ax.get_yaxis_transform(), ha="right", va="bottom", color="#cfd8dc", fontsize=7)
            if tp["target"] > 0:
                ax.axhline(tp["target"], color=tp["color"], linewidth=1.55, alpha=0.95, linestyle=tp["linestyle"], zorder=6)
                ax.text(
                    0.992,
                    tp["target"],
                    f"{tp['source']} +{tp['pct']:g}%",
                    transform=ax.get_yaxis_transform(),
                    ha="right",
                    va="top",
                    color=tp["color"],
                    fontsize=8,
                    fontweight="bold",
                    bbox={"facecolor": "#101418", "edgecolor": tp["color"], "alpha": 0.72, "boxstyle": "round,pad=0.16"},
                )
                if tp["hit_time_utc"]:
                    tp_hit = pd.Timestamp(tp["hit_time_utc"]).tz_convert("UTC")
                    tp_hit_naive = tp_hit.tz_convert(None)
                    ax.axvline(tp_hit_naive, color=tp["color"], linewidth=1.0, alpha=0.62, linestyle=tp["linestyle"], zorder=6)
                    ax.plot([entry_naive, tp_hit_naive], [tp["target"], tp["target"]], color=tp["color"], linewidth=1.25, alpha=0.58, linestyle=tp["linestyle"], zorder=7)
                    _draw_trade_move_arrow(ax, entry_naive, entry_price, tp_hit_naive, tp["target"], linewidth=2.0, alpha=0.88)
                    ax.scatter([tp_hit_naive], [tp["target"]], marker="*", s=210, color=tp["color"], edgecolors="#101418", linewidths=0.85, zorder=10)
                    ax.text(
                        tp_hit_naive,
                        tp["target"],
                        f" EXIT {tp['time_to_min']}m",
                        ha="left",
                        va="bottom",
                        color=tp["color"],
                        fontsize=8,
                        fontweight="bold",
                        bbox={"facecolor": "#101418", "edgecolor": tp["color"], "alpha": 0.82, "boxstyle": "round,pad=0.16"},
                    )
            mm = _visual_mfe_mae_fields(row)
            mfe_ts = _parse_utc_ts(mm["mfe_time_utc"])
            if mfe_ts is not None and start <= mfe_ts <= end and mm["mfe_price"] > 0:
                mfe_naive = mfe_ts.tz_convert(None)
                ax.scatter([mfe_naive], [mm["mfe_price"]], marker="D", s=120, color=MFE_COLOR, edgecolors="#101418", linewidths=0.75, zorder=10)
                ax.text(
                    mfe_naive,
                    mm["mfe_price"],
                    f" MFE diagnostic +{mm['mfe_pct']:.2f}%",
                    ha="left",
                    va="top",
                    color=MFE_COLOR,
                    fontsize=8,
                    fontweight="bold",
                    bbox={"facecolor": "#101418", "edgecolor": MFE_COLOR, "alpha": 0.82, "boxstyle": "round,pad=0.16"},
                )
            mae_ts = _parse_utc_ts(mm["mae_time_utc"])
            if mae_ts is not None and start <= mae_ts <= end and mm["mae_price"] > 0:
                mae_naive = mae_ts.tz_convert(None)
                ax.scatter([mae_naive], [mm["mae_price"]], marker="v", s=105, color=MAE_COLOR, edgecolors="#101418", linewidths=0.75, zorder=10)
                ax.text(
                    mae_naive,
                    mm["mae_price"],
                    f" MAE MIN {mm['mae_pct']:.2f}%",
                    ha="left",
                    va="bottom",
                    color=MAE_COLOR,
                    fontsize=8,
                    fontweight="bold",
                    bbox={"facecolor": "#101418", "edgecolor": MAE_COLOR, "alpha": 0.82, "boxstyle": "round,pad=0.16"},
                )
            ax.text(
                0.012,
                0.985,
                _row_text(row),
                transform=ax.transAxes,
                ha="left",
                va="top",
                color="#f5f5f5",
                fontsize=7,
                bbox={"facecolor": "#101418", "edgecolor": "#455a64", "alpha": 0.78, "boxstyle": "round,pad=0.28"},
            )
            ax.set_xlim(start.tz_convert(None).to_pydatetime(), end.tz_convert(None).to_pydatetime())
            ax.set_title(f"{symbol} {row.get('day_utc')} | entry {entry_ts.strftime('%H:%M')} | post-entry", color="white", fontsize=9)
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%H:%M"))
            ax.tick_params(axis="x", labelrotation=25)
        for ax in flat[len(page_rows) :]:
            ax.axis("off")
            ax.set_facecolor("#101418")
        fig.suptitle("STAS3 V2: entry -> review TP ladder | LONG-only | audit, not ML/scorer/Optuna.", color="white", fontsize=15)
        fig.subplots_adjust(left=0.045, right=0.985, top=0.955, bottom=0.045, hspace=0.28, wspace=0.12)
        path = out_dir / f"STAS3_ENTRY_LADDER_PAGE_{len(paths) + 1:02d}.png"
        fig.savefig(path, dpi=150, facecolor=fig.get_facecolor())
        plt.close(fig)
        paths.append(path)
    return paths


def _render_big_move_pages(
    *,
    market_df: pd.DataFrame,
    rows: list[dict[str, Any]],
    symbol: str,
    timeframe: str,
    out_dir: Path,
    render_limit: int = 48,
) -> list[Path]:
    review_order = {
        "EARLY_1PCT_TRAIL_REVIEW": 0,
        "CLEAN_MFE_RUNNER_REVIEW": 1,
        "BIG_MFE_BUT_DEEP_MAE_REVIEW": 2,
        "LATE_MFE_PUMP_REVIEW": 3,
    }
    selected = [row for row in rows if row.get("exit_review_bucket") in review_order]
    selected = sorted(
        selected,
        key=lambda row: (
            review_order.get(str(row.get("exit_review_bucket")), 99),
            _safe_float(row.get("time_to_1p0pct_min"), 999999.0),
            -_safe_float(row.get("actual_mfe_pct"), 0.0),
        ),
    )
    if render_limit > 0:
        selected = selected[:render_limit]
    paths: list[Path] = []
    per_page = 4
    for page_start in range(0, len(selected), per_page):
        page_rows = selected[page_start : page_start + per_page]
        fig, axes = plt.subplots(2, 2, figsize=(24, 15), sharex=False)
        fig.patch.set_facecolor("#101418")
        flat = list(axes.flatten())
        for ax, row in zip(flat, page_rows):
            _style_axis(ax)
            se = _visual_signal_entry_fields(row)
            mm = _visual_mfe_mae_fields(row)
            entry_ts = _parse_utc_ts(se["entry_time_utc"])
            signal_ts = _parse_utc_ts(se["signal_time_utc"]) or entry_ts
            mfe_ts = _parse_utc_ts(mm["mfe_time_utc"])
            if entry_ts is None or signal_ts is None:
                ax.axis("off")
                continue
            start = min(signal_ts, entry_ts) - pd.Timedelta(minutes=20)
            end = (mfe_ts + pd.Timedelta(minutes=90)) if mfe_ts is not None else entry_ts + pd.Timedelta(hours=12)
            end = max(end, entry_ts + pd.Timedelta(hours=2))
            win = market_df[(market_df["open_time_utc"] >= start) & (market_df["open_time_utc"] <= end)].reset_index(drop=True)
            if win.empty:
                ax.axis("off")
                continue
            _draw_candles(ax, win, timeframe, linewidth=0.28)
            entry_naive = entry_ts.tz_convert(None)
            signal_naive = signal_ts.tz_convert(None)
            entry_price = _safe_float(se["entry_price"])
            signal_price = _safe_float(se["signal_price"], entry_price)
            ax.axvline(entry_naive, color=ENTRY_COLOR, linewidth=1.1, alpha=0.76)
            ax.scatter([signal_naive], [signal_price], marker="o", s=72, color=SIGNAL_COLOR, edgecolors="white", linewidths=0.55, zorder=9)
            ax.scatter([entry_naive], [entry_price], marker="^", s=105, color=ENTRY_COLOR, edgecolors="white", linewidths=0.65, zorder=10)
            ax.plot([signal_naive, entry_naive], [signal_price, entry_price], color=ENTRY_COLOR, alpha=0.70, linewidth=1.15, linestyle=":", zorder=7)
            tp = _visual_tp_fields(row)
            if tp["target"] > 0 and tp["hit_time_utc"]:
                tp_hit = _parse_utc_ts(tp["hit_time_utc"])
                if tp_hit is not None:
                    tp_hit_naive = tp_hit.tz_convert(None)
                    ax.axhline(tp["target"], color=tp["color"], linewidth=1.05, alpha=0.65, linestyle=tp["linestyle"])
                    _draw_trade_move_arrow(ax, entry_naive, entry_price, tp_hit_naive, tp["target"], linewidth=1.45, alpha=0.75, color=TP_MOVE_COLOR)
                    ax.scatter([tp_hit_naive], [tp["target"]], marker="*", s=135, color=tp["color"], edgecolors="#101418", linewidths=0.6, zorder=10)
            if mfe_ts is not None and mm["mfe_price"] > 0:
                mfe_naive = mfe_ts.tz_convert(None)
                _draw_trade_move_arrow(ax, signal_naive, signal_price, mfe_naive, mm["mfe_price"], linewidth=2.0, alpha=0.82, color=MFE_MOVE_COLOR)
                ax.scatter([mfe_naive], [mm["mfe_price"]], marker="D", s=150, color=MFE_COLOR, edgecolors="#101418", linewidths=0.8, zorder=11)
                ax.text(
                    mfe_naive,
                    mm["mfe_price"],
                    f" MFE diagnostic +{mm['mfe_pct']:.2f}%",
                    ha="left",
                    va="top",
                    color=MFE_COLOR,
                    fontsize=8,
                    fontweight="bold",
                    bbox={"facecolor": "#101418", "edgecolor": MFE_COLOR, "alpha": 0.82, "boxstyle": "round,pad=0.16"},
                )
            mae_ts = _parse_utc_ts(mm["mae_time_utc"])
            if mae_ts is not None and mm["mae_price"] > 0 and start <= mae_ts <= end:
                mae_naive = mae_ts.tz_convert(None)
                ax.scatter([mae_naive], [mm["mae_price"]], marker="v", s=120, color=MAE_COLOR, edgecolors="#101418", linewidths=0.75, zorder=10)
            info = "\n".join(
                [
                    f"{row.get('candidate_id')} | {row.get('exit_review_bucket')}",
                    f"signal->{row.get('mfe_from_signal_pct')}% | entry->{row.get('mfe_from_entry_pct')}%",
                    f"1% time {row.get('time_to_1p0pct_min')}m | MFE time {row.get('entry_to_mfe_min')}m",
                    f"MAE before MFE {row.get('mae_before_mfe_pct')}%",
                    f"extra after 1% {row.get('extra_after_1p0pct_pct')}%",
                ]
            )
            ax.text(
                0.012,
                0.985,
                info,
                transform=ax.transAxes,
                ha="left",
                va="top",
                color="#f5f5f5",
                fontsize=8,
                bbox={"facecolor": "#101418", "edgecolor": "#455a64", "alpha": 0.78, "boxstyle": "round,pad=0.28"},
            )
            ax.set_xlim(start.tz_convert(None).to_pydatetime(), end.tz_convert(None).to_pydatetime())
            ax.set_title(f"{symbol} {row.get('day_utc')} | big move review | entry {entry_ts.strftime('%m-%d %H:%M')}", color="white", fontsize=9)
            ax.xaxis.set_major_formatter(mdates.DateFormatter("%m-%d %H:%M"))
            ax.tick_params(axis="x", labelrotation=25)
        for ax in flat[len(page_rows) :]:
            ax.axis("off")
            ax.set_facecolor("#101418")
        fig.suptitle("STAS3 V2: MFE DIAGNOSTIC | hindsight fact, not TP/exit/strategy/ML", color="white", fontsize=14)
        fig.tight_layout(rect=[0, 0, 1, 0.975])
        path = out_dir / f"STAS3_BIG_MOVE_REVIEW_PAGE_{len(paths) + 1:02d}.png"
        fig.savefig(path, dpi=145, facecolor=fig.get_facecolor())
        plt.close(fig)
        paths.append(path)
    return paths


def _render_browse_index(*, out_path: Path, day_summary: list[dict[str, Any]], symbol: str, timeframe: str) -> None:
    fig, ax = plt.subplots(figsize=(15, max(7, 1.0 + len(day_summary) * 0.72)))
    fig.patch.set_facecolor("#101418")
    ax.set_facecolor("#101418")
    ax.axis("off")
    ax.set_title(f"STAS3 Percent Ladder Review | {symbol} {timeframe}", color="white", fontsize=15, loc="left")
    lines = ["Открыть папку дня, затем 00_OVERVIEW и страницы STAS3_ENTRY_LADDER.", "", "day | entries | hit 1% | avg max ladder | avg MAE 15m"]
    for row in day_summary:
        lines.append(
            f"{row.get('day_utc')} | {row.get('count')} | {row.get('hit_1pct_count')} | "
            f"{row.get('avg_max_ladder_hit_pct')}% | {row.get('avg_mae_15m_pct')}%"
        )
    ax.text(0.02, 0.92, "\n".join(lines), transform=ax.transAxes, color="#eceff1", fontsize=11, va="top", family="monospace")
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, facecolor=fig.get_facecolor())
    plt.close(fig)


def _write_report(path: Path, payload: dict[str, Any]) -> None:
    summary = payload["summary"]
    artifacts = payload["artifacts"]
    ladder = ", ".join(f"{pct:g}%" for pct in payload["percent_ladder"])
    phase_lines = [
        f"- `{row.get('phase_at_entry') or '<empty>'}`: rows `{row.get('count')}`, fast TP `{row.get('reasonable_tp_fast_pct')}`, hold TP `{row.get('reasonable_tp_hold_pct')}`, band `{row.get('reasonable_tp_band')}`."
        for row in payload.get("tp_ladder_by_phase", [])
    ]
    lines = [
        "# STAS3 V2 Percent Ladder Review",
        "",
        f"Статус: `{payload['status']}`.",
        "",
        "Назначение: отдельный LONG-only post-entry audit поверх входов Stas1 и контекста Stas2. Здесь считаются процентная лестница, MFE/MAE, фаза рынка на момент входа и первая проверочная версия реалистичного TP по фазам.",
        "",
        "## Главное правило",
        "",
        "Stas3 V2 смотрит свечи после `entry_time_utc`, поэтому его post-entry поля нельзя использовать как causal features входа, scorer, target-lock или ML-label без отдельного ручного approved-ledger. Это анализ и проверка, не торговая стратегия. `SHORT` используется только как процентный риск-фон для LONG.",
        "",
        "## Сетка",
        "",
        f"`{ladder}`",
        "",
        "## Сводка",
        "",
        f"- Диапазон: `{payload['date_range']['start_day']}` .. `{payload['date_range']['end_day']}`.",
        f"- Stas2 rows input: `{summary['stas2_rows_input']}`.",
        f"- Entry rows: `{summary['entry_rows']}`.",
        f"- Skipped rows: `{summary['skipped_rows']}`.",
        f"- Row-count parity: `{summary['row_count_parity_ok']}`.",
        f"- Hit 1%: `{summary['hit_1pct_count']}`.",
        f"- Reasonable TP available: `{summary['reasonable_tp_available_count']}`.",
        f"- 1% mismatch count: `{summary['tp_wrong_1pct_count']}`.",
        f"- Noise entry count: `{summary['noise_entry_count']}`.",
        f"- Fast clean: `{summary['fast_clean_count']}`.",
        f"- Late-pump dependent: `{summary['late_pump_count']}`.",
        f"- Clean MFE runner review: `{summary['clean_mfe_runner_count']}`.",
        f"- Early 1% trail review: `{summary['early_1pct_trail_review_count']}`.",
        f"- Late MFE pump review: `{summary['late_mfe_pump_review_count']}`.",
        f"- Big MFE but deep MAE review: `{summary['big_mfe_deep_mae_review_count']}`.",
        f"- Fast 0.3 only review: `{summary['fast_0p3_only_review_count']}`.",
        f"- Missing OHLCV sources: `{summary['missing_sources']}`.",
        "",
        "## TP по фазам",
        "",
        *(phase_lines if phase_lines else ["- Нет строк для фазовой сводки."]),
        "",
        "## Артефакты",
        "",
        f"- `run_dir`: `{artifacts['run_dir']}`",
        f"- `records_csv`: `{artifacts['records_csv']}`",
        f"- `v2_entry_tp_audit_csv`: `{artifacts['v2_entry_tp_audit_csv']}`",
        f"- `v2_context_bundle_csv`: `{artifacts['v2_context_bundle_csv']}`",
        f"- `v2_tp_ladder_by_phase_csv`: `{artifacts['v2_tp_ladder_by_phase_csv']}`",
        f"- `v2_wrong_tp_review_csv`: `{artifacts['v2_wrong_tp_review_csv']}`",
        f"- `v2_skipped_csv`: `{artifacts['v2_skipped_csv']}`",
        f"- `entry_phase_csv`: `{artifacts['entry_phase_csv']}`",
        f"- `actual_movement_csv`: `{artifacts['actual_movement_csv']}`",
        f"- `reasonable_tp_csv`: `{artifacts['reasonable_tp_csv']}`",
        f"- `low_signal_to_mfe_csv`: `{artifacts['low_signal_to_mfe_csv']}`",
        f"- `entry_to_tp_path_csv`: `{artifacts['entry_to_tp_path_csv']}`",
        f"- `exit_review_csv`: `{artifacts['exit_review_csv']}`",
        f"- `tp_ladder_v0_ru`: `{artifacts['tp_ladder_v0_ru']}`",
        f"- `xlsx`: `{artifacts['xlsx']}`",
        f"- `browse_dir`: `{artifacts['browse_dir']}`",
        f"- `report_ru`: `{artifacts['report_ru']}`",
        "",
        "## Граница",
        "",
        "Разрешено: visual review, TP/percent ladder audit, MFE/MAE, зависание сделки, DCA/risk review. `MFE MAX` разрешен только как diagnostic/hindsight fact, не как TP/exit.",
        "",
        "Запрещено без отдельного решения: ML/export/training, Optuna, scorer, target-lock, API.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_tp_ladder_report(path: Path, payload: dict[str, Any]) -> None:
    config = payload["tp_ladder_config"]
    lines = [
        "# STAS3 V2 TP Ladder",
        "",
        f"Статус: `{payload['status']}`.",
        "",
        "Это первая проверочная лестница TP по фазам для LONG-only V2. Она построена только из фактического post-entry движения и не является торговой стратегией.",
        "",
        "## Метод",
        "",
        f"- Полная сетка: `{', '.join(f'{pct:g}%' for pct in payload['percent_ladder'])}`.",
        f"- Fast TP считается по достижению порога за `{config['tp_fast_minutes']}` минут.",
        f"- Минимум строк в группе: `{config['tp_min_samples']}`.",
        f"- Hold hit-rate threshold: `{config['tp_hit_rate_min']}`.",
        f"- Fast hit-rate threshold: `{config['tp_fast_hit_rate_min']}`.",
        "",
        "## Лестница по фазам",
        "",
    ]
    for row in payload.get("tp_ladder_by_phase", []):
        lines.append(
            f"- `{row.get('phase_at_entry') or '<empty>'}`: rows `{row.get('count')}`, "
            f"fast TP `{row.get('reasonable_tp_fast_pct')}`, hold TP `{row.get('reasonable_tp_hold_pct')}`, "
            f"median move `{row.get('median_actual_move_pct')}`, band `{row.get('reasonable_tp_band')}`."
        )
    lines.extend(["", "## Лестница по профилям", ""])
    for row in payload.get("tp_ladder_by_profile", []):
        lines.append(
            f"- `{row.get('tp_phase_profile') or '<empty>'}`: rows `{row.get('count')}`, "
            f"fast TP `{row.get('reasonable_tp_fast_pct')}`, hold TP `{row.get('reasonable_tp_hold_pct')}`, "
            f"band `{row.get('reasonable_tp_band')}`."
        )
    lines.extend(
        [
            "",
            "## Ограничение",
            "",
            "Группы с малым числом строк помечаются как `INSUFFICIENT_DATA_REVIEW_ONLY`; такие проценты нельзя переносить в правила входа без отдельного утверждения.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_review(
    *,
    start_day: str,
    end_day: str,
    symbol: str,
    timeframe: str,
    out_dir: Path,
    run_label: str,
    stas2_run_dir: Path | None,
    hold_hours: float,
    post_plot_minutes: int,
    render_limit: int,
    tp_fast_minutes: int,
    tp_min_samples: int,
    tp_hit_rate_min: float,
    tp_fast_hit_rate_min: float,
) -> dict[str, Any]:
    root = Path.cwd()
    source_stas2_run = stas2_run_dir or _latest_stas2_run(root)
    stas2_csv = source_stas2_run / "STAS2_RECORDS.csv"
    stas2_rows_all = _read_csv_rows(stas2_csv)
    stas2_rows = [row for row in stas2_rows_all if start_day <= str(row.get("day_utc", "")) <= end_day]
    context_tables = _load_stas2_context_tables(source_stas2_run)

    market_df, missing_sources = _load_days(root, start_day, end_day, hold_hours, symbol, timeframe)
    run_id = f"{run_label}_{_run_stamp()}"
    run_dir = out_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=False)

    stas2_source_run_rel = _rel(root, source_stas2_run)
    skipped_rows: list[dict[str, Any]] = []
    rows = compute_stas3_rows(stas2_rows, market_df, hold_hours, skipped_rows=skipped_rows, context_tables=context_tables)
    for row in rows:
        row["stas2_source_run"] = stas2_source_run_rel
    for row in skipped_rows:
        row["stas2_source_run"] = stas2_source_run_rel

    profile_recommendations = _annotate_rows_with_tp_recommendations(
        rows,
        min_samples=tp_min_samples,
        fast_minutes=tp_fast_minutes,
        hit_rate_min=tp_hit_rate_min,
        fast_hit_rate_min=tp_fast_hit_rate_min,
    )
    _annotate_big_move_review(rows)
    by_day = _summary(rows, ["day_utc"])
    by_setup = _summary(rows, ["entry_setup_quality_label"])
    by_session = _summary(rows, ["effective_session_label"])
    by_verdict = _summary(rows, ["stas3_verdict"])
    by_exit_review = _summary(rows, ["exit_review_bucket"])
    tp_ladder_by_phase = _tp_summary(
        rows,
        ["phase_at_entry"],
        min_samples=tp_min_samples,
        fast_minutes=tp_fast_minutes,
        hit_rate_min=tp_hit_rate_min,
        fast_hit_rate_min=tp_fast_hit_rate_min,
    )
    tp_ladder_by_long_wave = _tp_summary(
        rows,
        ["long_wave_at_entry"],
        min_samples=tp_min_samples,
        fast_minutes=tp_fast_minutes,
        hit_rate_min=tp_hit_rate_min,
        fast_hit_rate_min=tp_fast_hit_rate_min,
    )
    tp_ladder_by_setup = _tp_summary(
        rows,
        ["entry_setup_quality_label"],
        min_samples=tp_min_samples,
        fast_minutes=tp_fast_minutes,
        hit_rate_min=tp_hit_rate_min,
        fast_hit_rate_min=tp_fast_hit_rate_min,
    )
    tp_ladder_by_session = _tp_summary(
        rows,
        ["effective_session_label"],
        min_samples=tp_min_samples,
        fast_minutes=tp_fast_minutes,
        hit_rate_min=tp_hit_rate_min,
        fast_hit_rate_min=tp_fast_hit_rate_min,
    )
    phase_percent_summary = _tp_summary(
        rows,
        ["phase_at_entry", "long_wave_at_entry", "entry_setup_quality_label"],
        min_samples=tp_min_samples,
        fast_minutes=tp_fast_minutes,
        hit_rate_min=tp_hit_rate_min,
        fast_hit_rate_min=tp_fast_hit_rate_min,
    )
    tp_ladder_by_profile = sorted(profile_recommendations.values(), key=lambda row: str(row.get("tp_phase_profile", "")))

    provenance_cols = ["stas1_source_run", "stas2_source_run", "stas1_record_id", "stas2_record_id", "candidate_id"]
    context_columns = [
        *provenance_cols,
        "day_utc",
        "entry_time_utc",
        "direction_scope",
        "long_only_flag",
        "short_context_only_flag",
        "context_scope",
        "wave_context_scope",
        "session_time_bucket_code",
        "session_time_bucket_label",
        "effective_session_code",
        "effective_session_label",
        "real_tradfi_session_open",
        "hour_background_phase",
        "hour_background_phase_rank",
        "hour_range_pct",
        "hour_path_pct",
        "hour_close_move_pct",
        "hour_long_wave_phase",
        "hour_long_wave_up_from_low_pct",
        "hour_long_wave_pullback_from_post_low_high_pct",
        "hour_short_wave_phase",
        "hour_short_wave_down_from_high_pct",
        "hour_direction_bias",
        "pre_5m_range_pct",
        "pre_15m_range_pct",
        "pre_30m_range_pct",
        "pre_60m_range_pct",
        "pre_5m_path_pct",
        "pre_15m_path_pct",
        "pre_30m_path_pct",
        "pre_60m_path_pct",
        "pre_5m_long_wave_phase",
        "pre_15m_long_wave_phase",
        "pre_30m_long_wave_phase",
        "pre_60m_long_wave_phase",
        "macro_wave_id",
        "macro_wave_segment_kind",
        "macro_wave_direction",
        "macro_wave_status",
        "macro_wave_visible_move_pct",
        "macro_wave_full_move_pct",
        "macro_wave_carry_from_prev_day",
        "macro_wave_carry_to_next_day",
        "continuous_wave_id",
        "continuous_wave_segment_kind",
        "continuous_wave_direction",
        "continuous_wave_move_pct",
        "volume_context_status",
        "volume_ratio20",
    ]
    v2_entry_tp_audit_columns = [
        *provenance_cols,
        "day_utc",
        "anchor_time_utc",
        "anchor_low_price",
        "entry_time_utc",
        "entry_open_price",
        "entry_price_5bps",
        "entry_price_for_calc",
        "entry_price_source",
        "entry_price_locked_flag",
        "direction_scope",
        "long_only_flag",
        "short_context_only_flag",
        "context_scope",
        "phase_at_entry",
        "hour_background_phase",
        "hour_long_wave_phase",
        "hour_short_wave_down_from_high_pct",
        "macro_wave_direction",
        "wave_context_scope",
        "actual_mfe_pct",
        "actual_mae_pct",
        "max_ladder_hit_pct",
        "max_feasible_review_tp_pct",
        "ideal_review_tp_pct",
        "ideal_review_tp_reason",
        "ideal_review_tp_warning",
        "reasonable_tp_pct",
        "move_scale_bucket",
        "noise_flag",
        "wrong_tp_flag",
        "tp_vs_1pct_label",
        "tp_vs_1pct_reason",
        "stas3_verdict",
    ]
    entry_phase_columns = [
        *provenance_cols,
        "day_utc",
        "anchor_time_utc",
        "stas1_confirmation_time_utc",
        "entry_time_utc",
        "anchor_low_price",
        "entry_open_price",
        "entry_price_5bps",
        "entry_price_for_calc",
        "entry_price_source",
        "entry_price_locked_flag",
        "context_before_entry_check",
        "phase_at_entry",
        "long_wave_at_entry",
        "pre_15m_background_phase",
        "pre_15m_long_wave_phase",
        "pre_30m_background_phase",
        "pre_30m_long_wave_phase",
        "entry_setup_quality_label",
        "entry_setup_quality_rank",
        "entry_setup_quality_reason",
        "effective_session_label",
        "day_type",
        "tp_phase_profile",
    ]
    actual_movement_columns = [
        *provenance_cols,
        "day_utc",
        "anchor_time_utc",
        "stas1_confirmation_time_utc",
        "entry_time_utc",
        "anchor_low_price",
        "entry_open_price",
        "entry_price_5bps",
        "entry_price_for_calc",
        "actual_move_pct",
        "actual_mfe_pct",
        "actual_mfe_time_utc",
        "actual_mae_pct",
        "actual_mae_time_utc",
        "actual_end_time_utc",
        "drawdown_before_first_hit_pct",
        "first_ladder_hit",
        "first_ladder_hit_time_utc",
        "first_post_entry_event",
        "post_5m_mfe_pct",
        "post_5m_mae_pct",
        "post_15m_mfe_pct",
        "post_15m_mae_pct",
        "post_30m_mfe_pct",
        "post_30m_mae_pct",
        "post_60m_mfe_pct",
        "post_60m_mae_pct",
        "max_ladder_hit_pct",
        "stas3_verdict",
    ]
    reasonable_tp_columns = [
        *provenance_cols,
        "day_utc",
        "anchor_time_utc",
        "stas1_confirmation_time_utc",
        "entry_time_utc",
        "anchor_low_price",
        "entry_open_price",
        "entry_price_5bps",
        "entry_price_for_calc",
        "phase_at_entry",
        "long_wave_at_entry",
        "entry_setup_quality_label",
        "tp_phase_profile",
        "phase_profile_sample_count",
        "data_status",
        "actual_move_pct",
        "reasonable_tp_pct",
        "reasonable_tp_price",
        "max_feasible_review_tp_pct",
        "ideal_review_tp_pct",
        "ideal_review_tp_reason",
        "ideal_review_tp_warning",
        "reasonable_tp_fast_pct",
        "reasonable_tp_hold_pct",
        "reasonable_tp_band",
        "reasonable_tp_source",
        "reasonable_tp_reason",
        "tp_vs_1pct_label",
        "tp_vs_1pct_reason",
        "stas3_verdict",
    ]
    low_signal_to_mfe_columns = [
        *provenance_cols,
        "day_utc",
        "anchor_time_utc",
        "anchor_low_price",
        "stas1_confirmation_time_utc",
        "entry_time_utc",
        "entry_price_5bps",
        "entry_price_for_calc",
        "signal_to_entry_move_pct",
        "mfe_max_price",
        "mfe_max_time_utc",
        "mfe_from_signal_pct",
        "mfe_from_entry_pct",
        "signal_to_mfe_min",
        "entry_to_mfe_min",
        "mae_min_price",
        "mae_min_time_utc",
        "mae_from_entry_pct",
        "entry_to_mae_min",
        "mae_before_mfe_price",
        "mae_before_mfe_time_utc",
        "mae_before_mfe_pct",
        "phase_at_entry",
        "long_wave_at_entry",
        "entry_setup_quality_label",
        "effective_session_label",
        "exit_review_bucket",
        "exit_review_reason",
    ]
    entry_to_tp_path_columns = [
        *provenance_cols,
        "day_utc",
        "entry_time_utc",
        "entry_price_5bps",
        "entry_price_for_calc",
        "hit_0p3pct",
        "time_to_0p3pct_min",
        "mae_before_0p3pct_pct",
        "mfe_after_0p3pct_pct",
        "extra_after_0p3pct_pct",
        "hit_0p5pct",
        "time_to_0p5pct_min",
        "mae_before_0p5pct_pct",
        "mfe_after_0p5pct_pct",
        "extra_after_0p5pct_pct",
        "hit_1p0pct",
        "time_to_1p0pct_min",
        "mae_before_1p0pct_pct",
        "mfe_after_1p0pct_pct",
        "extra_after_1p0pct_pct",
        "hit_2p0pct",
        "time_to_2p0pct_min",
        "mae_before_2p0pct_pct",
        "mfe_after_2p0pct_pct",
        "extra_after_2p0pct_pct",
        "actual_mfe_pct",
        "entry_to_mfe_min",
        "after_mfe_min_pct",
        "giveback_after_mfe_pct",
        "giveback_after_mfe_to_close_pct",
        "giveback_after_mfe_to_min_pct",
        "reasonable_tp_pct",
        "ideal_review_tp_pct",
        "ideal_review_tp_warning",
        "tp_to_mfe_extra_pct",
        "exit_review_bucket",
        "exit_review_reason",
    ]
    exit_review_columns = [
        *provenance_cols,
        "day_utc",
        "anchor_time_utc",
        "entry_time_utc",
        "entry_price_for_calc",
        "phase_at_entry",
        "long_wave_at_entry",
        "entry_setup_quality_label",
        "effective_session_label",
        "actual_mfe_pct",
        "actual_mae_pct",
        "entry_to_mfe_min",
        "mae_before_mfe_pct",
        "time_to_0p3pct_min",
        "time_to_0p5pct_min",
        "time_to_1p0pct_min",
        "mae_before_1p0pct_pct",
        "extra_after_1p0pct_pct",
        "extra_after_0p3pct_pct",
        "giveback_after_mfe_pct",
        "reasonable_tp_pct",
        "ideal_review_tp_pct",
        "ideal_review_tp_warning",
        "tp_to_mfe_extra_pct",
        "exit_review_bucket",
        "exit_review_reason",
        "stas3_verdict",
        "tp_vs_1pct_label",
    ]
    wrong_tp_columns = [col for col in v2_entry_tp_audit_columns if col not in {"ideal_review_tp_reason"}] + ["ideal_review_tp_reason"]
    skipped_columns = ["stas2_source_run", "stas2_record_id", "candidate_id", "day_utc", "entry_time_utc", "skip_reason"]
    context_bundle_rows = _project_rows(rows, context_columns)
    v2_entry_tp_audit_rows = _project_rows(rows, v2_entry_tp_audit_columns)
    wrong_tp_rows = _project_rows([row for row in rows if bool(row.get("wrong_tp_flag")) or bool(row.get("noise_flag"))], wrong_tp_columns)
    entry_phase_rows = _project_rows(rows, entry_phase_columns)
    actual_movement_rows = _project_rows(rows, actual_movement_columns)
    reasonable_tp_rows = _project_rows(rows, reasonable_tp_columns)
    low_signal_to_mfe_rows = _project_rows(rows, low_signal_to_mfe_columns)
    entry_to_tp_path_rows = _project_rows(rows, entry_to_tp_path_columns)
    exit_review_rows = _project_rows(rows, exit_review_columns)

    records_csv = run_dir / "STAS3_RECORDS.csv"
    entry_phase_csv = run_dir / "STAS3_ENTRY_PHASE_TABLE.csv"
    actual_movement_csv = run_dir / "STAS3_ACTUAL_MOVEMENT.csv"
    reasonable_tp_csv = run_dir / "STAS3_REASONABLE_TP.csv"
    low_signal_to_mfe_csv = run_dir / "STAS3_LOW_SIGNAL_TO_MFE_MAX.csv"
    entry_to_tp_path_csv = run_dir / "STAS3_ENTRY_TO_TP_PATH.csv"
    exit_review_csv = run_dir / "STAS3_EXIT_REVIEW_BUCKETS.csv"
    day_summary_csv = run_dir / "STAS3_DAY_SUMMARY.csv"
    setup_summary_csv = run_dir / "STAS3_SETUP_SUMMARY.csv"
    session_summary_csv = run_dir / "STAS3_SESSION_SUMMARY.csv"
    verdict_summary_csv = run_dir / "STAS3_VERDICT_SUMMARY.csv"
    exit_review_summary_csv = run_dir / "STAS3_EXIT_REVIEW_SUMMARY.csv"
    tp_by_phase_csv = run_dir / "STAS3_TP_LADDER_BY_PHASE.csv"
    tp_by_long_wave_csv = run_dir / "STAS3_TP_LADDER_BY_LONG_WAVE.csv"
    tp_by_profile_csv = run_dir / "STAS3_TP_LADDER_BY_PROFILE.csv"
    tp_by_setup_csv = run_dir / "STAS3_TP_LADDER_BY_SETUP.csv"
    tp_by_session_csv = run_dir / "STAS3_TP_LADDER_BY_SESSION.csv"
    phase_percent_summary_csv = run_dir / "STAS3_PHASE_PERCENT_SUMMARY.csv"
    skipped_csv = run_dir / "STAS3_SKIPPED_ROWS.csv"
    v2_entry_tp_audit_csv = run_dir / "STAS3_V2_ENTRY_TP_AUDIT.csv"
    v2_context_bundle_csv = run_dir / "STAS3_V2_CONTEXT_BUNDLE.csv"
    v2_tp_ladder_by_phase_csv = run_dir / "STAS3_V2_TP_LADDER_BY_PHASE.csv"
    v2_wrong_tp_review_csv = run_dir / "STAS3_V2_WRONG_TP_REVIEW.csv"
    v2_skipped_csv = run_dir / "STAS3_V2_SKIPPED_ROWS.csv"
    xlsx_path = run_dir / "STAS3_PERCENT_LADDER_TABLES.xlsx"
    payload_path = run_dir / "STAS3_PAYLOAD.json"
    report_path = run_dir / "STAS3_REPORT_RU.md"
    v2_report_path = run_dir / "STAS3_V2_REPORT_RU.md"
    tp_ladder_report_path = run_dir / "STAS3_V2_TP_LADDER_RU.md"
    _write_csv(records_csv, rows)
    _write_csv_columns(entry_phase_csv, entry_phase_rows, entry_phase_columns)
    _write_csv_columns(actual_movement_csv, actual_movement_rows, actual_movement_columns)
    _write_csv_columns(reasonable_tp_csv, reasonable_tp_rows, reasonable_tp_columns)
    _write_csv_columns(low_signal_to_mfe_csv, low_signal_to_mfe_rows, low_signal_to_mfe_columns)
    _write_csv_columns(entry_to_tp_path_csv, entry_to_tp_path_rows, entry_to_tp_path_columns)
    _write_csv_columns(exit_review_csv, exit_review_rows, exit_review_columns)
    _write_csv(day_summary_csv, by_day)
    _write_csv(setup_summary_csv, by_setup)
    _write_csv(session_summary_csv, by_session)
    _write_csv(verdict_summary_csv, by_verdict)
    _write_csv(exit_review_summary_csv, by_exit_review)
    _write_csv(tp_by_phase_csv, tp_ladder_by_phase)
    _write_csv(tp_by_long_wave_csv, tp_ladder_by_long_wave)
    _write_csv(tp_by_profile_csv, tp_ladder_by_profile)
    _write_csv(tp_by_setup_csv, tp_ladder_by_setup)
    _write_csv(tp_by_session_csv, tp_ladder_by_session)
    _write_csv(phase_percent_summary_csv, phase_percent_summary)
    _write_csv_columns(skipped_csv, skipped_rows, skipped_columns)
    _write_csv_columns(v2_entry_tp_audit_csv, v2_entry_tp_audit_rows, v2_entry_tp_audit_columns)
    _write_csv_columns(v2_context_bundle_csv, context_bundle_rows, context_columns)
    _write_csv(v2_tp_ladder_by_phase_csv, tp_ladder_by_phase)
    _write_csv_columns(v2_wrong_tp_review_csv, wrong_tp_rows, wrong_tp_columns)
    _write_csv_columns(v2_skipped_csv, skipped_rows, skipped_columns)
    _write_xlsx(
        xlsx_path,
        {
            "V2 Entry TP Audit": v2_entry_tp_audit_rows,
            "V2 Context Bundle": context_bundle_rows,
            "V2 Wrong TP Review": wrong_tp_rows,
            "V2 TP by phase": tp_ladder_by_phase,
            "Records": rows,
            "Entry phase": entry_phase_rows,
            "Actual movement": actual_movement_rows,
            "Reasonable TP": reasonable_tp_rows,
            "Low to MFE max": low_signal_to_mfe_rows,
            "Entry to TP path": entry_to_tp_path_rows,
            "Exit review": exit_review_rows,
            "TP by phase": tp_ladder_by_phase,
            "TP by long wave": tp_ladder_by_long_wave,
            "TP by profile": tp_ladder_by_profile,
            "TP by setup": tp_ladder_by_setup,
            "TP by session": tp_ladder_by_session,
            "Phase percent": phase_percent_summary,
            "Day summary": by_day,
            "Setup summary": by_setup,
            "Session summary": by_session,
            "Verdict summary": by_verdict,
            "Exit review summary": by_exit_review,
            "Skipped": skipped_rows,
        },
    )

    day_overviews: list[str] = []
    for day in _iter_days(start_day, end_day):
        day_df = market_df[
            (market_df["open_time_utc"] >= pd.Timestamp(day, tz="UTC"))
            & (market_df["open_time_utc"] < pd.Timestamp(day, tz="UTC") + pd.Timedelta(days=1))
        ].reset_index(drop=True)
        day_rows = [row for row in rows if row.get("day_utc") == day]
        if day_df.empty:
            continue
        path = run_dir / f"STAS3_DAY_OVERVIEW_{day.replace('-', '')}.png"
        _render_day_overview(df=day_df, rows=day_rows, day=day, symbol=symbol, timeframe=timeframe, out_path=path)
        day_overviews.append(_rel(root, path))
    entry_pages = _render_entry_pages(
        market_df=market_df,
        rows=rows,
        symbol=symbol,
        timeframe=timeframe,
        out_dir=run_dir,
        post_minutes=post_plot_minutes,
        render_limit=render_limit,
    )
    big_move_pages = _render_big_move_pages(
        market_df=market_df,
        rows=rows,
        symbol=symbol,
        timeframe=timeframe,
        out_dir=run_dir,
        render_limit=48,
    )

    browse_dir = run_dir / "BROWSE_BY_DAY"
    browse_dir.mkdir(exist_ok=True)
    browse_index = browse_dir / "00_RUN_INDEX.png"
    _render_browse_index(out_path=browse_index, day_summary=by_day, symbol=symbol, timeframe=timeframe)
    for day in _iter_days(start_day, end_day):
        folder = browse_dir / day
        folder.mkdir(exist_ok=True)
        overview = run_dir / f"STAS3_DAY_OVERVIEW_{day.replace('-', '')}.png"
        if overview.exists():
            (folder / f"00_{day.replace('-', '')}_OVERVIEW.png").write_bytes(overview.read_bytes())
        day_rows = [row for row in rows if row.get("day_utc") == day]
        _write_csv(folder / f"{day.replace('-', '')}_STAS3_RECORDS.csv", day_rows)

    tp_wrong_labels = {"GOOD_OR_OK_ENTRY_WRONG_1PCT_TP", "GOOD_OR_OK_ENTRY_BUT_1PCT_LATE", "HIT_1PCT_BUT_PHASE_TP_LOWER"}
    summary = {
        "stas2_rows_input": len(stas2_rows),
        "entry_rows": len(rows),
        "skipped_rows": len(skipped_rows),
        "row_count_parity_ok": len(stas2_rows) == len(rows) + len(skipped_rows),
        "hit_1pct_count": sum(1 for row in rows if bool(row.get("hit_1p0pct"))),
        "reasonable_tp_available_count": sum(1 for row in rows if _safe_float(row.get("reasonable_tp_pct"), 0.0) > 0),
        "tp_wrong_1pct_count": sum(1 for row in rows if row.get("tp_vs_1pct_label") in tp_wrong_labels),
        "noise_entry_count": sum(1 for row in rows if str(row.get("tp_vs_1pct_label", "")).startswith("NOISE_ENTRY")),
        "review_manually_count": sum(1 for row in rows if row.get("tp_vs_1pct_label") == "REVIEW_MANUALLY"),
        "no_tp_room_count": sum(1 for row in rows if row.get("tp_vs_1pct_label") == "NO_TP_ROOM"),
        "fast_clean_count": sum(1 for row in rows if row.get("stas3_verdict") == "FAST_CLEAN_1PCT"),
        "late_pump_count": sum(1 for row in rows if row.get("stas3_verdict") == "LATE_PUMP_DEPENDENT"),
        "clean_mfe_runner_count": sum(1 for row in rows if row.get("exit_review_bucket") == "CLEAN_MFE_RUNNER_REVIEW"),
        "early_1pct_trail_review_count": sum(1 for row in rows if row.get("exit_review_bucket") == "EARLY_1PCT_TRAIL_REVIEW"),
        "late_mfe_pump_review_count": sum(1 for row in rows if row.get("exit_review_bucket") == "LATE_MFE_PUMP_REVIEW"),
        "big_mfe_deep_mae_review_count": sum(1 for row in rows if row.get("exit_review_bucket") == "BIG_MFE_BUT_DEEP_MAE_REVIEW"),
        "fast_0p3_only_review_count": sum(1 for row in rows if row.get("exit_review_bucket") == "FAST_0P3_ONLY_REVIEW"),
        "missing_sources": len(missing_sources),
        "png_count": len(list(run_dir.glob("*.png"))) + len(list(browse_dir.rglob("*.png"))),
    }
    artifacts = {
        "run_dir": _rel(root, run_dir),
        "source_stas2_run": _rel(root, source_stas2_run),
        "records_csv": _rel(root, records_csv),
        "entry_phase_csv": _rel(root, entry_phase_csv),
        "actual_movement_csv": _rel(root, actual_movement_csv),
        "reasonable_tp_csv": _rel(root, reasonable_tp_csv),
        "low_signal_to_mfe_csv": _rel(root, low_signal_to_mfe_csv),
        "entry_to_tp_path_csv": _rel(root, entry_to_tp_path_csv),
        "exit_review_csv": _rel(root, exit_review_csv),
        "day_summary_csv": _rel(root, day_summary_csv),
        "setup_summary_csv": _rel(root, setup_summary_csv),
        "session_summary_csv": _rel(root, session_summary_csv),
        "verdict_summary_csv": _rel(root, verdict_summary_csv),
        "exit_review_summary_csv": _rel(root, exit_review_summary_csv),
        "tp_by_phase_csv": _rel(root, tp_by_phase_csv),
        "tp_by_long_wave_csv": _rel(root, tp_by_long_wave_csv),
        "tp_by_profile_csv": _rel(root, tp_by_profile_csv),
        "tp_by_setup_csv": _rel(root, tp_by_setup_csv),
        "tp_by_session_csv": _rel(root, tp_by_session_csv),
        "phase_percent_summary_csv": _rel(root, phase_percent_summary_csv),
        "skipped_csv": _rel(root, skipped_csv),
        "v2_entry_tp_audit_csv": _rel(root, v2_entry_tp_audit_csv),
        "v2_context_bundle_csv": _rel(root, v2_context_bundle_csv),
        "v2_tp_ladder_by_phase_csv": _rel(root, v2_tp_ladder_by_phase_csv),
        "v2_wrong_tp_review_csv": _rel(root, v2_wrong_tp_review_csv),
        "v2_skipped_csv": _rel(root, v2_skipped_csv),
        "tp_ladder_v0_ru": _rel(root, tp_ladder_report_path),
        "v2_report_ru": _rel(root, v2_report_path),
        "xlsx": _rel(root, xlsx_path),
        "json": _rel(root, payload_path),
        "report_ru": _rel(root, report_path),
        "browse_dir": _rel(root, browse_dir),
        "browse_index_png": _rel(root, browse_index),
        "day_overviews": day_overviews,
        "entry_pages": [_rel(root, path) for path in entry_pages],
        "big_move_pages": [_rel(root, path) for path in big_move_pages],
    }
    payload = {
        "schema_version": 4,
        "status": STATUS,
        "created_utc": _utc_now(),
        "run_id": run_id,
        "symbol": symbol,
        "timeframe": timeframe,
        "date_range": {"start_day": start_day, "end_day": end_day},
        "hold_hours": hold_hours,
        "tp_ladder_config": {
            "tp_fast_minutes": tp_fast_minutes,
            "tp_min_samples": tp_min_samples,
            "tp_hit_rate_min": tp_hit_rate_min,
            "tp_fast_hit_rate_min": tp_fast_hit_rate_min,
        },
        "percent_ladder": PERCENT_LADDER,
        "v2_contract": {
            "source_stas2_run_required": V2_SOURCE_STAS2_RUN,
            "work_days": V2_WORK_DAYS,
            "entry_price_for_calc": "entry_price_5bps",
            "direction_scope": "LONG_ONLY",
            "short_usage": "RISK_CONTEXT_ONLY",
            "mfe_max_usage": "HINDSIGHT_DIAGNOSTIC_NOT_TP",
        },
        "post_windows_min": POST_WINDOWS_MIN,
        "summary": summary,
        "missing_sources": missing_sources,
        "skipped_rows": skipped_rows,
        "day_summary": by_day,
        "setup_summary": by_setup,
        "session_summary": by_session,
        "verdict_summary": by_verdict,
        "exit_review_summary": by_exit_review,
        "tp_ladder_by_phase": tp_ladder_by_phase,
        "tp_ladder_by_long_wave": tp_ladder_by_long_wave,
        "tp_ladder_by_profile": tp_ladder_by_profile,
        "tp_ladder_by_setup": tp_ladder_by_setup,
        "tp_ladder_by_session": tp_ladder_by_session,
        "phase_percent_summary": phase_percent_summary,
        "artifacts": artifacts,
        "guardrails": [
            "LONG_ONLY",
            "SHORT_RISK_CONTEXT_ONLY",
            "NO_SHORT_ENTRY",
            "NO_SHORT_TP",
            "NO_SHORT_LADDER",
            "NO_ML_EXPORT_TRAINING",
            "NO_OPTUNA",
            "NO_SCORER",
            "NO_TARGET_LOCK",
            "NO_API",
            "POST_ENTRY_AUDIT_ONLY",
        ],
    }
    _write_json(payload_path, payload)
    _write_report(report_path, payload)
    _write_report(v2_report_path, payload)
    _write_tp_ladder_report(tp_ladder_report_path, payload)
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description="STAS3 post-entry percent ladder review, no ML/Optuna/API.")
    parser.add_argument("--start-day", default="2026-05-02")
    parser.add_argument("--end-day", default="2026-05-03")
    parser.add_argument("--symbol", default="SOLUSDT")
    parser.add_argument("--timeframe", default="1m")
    parser.add_argument("--out-dir", default="STAS3_PERCENT_LADDER_REVIEW/runs")
    parser.add_argument("--run-label", default="stas3_tp_ladder_v0")
    parser.add_argument("--stas2-run-dir", default="")
    parser.add_argument("--hold-hours", type=float, default=48.0)
    parser.add_argument("--post-plot-minutes", type=int, default=360)
    parser.add_argument("--render-limit", type=int, default=0, help="0 means render all entry pages")
    parser.add_argument("--tp-fast-minutes", type=int, default=120)
    parser.add_argument("--tp-min-samples", type=int, default=5)
    parser.add_argument("--tp-hit-rate-min", type=float, default=0.60)
    parser.add_argument("--tp-fast-hit-rate-min", type=float, default=0.50)
    args = parser.parse_args()

    payload = run_review(
        start_day=args.start_day,
        end_day=args.end_day,
        symbol=args.symbol,
        timeframe=args.timeframe,
        out_dir=Path(args.out_dir),
        run_label=args.run_label,
        stas2_run_dir=Path(args.stas2_run_dir) if args.stas2_run_dir else None,
        hold_hours=args.hold_hours,
        post_plot_minutes=args.post_plot_minutes,
        render_limit=args.render_limit,
        tp_fast_minutes=args.tp_fast_minutes,
        tp_min_samples=args.tp_min_samples,
        tp_hit_rate_min=args.tp_hit_rate_min,
        tp_fast_hit_rate_min=args.tp_fast_hit_rate_min,
    )
    print(
        json.dumps(
            {
                "status": payload["status"],
                "run_id": payload["run_id"],
                "summary": payload["summary"],
                "artifacts": payload["artifacts"],
                "missing_sources": payload["missing_sources"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
