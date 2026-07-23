from __future__ import annotations

import argparse
import csv
import json
import math
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mlbotnav.visual_entry_low_anchor_suggester import _load_ohlcv, _source_csv


STATUS = "STAS2_MARKET_PHASE_SESSION_AUDIT_READY_NO_ML_NO_OPTUNA"


PHASES = [
    ("Мертвая", 0.20),
    ("Слабая", 0.40),
    ("Средняя", 0.70),
    ("Выше средней", 1.00),
    ("Большая", 1.50),
    ("Сильная", 2.20),
    ("Очень сильная", math.inf),
]


SESSION_TIME_BUCKETS = [
    ("ASIA_PACIFIC", "Азия/Pacific time", 0, 390),
    ("PRE_LONDON", "Перед Лондоном", 390, 420),
    ("LONDON_ONLY", "Лондон без Нью-Йорка", 420, 810),
    ("LONDON_NY_OVERLAP", "Пересечение Лондон/Нью-Йорк", 810, 930),
    ("NY_ONLY", "Нью-Йорк без Лондона", 930, 1200),
    ("POST_NY", "После Нью-Йорка/слабые часы", 1200, 1440),
]


REAL_TRADFI_BUCKETS = {"ASIA_PACIFIC", "LONDON_ONLY", "LONDON_NY_OVERLAP", "NY_ONLY"}


@dataclass(frozen=True)
class Stas1Run:
    run_dir: Path
    csv_path: Path


def _utc_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _run_stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")


def _rel(root: Path, path: Path) -> str:
    try:
        return path.resolve().relative_to(root.resolve()).as_posix()
    except Exception:
        return path.as_posix()


def _iter_days(start_day: str, end_day: str) -> list[str]:
    start = pd.Timestamp(start_day)
    end = pd.Timestamp(end_day)
    if end < start:
        raise ValueError("--end-day must be >= --start-day")
    return [item.strftime("%Y-%m-%d") for item in pd.date_range(start, end, freq="D")]


def _session_for_timestamp(ts: pd.Timestamp) -> tuple[str, str]:
    utc = ts.tz_convert("UTC") if ts.tzinfo is not None else ts.tz_localize("UTC")
    minute_of_day = int(utc.hour) * 60 + int(utc.minute)
    for code, label, start, end in SESSION_TIME_BUCKETS:
        if start <= minute_of_day < end:
            return code, label
    raise ValueError(f"unexpected minute_of_day: {minute_of_day}")


def _day_type(ts: pd.Timestamp) -> str:
    utc = ts.tz_convert("UTC") if ts.tzinfo is not None else ts.tz_localize("UTC")
    return "weekend" if utc.dayofweek >= 5 else "weekday"


def _effective_session(code: str, label: str, day_type: str) -> tuple[str, str, bool]:
    if day_type == "weekend":
        return f"WEEKEND_{code}", f"Выходной: {label}", False
    is_real_tradfi_open = code in REAL_TRADFI_BUCKETS
    prefix = "Будни" if is_real_tradfi_open else "Будни, переход"
    return f"WEEKDAY_{code}", f"{prefix}: {label}", is_real_tradfi_open


def _phase_from_pct(value: float) -> str:
    for name, upper in PHASES:
        if value < upper:
            return name
    return PHASES[-1][0]


def _phase_rank(name: str) -> int:
    for idx, (phase, _) in enumerate(PHASES):
        if phase == name:
            return idx
    return -1


def _load_days(root: Path, days: list[str], symbol: str, timeframe: str) -> tuple[pd.DataFrame, list[str]]:
    frames: list[pd.DataFrame] = []
    missing: list[str] = []
    for day in days:
        source = _source_csv(root, day, timeframe, symbol)
        if not source.exists():
            missing.append(_rel(root, source))
            continue
        df = _load_ohlcv(source)
        df["day_utc"] = day
        df["source_csv"] = _rel(root, source)
        frames.append(df)
    if not frames:
        raise FileNotFoundError("no OHLCV sources found for requested period")
    out = pd.concat(frames, ignore_index=True).sort_values("open_time_utc").reset_index(drop=True)
    out["hour_utc"] = out["open_time_utc"].dt.hour
    out["date_hour_utc"] = out["open_time_utc"].dt.floor("h")
    out["candle_range_pct"] = (out["high"] - out["low"]) / out["open"].replace(0, np.nan) * 100.0
    out["candle_body_pct"] = (out["close"] - out["open"]).abs() / out["open"].replace(0, np.nan) * 100.0
    out["close_step_abs_pct"] = out["close"].pct_change().abs() * 100.0
    out["close_step_abs_pct"] = out["close_step_abs_pct"].fillna(0.0)
    out["volume_median_prior_24h"] = out["volume"].rolling(1440, min_periods=60).median().shift(1)
    out["range_median_prior_24h"] = out["candle_range_pct"].rolling(1440, min_periods=60).median().shift(1)
    return out, missing


def _safe_mean(values: pd.Series) -> float:
    if values.empty:
        return 0.0
    value = values.mean()
    return 0.0 if pd.isna(value) else float(value)


def _hour_rows(df: pd.DataFrame) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for hour_ts, group in df.groupby("date_hour_utc", sort=True):
        group = group.sort_values("open_time_utc")
        open_price = float(group.iloc[0]["open"])
        close_price = float(group.iloc[-1]["close"])
        high = float(group["high"].max())
        low = float(group["low"].min())
        hour_range_pct = (high - low) / max(open_price, 1e-12) * 100.0
        close_move_pct = (close_price - open_price) / max(open_price, 1e-12) * 100.0
        path_pct = float(group["close_step_abs_pct"].sum())
        median_candle_pct = _safe_mean(group["candle_range_pct"])
        median_body_pct = _safe_mean(group["candle_body_pct"])
        active_candle_share = float((group["candle_range_pct"] >= 0.05).mean())
        prior_vol_median = float(group["volume_median_prior_24h"].dropna().iloc[-1]) if group["volume_median_prior_24h"].notna().any() else 0.0
        prior_range_median = float(group["range_median_prior_24h"].dropna().iloc[-1]) if group["range_median_prior_24h"].notna().any() else 0.0
        volume_sum = float(group["volume"].sum())
        volume_ratio_prior = volume_sum / max(prior_vol_median * max(len(group), 1), 1e-12) if prior_vol_median > 0 else 1.0
        range_ratio_prior = median_candle_pct / max(prior_range_median, 1e-12) if prior_range_median > 0 else 1.0
        trend_efficiency = abs(close_move_pct) / max(path_pct, 1e-12)
        direction = "UP" if close_move_pct > 0.03 else "DOWN" if close_move_pct < -0.03 else "FLAT"
        trend_state = "тренд" if trend_efficiency >= 0.42 and abs(close_move_pct) >= 0.25 else "боковик/пила"
        density_score = path_pct / max(hour_range_pct, 1e-12)
        phase_metric_pct = max(
            hour_range_pct,
            abs(close_move_pct) * 1.25,
            median_candle_pct * 8.0,
        )
        if volume_ratio_prior >= 2.0 and range_ratio_prior >= 1.4:
            phase_metric_pct *= 1.08
        phase = _phase_from_pct(phase_metric_pct)
        session_code, session_label = _session_for_timestamp(hour_ts + pd.Timedelta(minutes=30))
        day_type = _day_type(hour_ts)
        effective_code, effective_label, real_tradfi_open = _effective_session(session_code, session_label, day_type)
        rows.append(
            {
                "day_utc": hour_ts.strftime("%Y-%m-%d"),
                "hour_start_utc": hour_ts.strftime("%Y-%m-%dT%H:00:00Z"),
                "hour_utc": int(hour_ts.hour),
                "weekday": hour_ts.day_name(),
                "is_weekend": bool(hour_ts.dayofweek >= 5),
                "day_type": day_type,
                "session_assignment_rule": "hour_midpoint_utc",
                "session_time_bucket_code": session_code,
                "session_time_bucket_label": session_label,
                "session_code": effective_code,
                "session_label": effective_label,
                "effective_session_code": effective_code,
                "effective_session_label": effective_label,
                "real_tradfi_session_open": real_tradfi_open,
                "open": round(open_price, 8),
                "high": round(high, 8),
                "low": round(low, 8),
                "close": round(close_price, 8),
                "hour_range_pct": round(hour_range_pct, 6),
                "close_move_pct": round(close_move_pct, 6),
                "path_pct": round(path_pct, 6),
                "median_candle_range_pct": round(median_candle_pct, 6),
                "median_body_pct": round(median_body_pct, 6),
                "active_candle_share": round(active_candle_share, 6),
                "volume_sum": round(volume_sum, 6),
                "volume_ratio_prior_24h": round(volume_ratio_prior, 6),
                "range_ratio_prior_24h": round(range_ratio_prior, 6),
                "density_score": round(density_score, 6),
                "trend_efficiency": round(trend_efficiency, 6),
                "direction": direction,
                "trend_state": trend_state,
                "phase_metric_pct": round(phase_metric_pct, 6),
                "market_phase": phase,
                "market_phase_rank": _phase_rank(phase),
                "causal_note": "hour_closed_audit; для входа использовать только закрытые свечи до entry",
            }
        )
    return rows


def _aggregate(rows: list[dict[str, Any]], group_keys: list[str]) -> list[dict[str, Any]]:
    buckets: dict[tuple[Any, ...], list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        buckets[tuple(row[key] for key in group_keys)].append(row)
    out: list[dict[str, Any]] = []
    for key, items in sorted(buckets.items(), key=lambda x: x[0]):
        phase_counts = Counter(str(item["market_phase"]) for item in items)
        avg_rank = sum(float(item["market_phase_rank"]) for item in items) / max(len(items), 1)
        record = {group_keys[idx]: key[idx] for idx in range(len(group_keys))}
        record.update(
            {
                "hours": len(items),
                "avg_phase_rank": round(avg_rank, 4),
                "dominant_phase": phase_counts.most_common(1)[0][0],
                "strong_plus_hours": sum(1 for item in items if int(item["market_phase_rank"]) >= _phase_rank("Сильная")),
                "dead_weak_hours": sum(1 for item in items if int(item["market_phase_rank"]) <= _phase_rank("Слабая")),
                "avg_hour_range_pct": round(sum(float(item["hour_range_pct"]) for item in items) / max(len(items), 1), 6),
                "avg_abs_close_move_pct": round(sum(abs(float(item["close_move_pct"])) for item in items) / max(len(items), 1), 6),
                "avg_volume_ratio_prior_24h": round(sum(float(item["volume_ratio_prior_24h"]) for item in items) / max(len(items), 1), 6),
                "trend_hours": sum(1 for item in items if item["trend_state"] == "тренд"),
                "phase_counts": dict(phase_counts),
            }
        )
        out.append(record)
    return out


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _discover_stas1_runs(root: Path, run_dirs: list[str]) -> list[Stas1Run]:
    runs: list[Stas1Run] = []
    if run_dirs:
        for item in run_dirs:
            run_dir = (root / item).resolve() if not Path(item).is_absolute() else Path(item)
            csv_path = run_dir / "GOOD_1PCT_REVIEW_POOL_RECORDS.csv"
            if csv_path.exists():
                runs.append(Stas1Run(run_dir=run_dir, csv_path=csv_path))
        return runs

    base = root / "STAS1_GOOD_LOW_REVIEW" / "runs"
    if not base.exists():
        return []
    for csv_path in base.glob("*/GOOD_1PCT_REVIEW_POOL_RECORDS.csv"):
        runs.append(Stas1Run(run_dir=csv_path.parent, csv_path=csv_path))
    return runs


def _entry_phase_rows(
    *,
    root: Path,
    df: pd.DataFrame,
    start_day: str,
    end_day: str,
    stas1_runs: list[Stas1Run],
) -> list[dict[str, Any]]:
    start = pd.Timestamp(start_day, tz="UTC")
    end = pd.Timestamp(end_day, tz="UTC") + pd.Timedelta(days=1)
    loaded: list[dict[str, Any]] = []
    seen_ids: set[tuple[str, str]] = set()
    for run in stas1_runs:
        for row in _read_csv_rows(run.csv_path):
            entry_raw = row.get("entry_time_utc") or ""
            if not entry_raw:
                continue
            entry_ts = pd.Timestamp(entry_raw)
            if entry_ts.tzinfo is None:
                entry_ts = entry_ts.tz_localize("UTC")
            else:
                entry_ts = entry_ts.tz_convert("UTC")
            if entry_ts < start or entry_ts >= end:
                continue
            key = (str(row.get("record_id") or ""), entry_ts.isoformat())
            if key in seen_ids:
                continue
            seen_ids.add(key)
            prior = df[(df["open_time_utc"] >= entry_ts - pd.Timedelta(minutes=60)) & (df["open_time_utc"] < entry_ts)]
            if prior.empty:
                continue
            open_price = float(prior.iloc[0]["open"])
            close_price = float(prior.iloc[-1]["close"])
            high = float(prior["high"].max())
            low = float(prior["low"].min())
            range_pct = (high - low) / max(open_price, 1e-12) * 100.0
            close_move_pct = (close_price - open_price) / max(open_price, 1e-12) * 100.0
            path_pct = float(prior["close_step_abs_pct"].sum())
            median_candle_pct = _safe_mean(prior["candle_range_pct"])
            phase_metric_pct = max(range_pct, abs(close_move_pct) * 1.25, median_candle_pct * 8.0)
            phase = _phase_from_pct(phase_metric_pct)
            session_code, session_label = _session_for_timestamp(entry_ts)
            day_type = _day_type(entry_ts)
            effective_code, effective_label, real_tradfi_open = _effective_session(session_code, session_label, day_type)
            loaded.append(
                {
                    "source_run": _rel(root, run.run_dir),
                    "record_id": row.get("record_id"),
                    "candidate_id": row.get("candidate_id"),
                    "day_utc": entry_ts.strftime("%Y-%m-%d"),
                    "entry_time_utc": entry_ts.strftime("%Y-%m-%dT%H:%M:%SZ"),
                    "review_label": row.get("review_label"),
                    "outcome_status": row.get("outcome_status"),
                    "is_good": not str(row.get("review_label") or "").startswith("BAD"),
                    "day_type": day_type,
                    "is_weekend": day_type == "weekend",
                    "session_time_bucket_code": session_code,
                    "session_time_bucket_label": session_label,
                    "session_code": effective_code,
                    "session_label": effective_label,
                    "effective_session_code": effective_code,
                    "effective_session_label": effective_label,
                    "real_tradfi_session_open": real_tradfi_open,
                    "prior_60m_range_pct": round(range_pct, 6),
                    "prior_60m_close_move_pct": round(close_move_pct, 6),
                    "prior_60m_path_pct": round(path_pct, 6),
                    "prior_60m_median_candle_pct": round(median_candle_pct, 6),
                    "phase_metric_pct_before_entry": round(phase_metric_pct, 6),
                    "market_phase_before_entry": phase,
                    "market_phase_rank_before_entry": _phase_rank(phase),
                    "causal_note": "расчет только по свечам с open_time_utc < entry_time_utc",
                }
            )
    return sorted(loaded, key=lambda row: (row["entry_time_utc"], str(row.get("record_id") or "")))


SUMMARY_COLUMNS = [
    "session_time_bucket_code",
    "session_time_bucket_label",
    "day_type",
    "effective_session_code",
    "effective_session_label",
    "hours",
    "avg_phase_rank",
    "dominant_phase",
    "strong_plus_hours",
    "dead_weak_hours",
    "avg_hour_range_pct",
    "avg_abs_close_move_pct",
    "avg_volume_ratio_prior_24h",
    "trend_hours",
    "phase_counts",
]


def _write_csv(path: Path, rows: list[dict[str, Any]], columns_hint: list[str] | None = None) -> None:
    columns: list[str] = list(columns_hint or [])
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            out = dict(row)
            for key, value in list(out.items()):
                if isinstance(value, (dict, list)):
                    out[key] = json.dumps(value, ensure_ascii=False, sort_keys=True)
            writer.writerow(out)


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n", encoding="utf-8")


def _xlsx_value(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, (np.integer, np.floating)):
        return value.item()
    if isinstance(value, (bool, int, float)) or value is None:
        return value
    return str(value)


def _write_sheet(ws: Any, rows: list[dict[str, Any]], columns: list[str], title: str) -> None:
    ws.title = title
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, column in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, row in enumerate(rows, 2):
        for col_idx, column in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=_xlsx_value(row.get(column)))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, column in enumerate(columns, 1):
        values = [str(row.get(column, "")) for row in rows[:200]]
        width = min(max([len(column), *(len(value) for value in values)]) + 2, 48)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _columns_for(rows: list[dict[str, Any]], columns_hint: list[str] | None = None) -> list[str]:
    columns: list[str] = list(columns_hint or [])
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    return columns


def _write_xlsx(
    path: Path,
    *,
    hour_rows: list[dict[str, Any]],
    session_time_summary: list[dict[str, Any]],
    effective_session_summary: list[dict[str, Any]],
    weekday_session_summary: list[dict[str, Any]],
    weekend_session_summary: list[dict[str, Any]],
    day_type_summary: list[dict[str, Any]],
    entry_rows: list[dict[str, Any]],
) -> None:
    wb = Workbook()
    sheet_specs = [
        ("Hourly phases", hour_rows, _columns_for(hour_rows)),
        ("Session buckets", session_time_summary, _columns_for(session_time_summary, SUMMARY_COLUMNS)),
        ("Effective sessions", effective_session_summary, _columns_for(effective_session_summary, SUMMARY_COLUMNS)),
        ("Weekday sessions", weekday_session_summary, _columns_for(weekday_session_summary, SUMMARY_COLUMNS)),
        ("Weekend buckets", weekend_session_summary, _columns_for(weekend_session_summary, SUMMARY_COLUMNS)),
        ("Weekday weekend", day_type_summary, _columns_for(day_type_summary, SUMMARY_COLUMNS)),
        ("Entry context", entry_rows, _columns_for(entry_rows)),
    ]
    first = True
    for title, rows, columns in sheet_specs:
        ws = wb.active if first else wb.create_sheet()
        first = False
        _write_sheet(ws, rows, columns, title)
    wb.save(path)


def _render_heatmap(path: Path, hour_rows: list[dict[str, Any]], session_rows: list[dict[str, Any]]) -> None:
    days = sorted({row["day_utc"] for row in hour_rows})
    grid = np.full((len(days), 24), np.nan)
    labels = {day: idx for idx, day in enumerate(days)}
    for row in hour_rows:
        grid[labels[row["day_utc"]], int(row["hour_utc"])] = float(row["market_phase_rank"])

    fig = plt.figure(figsize=(18, max(7, len(days) * 0.75 + 4)))
    gs = fig.add_gridspec(2, 1, height_ratios=[3, 1.6], hspace=0.32)
    ax = fig.add_subplot(gs[0])
    cmap = plt.get_cmap("viridis", len(PHASES))
    im = ax.imshow(grid, aspect="auto", cmap=cmap, vmin=0, vmax=len(PHASES) - 1)
    ax.set_xticks(range(24))
    ax.set_xticklabels([f"{hour:02d}" for hour in range(24)])
    ax.set_yticks(range(len(days)))
    ax.set_yticklabels(days)
    ax.set_xlabel("UTC hour")
    ax.set_ylabel("day")
    ax.set_title("STAS2 фазы рынка по часам: 0=мертвая, 6=очень сильная")
    cbar = fig.colorbar(im, ax=ax, ticks=range(len(PHASES)))
    cbar.ax.set_yticklabels([name for name, _ in PHASES])
    for row in hour_rows:
        ax.text(int(row["hour_utc"]), labels[row["day_utc"]], int(row["market_phase_rank"]), ha="center", va="center", color="white", fontsize=7)

    ax2 = fig.add_subplot(gs[1])
    labels2 = [row.get("session_time_bucket_label") or row.get("session_label") for row in session_rows]
    avg_ranks = [float(row["avg_phase_rank"]) for row in session_rows]
    colors = [cmap(min(max(int(round(value)), 0), len(PHASES) - 1)) for value in avg_ranks]
    ax2.barh(labels2, avg_ranks, color=colors)
    ax2.set_xlim(0, len(PHASES) - 1)
    ax2.set_xlabel("средний phase rank")
    ax2.set_title("Сводка по сессиям")
    for idx, value in enumerate(avg_ranks):
        ax2.text(value + 0.04, idx, f"{value:.2f}", va="center", fontsize=9)
    fig.subplots_adjust(left=0.20, right=0.88, top=0.92, bottom=0.08, hspace=0.35)
    fig.savefig(path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def _phase_table_markdown(rows: list[dict[str, Any]], limit: int = 999) -> list[str]:
    lines = [
        "| День | Тип | Час UTC | UTC-корзина | Effective-сессия | Фаза | Range % | Ход close % | Тренд/боковик | Объем к prior |",
        "|---|---|---:|---|---|---|---:|---:|---|---:|",
    ]
    for row in rows[:limit]:
        lines.append(
            f"| {row['day_utc']} | {row['day_type']} | {row['hour_utc']:02d} | {row['session_time_bucket_label']} | {row['effective_session_label']} | {row['market_phase']} | {float(row['hour_range_pct']):.3f} | {float(row['close_move_pct']):+.3f} | {row['trend_state']} | {float(row['volume_ratio_prior_24h']):.2f} |"
        )
    return lines


def _summary_table_markdown(rows: list[dict[str, Any]], first_key: str, label: str) -> list[str]:
    lines = [
        f"| {label} | Часов | Доминирующая фаза | Avg rank | Сильных+ | Мертвых/слабых | Avg range % | Тренд-часы |",
        "|---|---:|---|---:|---:|---:|---:|---:|",
    ]
    for row in rows:
        lines.append(
            f"| {row[first_key]} | {row['hours']} | {row['dominant_phase']} | {float(row['avg_phase_rank']):.2f} | {row['strong_plus_hours']} | {row['dead_weak_hours']} | {float(row['avg_hour_range_pct']):.3f} | {row['trend_hours']} |"
        )
    return lines


def _effective_summary_table_markdown(rows: list[dict[str, Any]]) -> list[str]:
    lines = [
        "| Тип дня | Effective-сессия | Часов | Доминирующая фаза | Avg rank | Сильных+ | Мертвых/слабых | Avg range % | Тренд-часы |",
        "|---|---|---:|---|---:|---:|---:|---:|---:|",
    ]
    for row in rows:
        lines.append(
            f"| {row['day_type']} | {row['effective_session_label']} | {row['hours']} | {row['dominant_phase']} | {float(row['avg_phase_rank']):.2f} | {row['strong_plus_hours']} | {row['dead_weak_hours']} | {float(row['avg_hour_range_pct']):.3f} | {row['trend_hours']} |"
        )
    return lines


def _write_report(path: Path, payload: dict[str, Any]) -> None:
    hour_rows = payload["hour_rows"]
    session_rows = payload["session_time_summary"]
    effective_session_rows = payload["effective_session_summary"]
    weekday_session_rows = payload["weekday_session_summary"]
    weekend_session_rows = payload["weekend_session_summary"]
    weekend_rows = payload["day_type_summary"]
    entry_rows = payload["entry_phase_rows"]
    entry_phase_counts = Counter(row["market_phase_before_entry"] for row in entry_rows)
    good_by_phase = Counter(row["market_phase_before_entry"] for row in entry_rows if row["is_good"])
    bad_by_phase = Counter(row["market_phase_before_entry"] for row in entry_rows if not row["is_good"])

    best_sessions = [
        row for row in weekday_session_rows
        if int(row["strong_plus_hours"]) > 0 or float(row["avg_phase_rank"]) >= _phase_rank("Средняя")
    ]
    weak_sessions = [
        row for row in effective_session_rows
        if int(row["dead_weak_hours"]) / max(int(row["hours"]), 1) >= 0.45
    ]

    lines = [
        "# STAS 2: фазы рынка и сессии",
        "",
        f"Статус: `{payload['status']}`.",
        "",
        "Назначение: разобрать Stas 1 базу по фазам рынка, часам и торговым сессиям без запуска ML/Optuna и без использования будущих данных как признаков входа.",
        "",
        "## Stas 1 inventory",
        "",
        "- Текущий Stas 1 находится в `STAS1_GOOD_LOW_REVIEW/`.",
        "- Основной движок: `src/mlbotnav/visual_entry_good_1pct_review_pool.py`.",
        "- Источник low-кандидатов: `src/mlbotnav/visual_entry_low_anchor_suggester.py`.",
        "- Контракт входа: фактическая low-свеча = `signal`, вход = следующая свеча `next open`, расчетная цена = `entry open +5bps`.",
        "- Графики Stas 1: дневной overview, `GOOD_1PCT_REVIEW_POOL_GOOD_CLOSEUPS_PAGE_*.png`, `GOOD_1PCT_REVIEW_POOL_ALL_CLOSEUPS_PAGE_*.png`, `BROWSE_BY_DAY/`.",
        "- Хорошие входы в Stas 1 сейчас означают только offline outcome hit `+1%/+0.5%` от execution-цены, а не ручной ML-good.",
        "- Шум виден в `BAD` красных крестах, микролоях, дублях basin, late-pump-dependent входах и местах, где пользователь просил переякорить low.",
        "",
        "Актуальные Stas 1 runs для продолжения:",
    ]
    for run in payload["stas1_runs_used"]:
        lines.append(f"- `{run}`")

    lines.extend(
        [
            "",
            "## Правило фаз",
            "",
            "Фаза считается по закрытым часовым блокам для аудита дня. Для контекста конкретного входа дополнительно считается `phase_before_entry` только по свечам `open_time_utc < entry_time_utc` за предыдущие 60 минут.",
            "",
            "## Сессионная модель",
            "",
            "Используется `6` внутридневных UTC-корзин времени и отдельный `day_type`:",
            "",
            "- `session_time_bucket` всегда размечает время суток, даже в выходные.",
            "- `day_type = weekday/weekend` хранится отдельно.",
            "- `effective_session` = комбинация `day_type + session_time_bucket`, чтобы не смешивать будний Лондон с выходным временем Лондона.",
            "- `real_tradfi_session_open=false` для выходных и для переходных/после-NY окон.",
            "",
            "| UTC-корзина | Старт UTC | Конец UTC |",
            "|---|---:|---:|",
        ]
    )
    for item in payload["session_time_buckets_utc"]:
        lines.append(
            f"| {item['session_time_bucket_label']} | {item['start_time_utc']} | {item['end_time_utc']} |"
        )
    lines.extend(
        [
            "",
            "| Rank | Фаза | Порог phase_metric_pct |",
            "|---:|---|---:|",
        ]
    )
    prev = 0.0
    for idx, (name, upper) in enumerate(PHASES):
        bound = f">= {prev:.2f}" if math.isinf(upper) else f"{prev:.2f}..{upper:.2f}"
        lines.append(f"| {idx} | {name} | {bound} |")
        prev = upper

    lines.extend(
        [
            "",
            "## Таблица фаз по часам",
            "",
            *_phase_table_markdown(hour_rows),
            "",
            "## Таблица фаз по UTC-корзинам",
            "",
            *_summary_table_markdown(session_rows, "session_time_bucket_label", "UTC-корзина"),
            "",
            "## Таблица effective-сессий",
            "",
            *_effective_summary_table_markdown(effective_session_rows),
            "",
            "## Будние сессии отдельно",
            "",
            *_summary_table_markdown(weekday_session_rows, "session_time_bucket_label", "Будняя UTC-корзина"),
            "",
            "## Выходные по тем же UTC-корзинам",
            "",
            *_summary_table_markdown(weekend_session_rows, "session_time_bucket_label", "Выходная UTC-корзина"),
            "",
            "## Рабочий день против выходного",
            "",
            *_summary_table_markdown(weekend_rows, "day_type", "Тип дня"),
            "",
            "## Фаза перед Stas 1 входами",
            "",
            f"- Всего Stas1 entry-context строк: `{len(entry_rows)}`.",
            f"- Распределение фаз перед входом: `{dict(entry_phase_counts)}`.",
            f"- GOOD по фазам: `{dict(good_by_phase)}`.",
            f"- BAD по фазам: `{dict(bad_by_phase)}`.",
            "",
            "## Вывод по сессиям",
            "",
        ]
    )
    if best_sessions:
        lines.append("Будние окна, которые подходят для поиска входов в текущем срезе:")
        for row in best_sessions:
            lines.append(
                f"- `{row['session_time_bucket_label']}`: dominant `{row['dominant_phase']}`, avg rank `{float(row['avg_phase_rank']):.2f}`, strong+ hours `{row['strong_plus_hours']}`."
            )
    else:
        lines.append("В текущем срезе нет сессии, которая стабильно сильная; входы нужно разрешать только при локальной фазе не ниже `Средняя`.")
    lines.append("")
    if weak_sessions:
        lines.append("Шумовые/опасные effective-окна для автодобора:")
        for row in weak_sessions:
            lines.append(
                f"- `{row['effective_session_label']}`: dead/weak hours `{row['dead_weak_hours']}/{row['hours']}`, dominant `{row['dominant_phase']}`."
            )
    else:
        lines.append("Явно мертвой сессии на всем срезе нет, но слабые часы нужно фильтровать по hour phase, а не по названию сессии.")

    lines.extend(
        [
            "",
            "## Признаки фазы",
            "",
            "- `hour_range_pct`: high-low часа в процентах от open часа.",
            "- `close_move_pct`: чистый ход close-open часа.",
            "- `path_pct`: сумма минутных абсолютных шагов close внутри часа, показывает плотность движения.",
            "- `median_candle_range_pct` и `median_body_pct`: размер свечей.",
            "- `volume_ratio_prior_24h`: объем часа к rolling prior 24h median-volume.",
            "- `range_ratio_prior_24h`: размер свечей к prior 24h median-range.",
            "- `trend_efficiency`: доля направленного движения в общем пути; отделяет тренд от пилы.",
            "- `active_candle_share`: доля свечей с заметным range.",
            "",
            "## Артефакты",
            "",
        ]
    )
    for key, value in payload["artifacts"].items():
        lines.append(f"- `{key}`: `{value}`")
    lines.extend(
        [
            "",
            "## Границы",
            "",
            "- ML/export/training, Optuna, scorer, target-lock и API не запускались.",
            "- `hit_1pct`, `target_1pct` и future outcome не используются для определения фазы.",
            "- Следующий этап Stas 3 можно делать только после ручного принятия этой схемы фаз/сессий.",
        ]
    )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_audit(
    *,
    start_day: str,
    end_day: str,
    symbol: str,
    timeframe: str,
    out_dir: Path,
    run_label: str,
    stas1_run_dirs: list[str],
) -> dict[str, Any]:
    root = Path(__file__).resolve().parents[2]
    if not out_dir.is_absolute():
        out_dir = root / out_dir
    run_id = f"{run_label}_{_run_stamp()}" if run_label else f"stas2_{start_day.replace('-', '')}_{end_day.replace('-', '')}_{_run_stamp()}"
    run_dir = out_dir / run_id
    run_dir.mkdir(parents=True, exist_ok=True)

    days = _iter_days(start_day, end_day)
    df, missing_sources = _load_days(root, days, symbol, timeframe)
    hour_rows = _hour_rows(df)
    session_time_summary = _aggregate(hour_rows, ["session_time_bucket_code", "session_time_bucket_label"])
    effective_session_summary = _aggregate(hour_rows, ["day_type", "effective_session_code", "effective_session_label"])
    weekday_session_summary = _aggregate(
        [row for row in hour_rows if row["day_type"] == "weekday"],
        ["session_time_bucket_code", "session_time_bucket_label"],
    )
    weekend_session_summary = _aggregate(
        [row for row in hour_rows if row["day_type"] == "weekend"],
        ["session_time_bucket_code", "session_time_bucket_label"],
    )
    day_type_summary = _aggregate(hour_rows, ["day_type"])
    stas1_runs = _discover_stas1_runs(root, stas1_run_dirs)
    entry_rows = _entry_phase_rows(root=root, df=df, start_day=start_day, end_day=end_day, stas1_runs=stas1_runs)

    hour_csv = run_dir / "STAS2_HOURLY_PHASES.csv"
    session_csv = run_dir / "STAS2_SESSION_PHASE_SUMMARY.csv"
    effective_session_csv = run_dir / "STAS2_EFFECTIVE_SESSION_SUMMARY.csv"
    weekday_session_csv = run_dir / "STAS2_WEEKDAY_SESSION_SUMMARY.csv"
    weekend_session_csv = run_dir / "STAS2_WEEKEND_TIME_BUCKET_SUMMARY.csv"
    weekend_csv = run_dir / "STAS2_WEEKDAY_WEEKEND_SUMMARY.csv"
    entry_csv = run_dir / "STAS2_STAS1_ENTRY_PHASE_CONTEXT.csv"
    json_path = run_dir / "STAS2_MARKET_PHASE_AUDIT_PAYLOAD.json"
    report_path = run_dir / "STAS2_MARKET_PHASE_AUDIT_RU.md"
    heatmap_png = run_dir / "STAS2_MARKET_PHASE_HEATMAP.png"
    xlsx_path = run_dir / "STAS2_MARKET_PHASE_TABLES.xlsx"

    _write_csv(hour_csv, hour_rows)
    _write_csv(session_csv, session_time_summary, SUMMARY_COLUMNS)
    _write_csv(effective_session_csv, effective_session_summary, SUMMARY_COLUMNS)
    _write_csv(weekday_session_csv, weekday_session_summary, SUMMARY_COLUMNS)
    _write_csv(weekend_session_csv, weekend_session_summary, SUMMARY_COLUMNS)
    _write_csv(weekend_csv, day_type_summary, SUMMARY_COLUMNS)
    _write_csv(entry_csv, entry_rows)
    _write_xlsx(
        xlsx_path,
        hour_rows=hour_rows,
        session_time_summary=session_time_summary,
        effective_session_summary=effective_session_summary,
        weekday_session_summary=weekday_session_summary,
        weekend_session_summary=weekend_session_summary,
        day_type_summary=day_type_summary,
        entry_rows=entry_rows,
    )
    _render_heatmap(heatmap_png, hour_rows, session_time_summary)

    payload = {
        "schema_version": 1,
        "status": STATUS,
        "created_utc": _utc_now(),
        "run_id": run_id,
        "symbol": symbol,
        "timeframe": timeframe,
        "date_range": {"start_day": start_day, "end_day": end_day},
        "phase_ladder": [
            {"rank": idx, "phase": name, "upper_phase_metric_pct": None if math.isinf(upper) else upper}
            for idx, (name, upper) in enumerate(PHASES)
        ],
        "session_time_buckets_utc": [
            {
                "session_time_bucket_code": code,
                "session_time_bucket_label": label,
                "start_minute_utc": start,
                "end_minute_utc": end,
                "start_time_utc": f"{start // 60:02d}:{start % 60:02d}",
                "end_time_utc": "24:00" if end == 1440 else f"{end // 60:02d}:{end % 60:02d}",
            }
            for code, label, start, end in SESSION_TIME_BUCKETS
        ],
        "hour_rows": hour_rows,
        "session_summary": session_time_summary,
        "session_time_summary": session_time_summary,
        "effective_session_summary": effective_session_summary,
        "weekday_session_summary": weekday_session_summary,
        "weekend_session_summary": weekend_session_summary,
        "weekend_summary": day_type_summary,
        "day_type_summary": day_type_summary,
        "entry_phase_rows": entry_rows,
        "missing_sources": missing_sources,
        "stas1_runs_used": [_rel(root, run.run_dir) for run in stas1_runs],
        "artifacts": {
            "run_dir": _rel(root, run_dir),
            "report_ru": _rel(root, report_path),
            "json": _rel(root, json_path),
            "hourly_csv": _rel(root, hour_csv),
            "session_csv": _rel(root, session_csv),
            "effective_session_csv": _rel(root, effective_session_csv),
            "weekday_session_csv": _rel(root, weekday_session_csv),
            "weekend_session_csv": _rel(root, weekend_session_csv),
            "weekday_weekend_csv": _rel(root, weekend_csv),
            "entry_phase_context_csv": _rel(root, entry_csv),
            "xlsx": _rel(root, xlsx_path),
            "heatmap_png": _rel(root, heatmap_png),
        },
        "guardrails": [
            "NO_ML_EXPORT_TRAINING",
            "NO_OPTUNA",
            "NO_SCORER",
            "NO_TARGET_LOCK",
            "ENTRY_PHASE_USES_ONLY_PRIOR_CANDLES",
        ],
    }
    _write_json(json_path, payload)
    _write_report(report_path, payload)
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description="STAS2 market phase and session audit, no ML/Optuna.")
    parser.add_argument("--start-day", default="2026-05-02")
    parser.add_argument("--end-day", default="2026-05-08")
    parser.add_argument("--symbol", default="SOLUSDT")
    parser.add_argument("--timeframe", default="1m")
    parser.add_argument("--out-dir", default="reports/final_review/visual_entry_v3/fresh_target_led/stas2_market_phase_percent_ladder")
    parser.add_argument("--run-label", default="stas2_20260502_20260508_v0")
    parser.add_argument("--stas1-run-dir", action="append", default=[])
    args = parser.parse_args()

    payload = run_audit(
        start_day=args.start_day,
        end_day=args.end_day,
        symbol=args.symbol,
        timeframe=args.timeframe,
        out_dir=Path(args.out_dir),
        run_label=args.run_label,
        stas1_run_dirs=args.stas1_run_dir,
    )
    print(
        json.dumps(
            {
                "status": payload["status"],
                "run_id": payload["run_id"],
                "artifacts": payload["artifacts"],
                "missing_sources": payload["missing_sources"],
                "stas1_runs_used": payload["stas1_runs_used"],
            },
            ensure_ascii=False,
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
